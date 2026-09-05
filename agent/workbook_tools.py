"""Small structured workbook tools. Python remains available for complex transformations."""

from __future__ import annotations

import openpyxl

from verify import _range_cells


def inspect_range(path: str, sheet: str, cell_range: str, *, styles: bool = False, limit: int = 100) -> tuple[bool, str]:
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        if sheet not in wb.sheetnames:
            return False, f"Sheet not found: {sheet!r}"
        rows = []
        for cell in _range_cells(wb[sheet], cell_range):
            value = repr(cell.value)
            text = f"{cell.coordinate}={value} [{type(cell.value).__name__}]"
            if styles:
                text += f" style={cell.style_id} format={cell.number_format!r}"
            rows.append(text)
            if len(rows) >= limit:
                rows.append("... cells omitted ...")
                break
        return True, "\n".join(rows) or "range is empty"
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}"


def assert_blank(path: str, sheet: str, cell_range: str) -> tuple[bool, str]:
    ok, details = inspect_range(path, sheet, cell_range)
    if not ok:
        return False, details
    wb = openpyxl.load_workbook(path, data_only=False)
    nonempty = [cell.coordinate for cell in _range_cells(wb[sheet], cell_range) if cell.value not in (None, "")]
    return (not nonempty), ("Range is blank" if not nonempty else f"Expected blank; found: {', '.join(nonempty[:20])}")


def assert_sorted(path: str, sheet: str, cell_range: str, keys: list[str]) -> tuple[bool, str]:
    """Check ascending order using column letters supplied by the model."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet]
        rows = list(_range_cells(ws, cell_range))
        # Group flattened cells into rows using the requested range width.
        from openpyxl.utils.cell import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        key_cols = [openpyxl.utils.column_index_from_string(key) for key in keys]
        values = []
        for row in range(min_row, (max_row or ws.max_row) + 1):
            values.append(tuple(ws.cell(row, col).value for col in key_cols))
        cleaned = [row for row in values if any(value not in (None, "") for value in row)]
        ordered = sorted(cleaned, key=lambda row: tuple("" if value is None else value for value in row))
        return cleaned == ordered, ("Range is sorted ascending" if cleaned == ordered else "Range is not sorted ascending")
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}"
