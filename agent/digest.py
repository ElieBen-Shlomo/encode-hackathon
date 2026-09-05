"""Workbook digest for prompts. Unlike the baseline's blind 120x30 cap, this is
answer-aware and budgeted: small sheets are dumped whole, big sheets show the head
plus a window around the answer range, and always say what was omitted. The model's
code sees the full file regardless — the digest only needs to convey structure.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from sb import answer_ranges, parse_answer_position

HEAD_ROWS = 30
WINDOW = 8         # rows around the answer range
ANSWER_SLICE = 15  # for big answer ranges: rows shown from each end
MAX_SHOWN = 120    # hard cap on rows shown per sheet
MAX_COLS = 40
CELL_CHARS = 60    # per-cell truncation


def _fmt(v) -> str:
    s = "" if v is None else str(v)
    return s if len(s) <= CELL_CHARS else s[:CELL_CHARS] + "…"


def _rows_tsv(ws, rows: list[int], cols: int) -> list[str]:
    return ["\t".join([str(r)] + [_fmt(ws.cell(row=r, column=c).value) for c in range(1, cols + 1)])
            for r in rows]


def _answer_rows(task, sheet_title: str, max_row: int) -> set[int]:
    rows = set()
    for sheet, rng in answer_ranges(task):
        if sheet and sheet != sheet_title:
            continue
        try:
            _, r0, _, r1 = range_boundaries(rng)
        except ValueError:
            continue
        r0, r1 = r0 or 1, min(r1 or max_row, max_row)
        lo, hi = max(1, r0 - WINDOW), min(max_row, r1 + WINDOW)
        if hi - lo + 1 <= 2 * ANSWER_SLICE:
            rows.update(range(lo, hi + 1))
        else:  # big range: first and last slice; the script reads the full file anyway
            rows.update(range(lo, lo + ANSWER_SLICE))
            rows.update(range(hi - ANSWER_SLICE + 1, hi + 1))
    return rows


def digest(path: str, task: dict) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    parts = []
    for ws in wb.worksheets:
        n_rows, n_cols = ws.max_row, ws.max_column
        cols = min(n_cols, MAX_COLS)
        header = "\t".join([""] + [get_column_letter(c) for c in range(1, cols + 1)])

        show = set(range(1, min(n_rows, HEAD_ROWS) + 1))
        show |= _answer_rows(task, ws.title, n_rows)
        shown = sorted(show)[:MAX_SHOWN]

        title = (f"### Sheet: {ws.title!r} — {n_rows} rows x {n_cols} cols"
                 + (f" (showing {len(shown)} rows; columns beyond {get_column_letter(cols)} omitted)"
                    if len(shown) < n_rows or cols < n_cols else ""))
        body, prev = [], 0
        for r in shown:
            if r > prev + 1:
                body.append(f"... rows {prev + 1}-{r - 1} omitted ({r - 1 - prev} rows) ...")
            body.extend(_rows_tsv(ws, [r], cols))
            prev = r
        if prev < n_rows:
            body.append(f"... rows {prev + 1}-{n_rows} omitted ({n_rows - prev} rows) ...")
        parts.append("\n".join([title, header] + body))
    return "\n\n".join(parts)


def _cell_description(cell) -> str:
    value = cell.value
    if value is None:
        rendered = "<empty>"
    elif isinstance(value, str) and value.startswith("="):
        rendered = f"formula {value!r}"
    else:
        rendered = repr(value)
    return f"{cell.coordinate}={rendered} [{type(value).__name__}]"


def _sample_range(ws, cell_range: str, limit: int) -> list[str]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    min_row = min_row or 1
    max_row = max_row or ws.max_row
    coordinates = [(row, col) for row in range(min_row, max_row + 1) for col in range(min_col, max_col + 1)]
    omitted = len(coordinates) > limit
    if omitted:
        head = max(1, limit // 2)
        coordinates = coordinates[:head] + coordinates[-(limit - head):]
    lines = [_cell_description(ws.cell(row=row, column=col)) for row, col in coordinates]
    if omitted:
        lines.insert(max(1, limit // 2), "... cells omitted ...")
    return lines


def verification_snapshot(path: str, task: dict, *, answer_limit: int = 100, source_limit: int = 80) -> str:
    """Small, type-aware answer/source view for the review and critic phases."""
    wb = openpyxl.load_workbook(path, data_only=False)
    parts = ["## Graded answer cells (current workbook)"]
    for sheet, cell_range in answer_ranges(task):
        if sheet not in wb.sheetnames:
            parts.append(f"{sheet!r}!{cell_range}: SHEET MISSING")
            continue
        parts.append(f"### {sheet!r}!{cell_range}")
        parts.extend(_sample_range(wb[sheet], cell_range, answer_limit))

    if data_range := task.get("data_position"):
        parts.append("\n## Relevant declared source cells")
        for sheet, cell_range in parse_answer_position(data_range):
            sheet = sheet or task.get("answer_sheet")
            if sheet not in wb.sheetnames:
                parts.append(f"{sheet!r}!{cell_range}: SHEET MISSING")
                continue
            parts.append(f"### {sheet!r}!{cell_range}")
            parts.extend(_sample_range(wb[sheet], cell_range, source_limit))
    return "\n".join(parts)
