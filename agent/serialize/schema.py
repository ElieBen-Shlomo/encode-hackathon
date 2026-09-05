"""`schema` view, for the code-writing solver: describe the columns, show a few rows, stop.

The model's code reads the real file, so the prompt only needs the layout: per sheet the
dimensions, the header row, a column table (letter, header, type, format, non-empty count,
distinct values or samples, R1C1 formula pattern), merged ranges, defined names, then the header
plus the first 8 and last 2 rows, and the raw answer region. Size-invariant: a 50k-row sheet costs
the same as a 50-row sheet. Budget 6k tokens by default; the raw rows go first when over budget.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from serialize import Rendered, register
from serialize.grid import (
    ANSWER_PAD, MAX_COLS, MAX_ROI_SHEETS, GridModel, SheetPlan, _answer_regions, _mentioned_sheets, _meta,
    _row_cells, build_model,
)
from serialize.tokens import count_tokens
from workbook import formula_text

DEFAULT_BUDGET = 6_000
FIRST_ROWS = 8
LAST_ROWS = 2
ANSWER_ROWS = 12


def _column_table(info, title: str, ncols: int) -> list[str]:
    s = info.sheet(title)
    ws_f = info.wb_f[title]
    lines = ["col | header | type | format | non-empty | values | formula (R1C1)"]
    for c in range(1, min(ncols, len(s.columns)) + 1):
        col = s.columns[c - 1]
        if col.n_nonempty == 0 and not col.header:
            continue
        if col.distinct is not None and len(col.distinct) <= 8:
            values = "distinct: " + ", ".join(str(d)[:20] for d in col.distinct)
        else:
            values = "e.g. " + ", ".join(str(x)[:20] for x in col.samples)
        formula = ""
        if col.formula_pattern:
            first = ""
            for r in range(s.header_row + 1, s.scanned_rows + 1):
                ft = formula_text(ws_f.cell(row=r, column=c).value)
                if ft:
                    first = f" e.g. {col.letter}{r}: {ft[:50]}"
                    break
            formula = f"{col.formula_pattern[:60]} ({col.formula_pattern_share:.0%} of {col.n_formula}){first}"
        fmt = "" if col.number_format in ("General", "", None) else col.number_format
        lines.append(f"{col.letter} | {col.header or ''} | {col.dtype} | {fmt} | {col.n_nonempty} | {values} | {formula}")
    return lines


def plan_schema(task: dict, info) -> tuple[list[SheetPlan], dict]:
    regions = _answer_regions(task, info)
    roi_titles = set(regions) | _mentioned_sheets(task, info)
    if len(info.sheets) <= MAX_ROI_SHEETS:
        roi_titles |= {s.title for s in info.sheets}
    plans = []
    for s in info.sheets:
        ncols = min(s.max_col, MAX_COLS if s.max_row > 20 else 2 * MAX_COLS)
        header = set(range(1, s.header_row + 1))
        if s.title not in roi_titles:
            plans.append(SheetPlan(s.title, rows=sorted(header), protected=set(header), ncols=ncols, roi=False, note="header only"))
            continue
        rows = set(header)
        rows |= set(range(s.header_row + 1, min(s.max_row, s.header_row + FIRST_ROWS) + 1))
        rows |= set(range(max(1, s.max_row - LAST_ROWS + 1), s.max_row + 1))
        protected = set(header)
        for c0, r0, c1, r1 in regions.get(s.title, []):
            lo, hi = max(1, r0 - 1), min(s.max_row, r1 + 1)
            span = set(range(lo, hi + 1)) if hi - lo + 1 <= ANSWER_ROWS else set(range(lo, lo + ANSWER_ROWS // 2)) | set(range(hi - ANSWER_ROWS // 2 + 1, hi + 1))
            protected |= span
        rows |= protected
        extra = [f"Header row {s.header_row}; merged: {', '.join(s.merged[:8]) or 'none'}", *_column_table(info, s.title, ncols),
                 "Sample rows (header, first rows, answer region, last rows):"]
        plans.append(SheetPlan(s.title, rows=sorted(rows), protected=protected, ncols=ncols, roi=True, extra_lines=extra))
    return plans, regions


def emit_schema(model: GridModel) -> str:
    parts = ["\n".join(model.preamble)]
    for plan in model.plans:
        s = model.info.sheet(plan.title)
        lines = [f"### Sheet {plan.title!r}: {s.max_row} rows x {s.max_col} cols" + (f" ({plan.note})" if plan.note else "")]
        lines += plan.extra_lines
        if plan.roi:
            lines.append("\t".join([""] + [get_column_letter(c) for c in range(1, plan.ncols + 1)]))
            prev = 0
            for r in plan.rows:
                if r > prev + 1:
                    lines.append(f"... rows {prev + 1}-{r - 1} not shown ...")
                lines.append("\t".join([str(r)] + _row_cells(model, plan, r)))
                prev = r
            if prev < s.max_row:
                lines.append(f"... rows {prev + 1}-{s.max_row} not shown ...")
        parts.append("\n".join(lines))
    return "\n\n".join(parts)


@register("schema")
def render_schema(path: str, task: dict, budget_tokens: int | None) -> Rendered:
    model: GridModel = build_model(path, task, budget_tokens or DEFAULT_BUDGET, planner=plan_schema)
    text = emit_schema(model)
    meta = _meta(model)
    meta["layout"] = "schema"
    return Rendered(text, meta)
