"""Workbook renderers: one signature, many views. A run differs only by --digest.

    from serialize import render, available
    r = render("grid", task["init_xlsx"], task, budget_tokens=10_000)
    r.text   # goes into the prompt
    r.meta   # digest, tokens, sheets_rendered, rows_omitted, answer_range_in_window, ...

Renderers register themselves with @register("name"). Only init workbooks are ever opened here.
"""

from __future__ import annotations

import importlib
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

HERE = Path(__file__).resolve().parent
for _candidate in (HERE.parent.parent, HERE.parent.parent / "research"):
    if (_candidate / "sb.py").exists():
        if str(_candidate) not in sys.path:
            sys.path.insert(0, str(_candidate))
        break
if str(HERE.parent) not in sys.path:  # agent/, for digest.py
    sys.path.insert(0, str(HERE.parent))

import openpyxl
from openpyxl.utils.cell import range_boundaries

from sb import answer_ranges, serialize_workbook
from serialize.tokens import count_tokens


@dataclass
class Rendered:
    text: str
    meta: dict = field(default_factory=dict)


RendererFn = Callable[[str, dict, "int | None"], Rendered]
REGISTRY: dict[str, RendererFn] = {}


def register(name: str):
    def deco(fn: RendererFn) -> RendererFn:
        REGISTRY[name] = fn
        return fn
    return deco


def available() -> list[str]:
    return sorted(REGISTRY)


def render(name: str, path: str, task: dict, budget_tokens: int | None = None) -> Rendered:
    if name not in REGISTRY:
        raise KeyError(f"unknown digest {name!r}; available: {available()}")
    r = REGISTRY[name](path, task, budget_tokens)
    r.meta.setdefault("digest", name)
    r.meta.setdefault("budget_tokens", budget_tokens)
    r.meta.setdefault("chars", len(r.text))
    r.meta.setdefault("tokens", count_tokens(r.text))
    return r


def answer_visibility(task: dict, wb, shown: dict[str, tuple[set[int], int]],
                      dims: dict[str, tuple[int, int]] | None = None) -> dict:
    """How much of the graded range the view actually shows.

    shown: sheet title -> (set of shown row numbers, number of shown columns).
    dims:  sheet title -> (max_row, max_col) as loaded. Pass this when the workbook has been touched
           since loading: openpyxl's ws.cell() creates cells on access and silently widens max_row /
           max_column, which would make an empty answer region look hidden.
    """
    total = shown_n = 0
    for sheet, rng in answer_ranges(task):
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        try:
            c0, r0, c1, r1 = range_boundaries(rng)
        except ValueError:
            continue
        max_row, max_col = (dims or {}).get(ws.title, (ws.max_row, ws.max_column))
        r0 = r0 or 1
        r1 = max(r0, r1 or max_row)
        rows, ncols = shown.get(ws.title, (set(), 0))
        # cells beyond the sheet's used area are empty, so they count as visible
        col_ok = min(c1 or 1, max_col) <= ncols
        for r in range(r0, r1 + 1):
            total += 1
            if r > max_row or (col_ok and r in rows):
                shown_n += 1
    return {
        "answer_rows_total": total,
        "answer_rows_shown": shown_n,
        "answer_range_in_window": total > 0 and shown_n == total,
    }


@register("tsv")
def render_tsv(path: str, task: dict, budget_tokens: int | None) -> Rendered:
    """The shipped baseline view: values only, every sheet, blind 120x30 cap."""
    text = serialize_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    shown = {ws.title: (set(range(1, min(ws.max_row, 120) + 1)), min(ws.max_column, 30)) for ws in wb.worksheets}
    meta = answer_visibility(task, wb, shown)
    meta.update(
        sheets_rendered=len(wb.worksheets),
        rows_omitted=sum(max(0, ws.max_row - 120) for ws in wb.worksheets),
        cols_omitted=sum(max(0, ws.max_column - 30) for ws in wb.worksheets),
        budget_applied=False,
    )
    return Rendered(text, meta)


@register("windowed")
def render_windowed(path: str, task: dict, budget_tokens: int | None) -> Rendered:
    """Lorcan's answer-aware digest: head rows plus a window around the answer range, values only."""
    import digest as dg

    text = dg.digest(path, task)
    wb = openpyxl.load_workbook(path, data_only=True)
    shown = {}
    rows_omitted = 0
    for ws in wb.worksheets:
        show = set(range(1, min(ws.max_row, dg.HEAD_ROWS) + 1)) | dg._answer_rows(task, ws.title, ws.max_row)
        kept = set(sorted(show)[: dg.MAX_SHOWN])
        shown[ws.title] = (kept, min(ws.max_column, dg.MAX_COLS))
        rows_omitted += ws.max_row - len(kept)
    meta = answer_visibility(task, wb, shown)
    meta.update(
        sheets_rendered=len(wb.worksheets),
        rows_omitted=rows_omitted,
        cols_omitted=sum(max(0, ws.max_column - dg.MAX_COLS) for ws in wb.worksheets),
        budget_applied=False,
    )
    return Rendered(text, meta)


# Optional renderer modules register themselves on import; missing ones are simply unavailable.
for _mod in ("grid", "layouts", "compact", "schema"):
    try:
        importlib.import_module(f"serialize.{_mod}")
    except ModuleNotFoundError as e:
        if e.name != f"serialize.{_mod}":
            raise
