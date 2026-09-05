"""`grid` view and its layout variants (markdown, html, json, addressed).

Content (knob K1): formulas shown as `=F -> value` with values from a recalculated init, typed
column headers with number formats, dates marked as dates, percents shown both ways, defined
names, merged ranges, header-row guess, sheet list, and the current contents of the answer range.

Window (knob K2): header rows, the first body rows, the answer range +-3, rows referenced by
formulas near the answer, the last rows; sheets named in the instruction or referenced from the
answer region are rendered too. Everything else is summarised as an omission marker. A token
budget trims droppable rows first (never header or answer rows).

Layout (knob K3): the same shown cells emitted as a tab grid, a markdown table, an HTML table,
JSON records, or address-tagged cells. Same content, different shape.
"""

from __future__ import annotations

import html
import json
import re
from dataclasses import dataclass, field

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from sb import answer_ranges
from serialize import Rendered, answer_visibility, register
from serialize.tokens import count_tokens
from workbook import (
    WorkbookInfo, cell_text, coord_to_rc, formula_text, load_info, referenced_rows, referenced_sheets,
)

DEFAULT_BUDGET = 10_000
ANSWER_SHEET_SHARE = 0.6
MAX_COLS = 40
FIRST_BODY_ROWS = 10
LAST_ROWS = 3
ANSWER_PAD = 3
SMALL_SHEET_ROWS = 40         # sheets this small are shown whole
MAX_ROI_SHEETS = 6
MAX_ANSWER_CELLS_LISTED = 20
CHARS_PER_TOKEN = 3.2         # for trimming; the exact count is taken at the end


@dataclass
class SheetPlan:
    title: str
    rows: list[int]                    # shown rows, ascending
    protected: set[int]                # header + answer rows: never dropped by the budget
    ncols: int
    roi: bool
    note: str = ""
    markers: dict = field(default_factory=dict)   # row -> custom omission marker (compact view)
    extra_lines: list = field(default_factory=list)  # lines printed under the sheet title (compact/schema)


@dataclass
class GridModel:
    info: WorkbookInfo
    task: dict
    plans: list[SheetPlan]
    answer_cells_listed: list[tuple[str, str, str]]   # (sheet, coord, current display)
    preamble: list[str] = field(default_factory=list)


# ----------------------------------------------------------------------------- planning

def _answer_regions(task: dict, info: WorkbookInfo) -> dict[str, list[tuple[int, int, int, int]]]:
    """sheet title -> [(min_col, min_row, max_col, max_row)] for the graded ranges."""
    regions: dict[str, list] = {}
    for sheet, rng in answer_ranges(task):
        title = sheet if sheet and sheet in info.wb_f.sheetnames else info.active_title
        try:
            c0, r0, c1, r1 = range_boundaries(rng)
        except ValueError:
            continue
        ws = info.wb_f[title]
        r0 = r0 or 1
        r1 = r1 or ws.max_row
        regions.setdefault(title, []).append((c0 or 1, r0, c1 or ws.max_column, r1))
    return regions


def _mentioned_sheets(task: dict, info: WorkbookInfo) -> set[str]:
    text = task["instruction"].casefold()
    return {s.title for s in info.sheets if len(s.title) >= 3 and s.title.casefold() in text}


def plan_sheets(task: dict, info: WorkbookInfo) -> tuple[list[SheetPlan], dict]:
    regions = _answer_regions(task, info)
    answer_titles = set(regions)
    roi_titles = set(answer_titles) | _mentioned_sheets(task, info)

    # sheets referenced by formulas in/near the answer region
    for title, regs in regions.items():
        ws = info.wb_f[title]
        for c0, r0, c1, r1 in regs:
            for r in range(max(1, r0 - ANSWER_PAD), min(ws.max_row, r1 + ANSWER_PAD) + 1):
                for c in range(c0, min(c1, ws.max_column) + 1):
                    ftext = formula_text(ws.cell(row=r, column=c).value)
                    if ftext:
                        roi_titles |= {s for s in referenced_sheets(ftext) if s in info.wb_f.sheetnames}
    if len(info.sheets) <= MAX_ROI_SHEETS:
        roi_titles |= {s.title for s in info.sheets}
    if len(roi_titles) > MAX_ROI_SHEETS:
        keep = set(answer_titles)
        for t in [s.title for s in info.sheets]:
            if len(keep) >= MAX_ROI_SHEETS:
                break
            if t in roi_titles:
                keep.add(t)
        roi_titles = keep

    plans = []
    for s in info.sheets:
        ws = info.wb_f[s.title]
        ncols = min(s.max_col, MAX_COLS if s.max_row > 20 else 2 * MAX_COLS)  # wide-but-short sheets fit
        if s.title not in roi_titles:
            plans.append(SheetPlan(s.title, rows=list(range(1, s.header_row + 1)), protected=set(range(1, s.header_row + 1)),
                                   ncols=ncols, roi=False, note="header only"))
            continue
        header = set(range(1, s.header_row + 1))
        if s.max_row <= SMALL_SHEET_ROWS:
            rows = set(range(1, s.max_row + 1))
            protected = set(rows)
        else:
            rows = set(header)
            rows |= set(range(s.header_row + 1, min(s.max_row, s.header_row + FIRST_BODY_ROWS) + 1))
            rows |= set(range(max(1, s.max_row - LAST_ROWS + 1), s.max_row + 1))
            protected = set(header)
            for c0, r0, c1, r1 in regions.get(s.title, []):
                lo, hi = max(1, r0 - ANSWER_PAD), min(s.max_row, r1 + ANSWER_PAD)
                if hi - lo + 1 <= 2 * (FIRST_BODY_ROWS + ANSWER_PAD):
                    span = set(range(lo, hi + 1))
                else:  # huge answer range: both ends
                    span = set(range(lo, lo + FIRST_BODY_ROWS + ANSWER_PAD)) | set(range(hi - FIRST_BODY_ROWS - ANSWER_PAD + 1, hi + 1))
                rows |= span
                protected |= span
                # rows referenced by formulas in the answer region (same sheet)
                refd = set()
                for r in range(max(1, r0 - ANSWER_PAD), min(ws.max_row, r1 + ANSWER_PAD) + 1):
                    for c in range(c0, min(c1, ws.max_column) + 1):
                        ftext = formula_text(ws.cell(row=r, column=c).value)
                        if ftext:
                            refd |= {x for x in referenced_rows(ftext) if 1 <= x <= s.max_row}
                    if len(refd) > 20:
                        break
                rows |= set(sorted(refd)[:20])
        plans.append(SheetPlan(s.title, rows=sorted(rows), protected=protected, ncols=ncols, roi=True))
    return plans, regions


# ----------------------------------------------------------------------------- model

def build_model(path: str, task: dict, budget_tokens: int | None, planner=None) -> GridModel:
    """Shared by grid/layouts (default planner), compact and schema (their own planners)."""
    info = load_info(path, recalc=True)
    plans, regions = (planner or plan_sheets)(task, info)

    listed = []
    for title, regs in regions.items():
        for c0, r0, c1, r1 in regs:
            for r in range(r0, r1 + 1):
                for c in range(c0, c1 + 1):
                    if len(listed) >= MAX_ANSWER_CELLS_LISTED:
                        break
                    listed.append((title, f"{get_column_letter(c)}{r}", cell_text(info, title, r, c) or "(empty)"))

    pre = []
    pre.append(f"Workbook: {len(info.sheets)} sheet(s). Active sheet: {info.active_title!r}.")
    for s in info.sheets:
        extra = []
        if s.n_formula:
            extra.append(f"{s.n_formula} formula cells")
        if s.merged:
            extra.append(f"merged: {', '.join(s.merged[:8])}" + (" …" if len(s.merged) > 8 else ""))
        pre.append(f"- {s.title!r}: {s.max_row} rows x {s.max_col} cols, header row {s.header_row}"
                   + (f" ({'; '.join(extra)})" if extra else ""))
    if info.defined_names:
        items = list(info.defined_names.items())[:15]
        pre.append("Defined names: " + "; ".join(f"{k} = {v}" for k, v in items) + (" …" if len(info.defined_names) > 15 else ""))
    if info.cached_values_missing:
        pre.append("Note: some formula cells had no saved values; values shown come from a LibreOffice recalculation."
                   if info.recalc_ok else "Note: some formula cells have no saved values (shown as (uncomputed)).")
    n_answer = sum((r1 - r0 + 1) * (c1 - c0 + 1) for regs in regions.values() for c0, r0, c1, r1 in regs)
    pre.append(f"Graded answer range: {task['answer_position']} on "
               f"{', '.join(repr(t) for t in regions) or 'the active sheet'} ({n_answer} cells).")
    if listed:
        shown = "; ".join(f"{coord}={txt}" for _, coord, txt in listed)
        pre.append(f"Current contents of the answer range{' (first %d cells)' % MAX_ANSWER_CELLS_LISTED if n_answer > MAX_ANSWER_CELLS_LISTED else ''}: {shown}")

    model = GridModel(info=info, task=task, plans=plans, answer_cells_listed=listed, preamble=pre)
    apply_budget(model, budget_tokens or DEFAULT_BUDGET)
    return model


def apply_budget(model: GridModel, budget_tokens: int) -> None:
    """Drop droppable rows (never header/answer rows), farthest-from-answer first, until under budget.

    Each row's character cost is computed once; rows are then removed in one pass in distance order,
    so the cost is linear in the rows shown rather than quadratic.
    """
    budget_chars = int(budget_tokens * CHARS_PER_TOKEN)
    pre_chars = sum(len(p) for p in model.preamble) + 50
    roi = [p for p in model.plans if p.roi]
    if not roi:
        return
    answer_titles = {t for t, _, _ in model.answer_cells_listed} or {model.info.active_title}
    shares = {}
    others = [p for p in roi if p.title not in answer_titles]
    for p in roi:
        if p.title in answer_titles:
            shares[p.title] = ANSWER_SHEET_SHARE / max(1, len([q for q in roi if q.title in answer_titles]))
        else:
            shares[p.title] = (1 - ANSWER_SHEET_SHARE) / max(1, len(others))
    for p in roi:
        limit = max(400, int((budget_chars - pre_chars) * shares[p.title]))
        cost = {r: len(_row_cells(model, p, r, joined=True)) + 4 for r in p.rows}
        total = sum(cost.values()) + 200
        if total <= limit:
            continue
        header_rows = set(range(1, model.info.sheet(p.title).header_row + 1))
        anchor = min(p.protected - header_rows or p.protected or {1})
        droppable = sorted((r for r in p.rows if r not in p.protected), key=lambda r: -abs(r - anchor))
        dropped = 0
        for victim in droppable:
            if total <= limit:
                break
            total -= cost[victim]
            p.rows.remove(victim)
            dropped += 1
        if dropped:
            p.note = f"{dropped} rows dropped for the token budget"


# ----------------------------------------------------------------------------- emitters

def _row_cells(model: GridModel, plan: SheetPlan, r: int, joined: bool = False):
    cells = [cell_text(model.info, plan.title, r, c) for c in range(1, plan.ncols + 1)]
    return "\t".join(cells) if joined else cells


def _col_labels(model: GridModel, plan: SheetPlan) -> list[str]:
    s = model.info.sheet(plan.title)
    return [s.columns[c - 1].label() if c - 1 < len(s.columns) else get_column_letter(c) for c in range(1, plan.ncols + 1)]


def _omission_lines(plan: SheetPlan, max_row: int) -> dict[int, str]:
    """row number -> marker to print before it."""
    marks = {}
    prev = 0
    for r in plan.rows:
        if r > prev + 1:
            marks[r] = plan.markers.get(r) or f"... rows {prev + 1}-{r - 1} omitted ({r - 1 - prev} rows) ..."
        prev = r
    if prev < max_row:
        marks[max_row + 1] = plan.markers.get(max_row + 1) or f"... rows {prev + 1}-{max_row} omitted ({max_row - prev} rows) ..."
    return marks


def _sheet_title_line(model: GridModel, plan: SheetPlan) -> str:
    s = model.info.sheet(plan.title)
    shown = f"showing {len(plan.rows)} of {s.max_row} rows"
    cols = f", columns A-{get_column_letter(plan.ncols)}" + (f" of {s.max_col}" if plan.ncols < s.max_col else "")
    note = f"; {plan.note}" if plan.note else ""
    return f"### Sheet {plan.title!r} ({shown}{cols}{note})"


def emit_tsv(model: GridModel) -> str:
    parts = ["\n".join(model.preamble)]
    for plan in model.plans:
        s = model.info.sheet(plan.title)
        lines = [_sheet_title_line(model, plan), *plan.extra_lines, "\t".join([""] + _col_labels(model, plan))]
        marks = _omission_lines(plan, s.max_row)
        for r in plan.rows:
            if r in marks:
                lines.append(marks[r])
            lines.append("\t".join([str(r)] + _row_cells(model, plan, r)))
        if s.max_row + 1 in marks:
            lines.append(marks[s.max_row + 1])
        parts.append("\n".join(lines))
    return "\n\n".join(parts)


def emit_markdown(model: GridModel) -> str:
    parts = ["\n".join(model.preamble)]
    for plan in model.plans:
        s = model.info.sheet(plan.title)
        labels = _col_labels(model, plan)
        lines = [_sheet_title_line(model, plan), "| row | " + " | ".join(labels) + " |",
                 "|---|" + "---|" * len(labels)]
        marks = _omission_lines(plan, s.max_row)
        for r in plan.rows:
            if r in marks:
                lines.append(f"| … | {marks[r]} |")
            cells = [c.replace("|", "\\|") for c in _row_cells(model, plan, r)]
            lines.append(f"| {r} | " + " | ".join(cells) + " |")
        if s.max_row + 1 in marks:
            lines.append(f"| … | {marks[s.max_row + 1]} |")
        parts.append("\n".join(lines))
    return "\n\n".join(parts)


def emit_html(model: GridModel) -> str:
    parts = ["\n".join(model.preamble)]
    for plan in model.plans:
        s = model.info.sheet(plan.title)
        labels = _col_labels(model, plan)
        lines = [f"<h3>{html.escape(_sheet_title_line(model, plan)[4:])}</h3>", "<table>",
                 "<tr><th></th>" + "".join(f"<th>{html.escape(l)}</th>" for l in labels) + "</tr>"]
        marks = _omission_lines(plan, s.max_row)
        for r in plan.rows:
            if r in marks:
                lines.append(f"<tr><td colspan={len(labels) + 1}>{html.escape(marks[r])}</td></tr>")
            cells = "".join(f"<td>{html.escape(c)}</td>" for c in _row_cells(model, plan, r))
            lines.append(f"<tr><th>{r}</th>{cells}</tr>")
        if s.max_row + 1 in marks:
            lines.append(f"<tr><td colspan={len(labels) + 1}>{html.escape(marks[s.max_row + 1])}</td></tr>")
        lines.append("</table>")
        parts.append("\n".join(lines))
    return "\n\n".join(parts)


def emit_json(model: GridModel) -> str:
    out = {"summary": model.preamble, "sheets": []}
    for plan in model.plans:
        s = model.info.sheet(plan.title)
        cols = {get_column_letter(c): lab for c, lab in enumerate(_col_labels(model, plan), start=1)}
        rows = []
        marks = _omission_lines(plan, s.max_row)
        for r in plan.rows:
            if r in marks:
                rows.append({"omitted": marks[r]})
            cells = _row_cells(model, plan, r)
            rows.append({"row": r, **{get_column_letter(c): v for c, v in enumerate(cells, start=1) if v != ""}})
        if s.max_row + 1 in marks:
            rows.append({"omitted": marks[s.max_row + 1]})
        out["sheets"].append({"name": plan.title, "rows_total": s.max_row, "cols_total": s.max_col,
                              "header_row": s.header_row, "columns": cols, "rows": rows, "note": plan.note})
    return json.dumps(out, ensure_ascii=False, indent=None, separators=(",", ":"))


def emit_addressed(model: GridModel) -> str:
    parts = ["\n".join(model.preamble)]
    for plan in model.plans:
        s = model.info.sheet(plan.title)
        lines = [_sheet_title_line(model, plan), "Columns: " + ", ".join(_col_labels(model, plan))]
        marks = _omission_lines(plan, s.max_row)
        for r in plan.rows:
            if r in marks:
                lines.append(marks[r])
            cells = _row_cells(model, plan, r)
            tagged = [f"{get_column_letter(c)}{r}={v}" for c, v in enumerate(cells, start=1) if v != ""]
            lines.append(" ".join(tagged) if tagged else f"(row {r} empty)")
        if s.max_row + 1 in marks:
            lines.append(marks[s.max_row + 1])
        parts.append("\n".join(lines))
    return "\n\n".join(parts)


EMITTERS = {"grid": emit_tsv, "markdown": emit_markdown, "html": emit_html, "json": emit_json, "addressed": emit_addressed}


def _meta(model: GridModel) -> dict:
    shown = {p.title: (set(p.rows), p.ncols) for p in model.plans}
    dims = {s.title: (s.max_row, s.max_col) for s in model.info.sheets}  # as loaded, before cell access widened them
    meta = answer_visibility(model.task, model.info.wb_f, shown, dims)
    meta.update(
        sheets_rendered=sum(1 for p in model.plans if p.roi),
        sheets_total=len(model.plans),
        rows_omitted=sum(model.info.sheet(p.title).max_row - len(p.rows) for p in model.plans),
        cols_omitted=sum(max(0, model.info.sheet(p.title).max_col - p.ncols) for p in model.plans),
        recalc_ok=model.info.recalc_ok,
        cached_values_missing=model.info.cached_values_missing,
        defined_names=len(model.info.defined_names),
        budget_applied=True,
    )
    return meta


def _make(layout: str):
    def _render(path: str, task: dict, budget_tokens: int | None) -> Rendered:
        model = build_model(path, task, budget_tokens)
        text = EMITTERS[layout](model)
        meta = _meta(model)
        meta["layout"] = layout
        return Rendered(text, meta)
    return _render


for _name in EMITTERS:
    register(_name)(_make(_name))
