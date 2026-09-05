"""`compact` view: SheetCompressor-lite (after Microsoft's SpreadsheetLLM).

Keeps the rows that carry structure and collapses the rest:
- structural anchors: header rows, rows whose type signature differs from both neighbours, blank
  separator rows next to content, the first data row, the last row, the answer range +-2;
- runs of rows with the same type signature collapse to one marker
  ("rows 14-1190 (1177 rows): same pattern as row 13");
- per-column aggregates under the sheet title (type, number format, count, min/max or examples,
  formula pattern in R1C1 with its share) so the model knows a column's shape without seeing it.

Same preamble, cell rendering and token budgeting as `grid`.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from serialize import Rendered, register
from serialize.grid import (
    ANSWER_PAD, DEFAULT_BUDGET, MAX_COLS, MAX_ROI_SHEETS, GridModel, SheetPlan, _answer_regions,
    _meta, _mentioned_sheets, apply_budget, build_model, emit_tsv,
)
from workbook import cell_kind, fmt_number, formula_text

ANCHOR_PAD = 2
MAX_RUN_KEEP = 3          # a run shorter than this is shown, not collapsed
SIGNATURE_COLS = 40
MAX_SCAN_ROWS = 5000      # rows scanned for structure; longer sheets summarise the tail
MAX_ANCHORS = 60          # cap on structural anchors per sheet (each costs ~5 shown rows)


def _signature(ws, r: int, ncols: int) -> tuple:
    sig = []
    for c in range(1, min(ncols, SIGNATURE_COLS) + 1):
        cell = ws.cell(row=r, column=c)
        k = cell_kind(cell.value, cell.number_format or "General")
        sig.append("num" if k in ("number", "currency", "percent") else k)
    return tuple(sig)


def _column_aggregates(info, title: str, ncols: int) -> list[str]:
    s = info.sheet(title)
    ws_v = info.wb_v[title] if title in info.wb_v.sheetnames else None
    ws_f = info.wb_f[title]
    lines = []
    for c in range(1, min(ncols, len(s.columns)) + 1):
        col = s.columns[c - 1]
        if col.n_nonempty == 0:
            continue
        head = f" {col.header!r}" if col.header else ""
        desc = f"{col.letter}{head}: {col.dtype}"
        if col.number_format not in ("General", "", None):
            desc += f' "{col.number_format}"'
        desc += f", {col.n_nonempty} values"
        blanks = max(0, s.scanned_rows - s.header_row - col.n_nonempty)
        if blanks:
            desc += f", {blanks} blanks"
        if col.dtype in ("number", "currency", "percent") and ws_v is not None:
            nums = []
            for r in range(s.header_row + 1, s.scanned_rows + 1):
                v = ws_v.cell(row=r, column=c).value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    nums.append(v)
            if nums:
                desc += f", min {fmt_number(min(nums))}, max {fmt_number(max(nums))}"
        elif col.dtype in ("text", "date", "datetime", "bool"):
            if col.distinct is not None and len(col.distinct) <= 8:
                desc += f", distinct: {', '.join(str(d)[:24] for d in col.distinct)}"
            elif col.samples:
                desc += f", e.g. {', '.join(str(x)[:24] for x in col.samples)}"
        if col.formula_pattern:
            first = None
            for r in range(s.header_row + 1, s.scanned_rows + 1):
                ft = formula_text(ws_f.cell(row=r, column=c).value)
                if ft:
                    first = f"{col.letter}{r}: {ft[:60]}"
                    break
            desc += f"; formula {col.formula_pattern[:80]} ({col.formula_pattern_share:.0%} of {col.n_formula}" + (f", e.g. {first})" if first else ")")
        lines.append("  " + desc)
    return lines


def plan_compact(task: dict, info) -> tuple[list[SheetPlan], dict]:
    regions = _answer_regions(task, info)
    roi_titles = set(regions) | _mentioned_sheets(task, info)
    if len(info.sheets) <= MAX_ROI_SHEETS:
        roi_titles |= {s.title for s in info.sheets}
    plans = []
    for s in info.sheets:
        ws = info.wb_f[s.title]
        ncols = min(s.max_col, MAX_COLS if s.max_row > 20 else 2 * MAX_COLS)
        header = set(range(1, s.header_row + 1))
        if s.title not in roi_titles:
            plans.append(SheetPlan(s.title, rows=sorted(header), protected=set(header), ncols=ncols, roi=False, note="header only"))
            continue
        n = s.max_row
        scan = min(n, MAX_SCAN_ROWS)  # beyond this the tail is summarised, not scanned
        sigs = {r: _signature(ws, r, ncols) for r in range(1, scan + 1)}
        if n > scan:
            sigs[n] = _signature(ws, n, ncols)
        blank = {r for r in sigs if all(k == "empty" for k in sigs[r])}
        anchors = set(header) | {min(n, s.header_row + 1), n}
        for r in range(2, scan):
            if sigs[r] != sigs[r - 1] and sigs[r] != sigs[r + 1]:
                anchors.add(r)
            if r in blank and (r - 1 not in blank or r + 1 not in blank):
                anchors.add(r)
            if r not in blank and (r - 1 in blank):  # first row of a block
                anchors.add(r)
        if len(anchors) > MAX_ANCHORS:  # very heterogeneous sheet: keep the anchors nearest the answer
            keep = set(header) | {n}
            focus = min(regions.get(s.title, [(1, s.header_row + 1, 1, s.header_row + 1)])[0][1], n)
            keep |= set(sorted(anchors, key=lambda a: abs(a - focus))[:MAX_ANCHORS])
            anchors = keep
        protected = set(header)
        for c0, r0, c1, r1 in regions.get(s.title, []):
            lo, hi = max(1, r0 - ANSWER_PAD), min(n, r1 + ANSWER_PAD)
            span = set(range(lo, hi + 1)) if hi - lo + 1 <= 30 else set(range(lo, lo + 15)) | set(range(hi - 14, hi + 1))
            protected |= span
        rows = set(protected)
        for a in anchors:
            rows |= set(range(max(1, a - ANCHOR_PAD), min(n, a + ANCHOR_PAD) + 1))
        # short gaps are cheaper to show than to describe
        srt = sorted(rows)
        for a, b in zip(srt, srt[1:]):
            if 1 < b - a - 1 < MAX_RUN_KEEP:
                rows |= set(range(a + 1, b))
        rows = sorted(rows)
        markers = {}
        prev = 0
        for r in rows + [n + 1]:
            if r > prev + 1:
                run = range(prev + 1, r)
                ref = prev if prev >= 1 else None
                same = sum(1 for x in run if ref and x in sigs and ref in sigs and sigs[x] == sigs[ref])
                if ref and any(x not in sigs for x in run):  # unscanned tail
                    markers[r] = f"... rows {run.start}-{run.stop - 1} ({len(run)} rows): not scanned, likely same pattern as row {ref} ..."
                    prev = r
                    continue
                if ref and same == len(run):
                    desc = f"same pattern as row {ref}"
                elif ref and same >= 0.6 * len(run):
                    desc = f"{same / len(run):.0%} same pattern as row {ref}, rest mixed"
                elif all(x in blank for x in run):
                    desc = "empty"
                else:
                    desc = "mixed rows"
                markers[r] = f"... rows {run.start}-{run.stop - 1} ({len(run)} rows): {desc} ..."
            prev = r
        plans.append(SheetPlan(s.title, rows=rows, protected=protected, ncols=ncols, roi=True, markers=markers,
                               extra_lines=["Columns:"] + _column_aggregates(info, s.title, ncols)))
    return plans, regions


@register("compact")
def render_compact(path: str, task: dict, budget_tokens: int | None) -> Rendered:
    model: GridModel = build_model(path, task, budget_tokens, planner=plan_compact)
    text = emit_tsv(model)
    meta = _meta(model)
    meta["layout"] = "compact"
    return Rendered(text, meta)
