"""Deterministic workbook checks used by the spreadsheet agent."""

from __future__ import annotations

from collections import defaultdict

import openpyxl
from openpyxl.utils.cell import range_boundaries

from sb import answer_ranges, parse_answer_position


def _cell_value(cell):
    """Include value type so ``1`` and ``"1"`` are not treated as identical."""
    return type(cell.value).__name__, cell.value


def _range_cells(ws, cell_range: str):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    min_row = min_row or 1
    max_row = max_row or ws.max_row
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        yield from row


def formula_cells(path: str, task: dict) -> list[str]:
    """List formulas in the graded area; these require LibreOffice cache refresh."""
    wb = openpyxl.load_workbook(path, data_only=False)
    found = []
    for sheet, cell_range in answer_ranges(task):
        if sheet not in wb.sheetnames:
            continue
        for cell in _range_cells(wb[sheet], cell_range):
            if isinstance(cell.value, str) and cell.value.startswith("="):
                found.append(f"{sheet}!{cell.coordinate}")
    return found


EXCEL_ERRORS = {"#NAME?", "#VALUE!", "#N/A", "#REF!", "#DIV/0!", "#NULL!", "#NUM!", "#SPILL!", "#CALC!"}


def error_cells_in_answer_range(path: str, task: dict) -> list[str]:
    """Graded cells whose recalculated value is an Excel error string.

    formula_cells only reports that a formula exists; this checks what it actually produced,
    using the cached (post-recalculation) value so a broken formula can't slip through unnoticed.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    found = []
    for sheet, cell_range in answer_ranges(task):
        if sheet not in wb.sheetnames:
            continue
        for cell in _range_cells(wb[sheet], cell_range):
            if isinstance(cell.value, str) and cell.value.strip() in EXCEL_ERRORS:
                found.append(f"{sheet}!{cell.coordinate}={cell.value}")
    return found


def answer_range_coverage(init_path: str, current_path: str, task: dict) -> str:
    """Per-column count of graded cells still identical to the initial workbook.

    A cumulative check against the task's starting state (not just this edit's diff), so a
    column the model never touched across any turn is surfaced explicitly instead of silently
    shipping unchanged.
    """
    init = openpyxl.load_workbook(init_path, data_only=False)
    cur = openpyxl.load_workbook(current_path, data_only=False)
    lines = []
    for sheet, cell_range in answer_ranges(task):
        if sheet not in cur.sheetnames or sheet not in init.sheetnames:
            lines.append(f"{sheet!r}!{cell_range}: SHEET MISSING")
            continue
        by_col: dict[str, list[int]] = defaultdict(lambda: [0, 0])  # col letter -> [untouched, total]
        for cell in _range_cells(cur[sheet], cell_range):
            col = "".join(ch for ch in cell.coordinate if ch.isalpha())
            init_cell = init[sheet][cell.coordinate]
            by_col[col][1] += 1
            if _cell_value(cell) == _cell_value(init_cell):
                by_col[col][0] += 1
        stale = [f"{col}: {untouched}/{total} unchanged" for col, (untouched, total) in sorted(by_col.items()) if untouched == total]
        if stale:
            lines.append(f"{sheet}: columns fully untouched since start (still identical to initial workbook) — {', '.join(stale)}")
    return "\n".join(lines) if lines else "All declared columns have at least one edited cell."


def _normalise_expected_ranges(task: dict, expected_changes: list[str] | None) -> dict[str, list[str]]:
    result: dict[str, list[str]] = defaultdict(list)
    for item in expected_changes or []:
        for sheet, cell_range in parse_answer_position(str(item)):
            result[sheet or task.get("answer_sheet")].append(cell_range)
    return result


def _in_expected(sheet: str, coordinate: str, expected: dict[str, list[str]]) -> bool:
    if sheet not in expected:
        return False
    ws_coordinate = coordinate
    for cell_range in expected[sheet]:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        min_row = min_row or 1
        max_row = max_row or 1_048_576
        col = openpyxl.utils.column_index_from_string("".join(ch for ch in ws_coordinate if ch.isalpha()))
        row = int("".join(ch for ch in ws_coordinate if ch.isdigit()))
        if min_col <= col <= max_col and min_row <= row <= max_row:
            return True
    return False


def diff_workbooks(before_path: str, after_path: str, task: dict, expected_changes: list[str] | None = None,
                   *, sample_limit: int = 30) -> dict:
    """Compare logical cell values/formulas and report unexpected mutations compactly."""
    before = openpyxl.load_workbook(before_path, data_only=False)
    after = openpyxl.load_workbook(after_path, data_only=False)
    expected = _normalise_expected_ranges(task, expected_changes)
    changed, unexpected = [], []
    sheet_names = sorted(set(before.sheetnames) | set(after.sheetnames))
    for sheet in sheet_names:
        if sheet not in before.sheetnames or sheet not in after.sheetnames:
            changed.append(f"{sheet}!<sheet-added-or-removed>")
            unexpected.append(changed[-1])
            continue
        left, right = before[sheet], after[sheet]
        max_row, max_col = max(left.max_row, right.max_row), max(left.max_column, right.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                if _cell_value(left.cell(row, col)) == _cell_value(right.cell(row, col)):
                    continue
                label = f"{sheet}!{left.cell(row, col).coordinate}"
                changed.append(label)
                if expected_changes is not None and not _in_expected(sheet, left.cell(row, col).coordinate, expected):
                    unexpected.append(label)
    return {
        "changed_count": len(changed),
        "changed_sample": changed[:sample_limit],
        "unexpected_count": len(unexpected),
        "unexpected_sample": unexpected[:sample_limit],
    }


def format_verification(diff: dict, formulas: list[str]) -> str:
    lines = [
        "## Deterministic verification",
        f"Changed cells: {diff['changed_count']}" + (f"; sample: {', '.join(diff['changed_sample'])}" if diff["changed_sample"] else ""),
        f"Unexpected changed cells: {diff['unexpected_count']}" + (
            f"; sample: {', '.join(diff['unexpected_sample'])}" if diff["unexpected_sample"] else ""
        ),
        f"Formula cells in graded range: {len(formulas)}" + (f"; sample: {', '.join(formulas[:20])}" if formulas else ""),
    ]
    return "\n".join(lines)
