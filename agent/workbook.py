"""Workbook inspection shared by the renderers: formula view + values view, column profiles,
header-row guess, merged ranges, defined names, and a cached LibreOffice-recalculated copy of the
init workbook so formula cells have values even when the file saved none.

Only init workbooks are opened here. Nothing in this module knows about golden files.
"""

from __future__ import annotations

import datetime
import hashlib
import os
import threading
import re
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

HERE = Path(__file__).resolve().parent
for _candidate in (HERE.parent, HERE.parent / "research"):
    if (_candidate / "sb.py").exists():
        if str(_candidate) not in sys.path:
            sys.path.insert(0, str(_candidate))
        break

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from sb import recalculate

RECALC_CACHE = Path(os.environ.get("RECALC_CACHE") or (Path(tempfile.gettempdir()) / "sb_recalc_cache"))
CELL_CHARS = 60
PROFILE_ROWS = 2000  # rows scanned per sheet for column statistics

ERROR_STRINGS = {"#NAME?", "#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NUM!", "#NULL!", "#SPILL!"}


# ----------------------------------------------------------------------------- recalculation

def recalculated_copy(path: str) -> tuple[str, bool]:
    """(path to a LibreOffice-recalculated copy, ok). Cached by content hash; on failure (path, False)."""
    p = Path(path)
    digest = hashlib.sha256(p.read_bytes()).hexdigest()[:16]
    out_dir = RECALC_CACHE / digest
    target = out_dir / p.name
    if target.exists():
        return str(target), True
    try:
        recalculate(path, out_dir)
    except Exception:
        return path, False
    return (str(target), True) if target.exists() else (path, False)


# ----------------------------------------------------------------------------- cell typing

def formula_text(value) -> str | None:
    """Formula source for a cell value: plain '=...' strings and openpyxl ArrayFormula/DataTableFormula objects."""
    if isinstance(value, str):
        return value if value.startswith("=") else None
    text = getattr(value, "text", None)
    if isinstance(text, str) and type(value).__name__.endswith("Formula"):
        text = text if text.startswith("=") else "=" + text
        ref = getattr(value, "ref", None)
        return f"{text} {{array {ref}}}" if ref else text
    return None


def cell_kind(value, number_format: str = "General") -> str:
    if value is None or value == "":
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if formula_text(value) is not None:
        return "formula"
    if isinstance(value, str):
        if value in ERROR_STRINGS:
            return "error"
        return "text"
    if isinstance(value, datetime.datetime):
        return "date" if (value.hour, value.minute, value.second) == (0, 0, 0) else "datetime"
    if isinstance(value, datetime.date):
        return "date"
    if isinstance(value, datetime.time):
        return "time"
    if isinstance(value, datetime.timedelta):
        return "duration"
    if isinstance(value, (int, float)):
        fmt = number_format or "General"
        if "%" in fmt:
            return "percent"
        if any(s in fmt for s in ("$", "£", "€", "[$")):
            return "currency"
        return "number"
    return "other"


def fmt_number(v) -> str:
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v.is_integer() and abs(v) < 1e15:
            return str(int(v))
        s = f"{v:.6f}".rstrip("0").rstrip(".")
        return s if len(s) <= 14 else f"{v:.6g}"
    return str(v)


def display_value(value, number_format: str = "General", cached=None, show_formula_value: bool = True) -> str:
    """Human-readable cell text: formulas as `=F -> value`, dates marked, percents shown both ways."""
    kind = cell_kind(value, number_format)
    if kind == "empty":
        return ""
    if kind == "formula":
        text = formula_text(value) or str(value)
        if not show_formula_value:
            return _trim(text)
        tail = "(uncomputed)" if cached is None or cached == "" else display_value(cached, number_format)
        return f"{_trim(text)} -> {tail}"
    if kind == "date":
        return f"{value:%Y-%m-%d} (date)"
    if kind == "datetime":
        return f"{value:%Y-%m-%d %H:%M} (datetime)"
    if kind == "time":
        return f"{value:%H:%M} (time)"
    if kind == "duration":
        return f"{value} (duration)"
    if kind == "percent":
        return f"{fmt_number(value)} ({value * 100:.6g}%)"
    if kind in ("number", "currency", "bool"):
        return fmt_number(value)
    return _trim(str(value))


def _trim(s: str, n: int = CELL_CHARS) -> str:
    s = s.replace("\n", " ").replace("\t", " ")
    return s if len(s) <= n else s[: n - 1] + "…"


# ----------------------------------------------------------------------------- formulas

_REF = re.compile(r"(?<![A-Za-z_\d.])(\$?)([A-Z]{1,3})(\$?)(\d{1,7})(?![\d(])")
_SHEET_REF = re.compile(r"(?:'([^']+)'|([A-Za-z_][\w. ]*))!\$?[A-Z]{1,3}\$?\d*(?::\$?[A-Z]{1,3}\$?\d*)?")


def _split_quoted(formula: str):
    """Yield (segment, is_string_literal) so refs inside "..." are left alone."""
    parts = re.split(r'("[^"]*")', formula)
    for part in parts:
        yield part, part.startswith('"') and part.endswith('"') and len(part) >= 2


def to_r1c1(formula: str, row: int, col: int) -> str:
    """A1 -> relative R1C1 so copied-down formulas collapse to one pattern."""
    def repl(m):
        c_abs, c_letters, r_abs, r_digits = m.groups()
        c = column_index_from_string(c_letters)
        r = int(r_digits)
        cpart = f"C{c}" if c_abs else ("C" if c == col else f"C[{c - col}]")
        rpart = f"R{r}" if r_abs else ("R" if r == row else f"R[{r - row}]")
        return rpart + cpart

    out = []
    for seg, is_str in _split_quoted(formula):
        out.append(seg if is_str else _REF.sub(repl, seg))
    return "".join(out)


def referenced_sheets(formula: str) -> set[str]:
    names = set()
    for seg, is_str in _split_quoted(formula):
        if is_str:
            continue
        for m in _SHEET_REF.finditer(seg):
            names.add(m.group(1) or m.group(2))
    return names


def referenced_rows(formula: str, same_sheet_only: bool = True) -> set[int]:
    """Row numbers referenced by a formula (same-sheet refs only by default)."""
    rows = set()
    for seg, is_str in _split_quoted(formula):
        if is_str:
            continue
        if same_sheet_only:
            seg = _SHEET_REF.sub("", seg)
        for m in _REF.finditer(seg):
            rows.add(int(m.group(4)))
    return rows


# ----------------------------------------------------------------------------- profiles

@dataclass
class ColumnProfile:
    letter: str
    header: str | None
    dtype: str                 # dominant kind among non-empty cells (excluding the header row)
    number_format: str         # dominant number format among non-empty cells
    n_nonempty: int
    n_formula: int
    samples: list = field(default_factory=list)
    distinct: list | None = None
    formula_pattern: str | None = None   # dominant R1C1 pattern
    formula_pattern_share: float = 0.0
    kinds: dict = field(default_factory=dict)

    def label(self) -> str:
        fmt = "" if self.number_format in ("General", "", None) else f' "{self.number_format}"'
        return f"{self.letter} [{self.dtype}{fmt}]"


@dataclass
class SheetInfo:
    title: str
    max_row: int
    max_col: int
    header_row: int
    merged: list[str]
    columns: list[ColumnProfile]
    n_formula: int
    is_active: bool
    scanned_rows: int


@dataclass
class WorkbookInfo:
    path: str
    values_path: str
    recalc_ok: bool
    cached_values_missing: bool
    sheets: list[SheetInfo]
    defined_names: dict[str, str]
    active_title: str
    wb_f: object   # openpyxl workbook, formulas preserved
    wb_v: object   # openpyxl workbook, values (from the recalculated copy when available)

    def sheet(self, title: str) -> SheetInfo | None:
        return next((s for s in self.sheets if s.title == title), None)


def guess_header_row(ws, max_scan: int = 10) -> int:
    """First row (within max_scan) that is mostly text and is followed by a row that looks different.

    Falls back to the first row with at least two non-empty cells, then to row 1. Never drifts to
    the last scanned row: a block of uniform text rows is data, not a header.
    """
    def signature(r):
        kinds = []
        for c in range(1, min(ws.max_column, 60) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and v != "":
                kinds.append("text" if isinstance(v, str) and not v.startswith("=") else "other")
        return kinds

    first_populated = None
    for r in range(1, min(ws.max_row, max_scan) + 1):
        sig = signature(r)
        if len(sig) < 2:
            continue
        if first_populated is None:
            first_populated = r
        text_share = sig.count("text") / len(sig)
        if text_share < 0.6:
            continue
        for r2 in range(r + 1, min(ws.max_row, r + 4) + 1):
            sig2 = signature(r2)
            if not sig2:
                continue
            if sig2.count("text") / len(sig2) < text_share or len(sig2) > len(sig):
                return r
            break
    return first_populated or 1


def profile_sheet(ws_f, ws_v, header_row: int) -> list[ColumnProfile]:
    cols = []
    last_row = min(ws_f.max_row, PROFILE_ROWS)
    for c in range(1, min(ws_f.max_column, 200) + 1):
        letter = get_column_letter(c)
        header_v = ws_f.cell(row=header_row, column=c).value
        header = _trim(str(header_v), 40) if header_v not in (None, "") else None
        kinds: dict[str, int] = {}
        formats: dict[str, int] = {}
        patterns: dict[str, int] = {}
        samples: list = []
        distinct: set = set()
        n_nonempty = n_formula = 0
        for r in range(header_row + 1, last_row + 1):
            cell = ws_f.cell(row=r, column=c)
            v = cell.value
            if v is None or v == "":
                continue
            n_nonempty += 1
            fmt = cell.number_format or "General"
            ftext = formula_text(v)
            if ftext is not None:
                n_formula += 1
                pat = to_r1c1(ftext, r, c)
                patterns[pat] = patterns.get(pat, 0) + 1
                vv = ws_v.cell(row=r, column=c).value if ws_v is not None else None
                k = cell_kind(vv, fmt) if vv not in (None, "") else "formula"
            else:
                k = cell_kind(v, fmt)
                vv = v
            kinds[k] = kinds.get(k, 0) + 1
            formats[fmt] = formats.get(fmt, 0) + 1
            if len(samples) < 3 and vv not in (None, ""):
                samples.append(display_value(vv, fmt))
            if len(distinct) <= 8 and vv not in (None, ""):
                distinct.add(vv if isinstance(vv, (str, int, float, bool)) else str(vv))
        dtype = max(kinds, key=kinds.get) if kinds else "empty"
        number_format = max(formats, key=formats.get) if formats else "General"
        pattern = share = None
        if patterns:
            pattern = max(patterns, key=patterns.get)
            share = patterns[pattern] / n_formula
        cols.append(ColumnProfile(
            letter=letter, header=header, dtype=dtype, number_format=number_format,
            n_nonempty=n_nonempty, n_formula=n_formula, samples=samples,
            distinct=sorted(distinct, key=str) if len(distinct) <= 8 else None,
            formula_pattern=pattern, formula_pattern_share=share or 0.0, kinds=kinds,
        ))
    return cols


_INFO_CACHE: dict[tuple, WorkbookInfo] = {}
_INFO_CACHE_MAX = 6
_INFO_LOCK = threading.Lock()   # render threads share the cache: lookups, eviction and insertion must not race


def load_info(path: str, recalc: bool = True) -> WorkbookInfo:
    """Cached per (path, mtime, recalc) within a process: probes and sweeps render several views of one file.

    Safe to share because every renderer only reads; the dims recorded in SheetInfo are taken at load
    time, so later ws.cell() accesses widening the openpyxl worksheet do not change what is reported.
    """
    key = (str(path), os.path.getmtime(path), recalc)
    with _INFO_LOCK:
        hit = _INFO_CACHE.get(key)
    if hit is not None:
        return hit
    info = _load_info_uncached(path, recalc)      # outside the lock so distinct files load in parallel
    with _INFO_LOCK:
        while len(_INFO_CACHE) >= _INFO_CACHE_MAX:
            _INFO_CACHE.pop(next(iter(_INFO_CACHE)), None)
        _INFO_CACHE[key] = info
    return info


def _load_info_uncached(path: str, recalc: bool = True) -> WorkbookInfo:
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v_raw = openpyxl.load_workbook(path, data_only=True)

    # Are cached values missing for formula cells? Then a recalculated copy is worth having.
    missing = 0
    total_formulas = 0
    for ws_f, ws_v in zip(wb_f.worksheets, wb_v_raw.worksheets):
        for row_f, row_v in zip(ws_f.iter_rows(max_row=min(ws_f.max_row, 500)),
                                ws_v.iter_rows(max_row=min(ws_v.max_row, 500))):
            for cf, cv in zip(row_f, row_v):
                if formula_text(cf.value) is not None:
                    total_formulas += 1
                    if cv.value is None:
                        missing += 1
    cached_missing = total_formulas > 0 and missing > 0

    values_path, recalc_ok = path, False
    wb_v = wb_v_raw
    if recalc and total_formulas > 0:
        values_path, recalc_ok = recalculated_copy(path)
        if recalc_ok:
            try:
                wb_v = openpyxl.load_workbook(values_path, data_only=True)
            except Exception:
                wb_v, values_path, recalc_ok = wb_v_raw, path, False

    sheets = []
    for ws_f in wb_f.worksheets:
        ws_v = wb_v[ws_f.title] if ws_f.title in wb_v.sheetnames else None
        header_row = guess_header_row(ws_f)
        columns = profile_sheet(ws_f, ws_v, header_row)
        sheets.append(SheetInfo(
            title=ws_f.title, max_row=ws_f.max_row, max_col=ws_f.max_column, header_row=header_row,
            merged=[str(m) for m in ws_f.merged_cells.ranges][:50], columns=columns,
            n_formula=sum(c.n_formula for c in columns), is_active=(ws_f is wb_f.active),
            scanned_rows=min(ws_f.max_row, PROFILE_ROWS),
        ))

    names = {}
    try:
        dn = wb_f.defined_names
        items = dn.items() if hasattr(dn, "items") else [(d.name, d) for d in dn.definedName]
        for name, d in items:
            names[name] = str(getattr(d, "attr_text", d))
    except Exception:
        pass

    return WorkbookInfo(
        path=path, values_path=values_path, recalc_ok=recalc_ok, cached_values_missing=cached_missing,
        sheets=sheets, defined_names=names, active_title=wb_f.active.title, wb_f=wb_f, wb_v=wb_v,
    )


def cell_text(info: WorkbookInfo, ws_title: str, row: int, col: int) -> str:
    ws_f = info.wb_f[ws_title]
    cell = ws_f.cell(row=row, column=col)
    cached = None
    if formula_text(cell.value) is not None and ws_title in info.wb_v.sheetnames:
        cached = info.wb_v[ws_title].cell(row=row, column=col).value
    return display_value(cell.value, cell.number_format or "General", cached)


def coord_to_rc(coord: str) -> tuple[int, int]:
    col, row = coordinate_from_string(coord)
    return row, column_index_from_string(col)
