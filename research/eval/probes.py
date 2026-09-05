"""Reading probes: can Qwen read a view at all, separately from solving the task?

    uv run eval/probes.py --digest tsv --digest grid --digest compact --model tinker --reasoning low \
        --ids-file eval/splits/dev100.txt --limit 100

For every init workbook, generate up to 20 questions whose answers are computed from the file itself
(value at an address, header of a column, which columns hold dates, non-empty count in a column,
formula text in a cell, sheet names, what a defined name refers to, last data row of a sheet,
formula-or-constant). Ask them all in one call per workbook per view, grade exactly, and report
accuracy by question type against prompt tokens. Writes private/probes/<digest>.json and appends a
row to experiments/probes.md. Only init workbooks are opened.
"""

from __future__ import annotations

import argparse
import asyncio
import datetime
import json
import random
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

RESEARCH = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RESEARCH))
sys.path.insert(0, str(RESEARCH.parent / "agent"))

from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.utils.cell import column_index_from_string  # noqa: E402

from baseline.common import load_env  # noqa: E402
from models import get_model  # noqa: E402
from sb import load_dataset, transform_value  # noqa: E402
from serialize import render  # noqa: E402
from workbook import cell_kind, formula_text, load_info  # noqa: E402

SYSTEM = ("You answer questions about a spreadsheet using ONLY the serialized workbook given to you. "
          "Answer each question with the bare value (no units, no explanation). If the information is not "
          "present in the text, answer \"unknown\". Reply with a single JSON object mapping question number to answer.")

QTYPES = ["cell_value", "cell_value_deep", "header", "date_columns", "nonempty_count", "formula_text",
          "sheet_names", "defined_name", "last_row", "is_formula"]


def _norm_text(s) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().casefold()


def _norm_formula(s) -> str:
    s = str(s).strip()
    s = s[1:] if s.startswith("=") else s
    return re.sub(r"\s+", "", s).casefold().replace("_xlfn.", "").replace("_xlws.", "")


def _as_list(actual) -> list[str]:
    if isinstance(actual, list):
        return [str(x) for x in actual]
    s = str(actual).strip()
    if s.startswith("["):
        try:
            return [str(x) for x in json.loads(s)]
        except json.JSONDecodeError:
            pass
    return [x.strip().strip("'\"") for x in re.split(r"[,;|]+", s.strip("[]{}")) if x.strip()]


def grade(qtype: str, expected, actual) -> bool:
    if actual is None:
        return False
    if qtype in ("date_columns", "sheet_names"):
        got = {_norm_text(x) for x in _as_list(actual) if str(x).strip()}
        exp = {_norm_text(x) for x in expected}
        return got == exp
    if isinstance(actual, (list, dict)):
        actual = json.dumps(actual)
    a = str(actual).strip()
    if qtype == "formula_text":
        return _norm_formula(a) == _norm_formula(expected)
    if qtype == "is_formula":
        return _norm_text(a).startswith(_norm_text(expected)[:5])
    if qtype in ("nonempty_count", "last_row"):
        try:
            return int(float(a.replace(",", ""))) == int(expected)
        except ValueError:
            return False
    # values, headers, defined names: scorer-style normalisation, then a text fallback for non-numeric strings only
    te, ta = transform_value(expected), transform_value(a)
    if type(te) == type(ta) and te == ta:
        return True
    if isinstance(expected, datetime.datetime):
        return _norm_text(expected.strftime("%Y-%m-%d")) in _norm_text(a) or _norm_text(a) in _norm_text(str(expected))
    if isinstance(expected, str) and isinstance(te, str):  # genuinely textual expected value
        ne, na = _norm_text(expected), _norm_text(a)
        return ne == na or (bool(ne) and ne in na and len(na) < len(ne) + 12)
    return False  # numeric expected: only the exact (2-dp) match above counts


def make_questions(task: dict, info, rng: random.Random, n: int = 20) -> list[dict]:
    qs: list[dict] = []
    sheets = info.sheets

    def add(qtype, text, expected, **extra):
        qs.append({"type": qtype, "q": text, "expected": expected, **extra})

    # sheet names
    add("sheet_names", "List all sheet names in the workbook, comma-separated.", [s.title for s in sheets])
    # defined names
    for name, ref in list(info.defined_names.items())[:2]:
        if "#REF" not in ref:
            add("defined_name", f"What cell range does the defined name '{name}' refer to?", ref)
    for s in sheets[:3]:
        ws_f = info.wb_f[s.title]
        ws_v = info.wb_v[s.title] if s.title in info.wb_v.sheetnames else ws_f
        # last row that actually holds a value (openpyxl's max_row also counts styled-but-empty rows)
        last_data_row = None
        if s.max_row <= 3000:
            for row in ws_f.iter_rows(min_row=1, max_row=s.max_row):
                if any(c.value not in (None, "") for c in row):
                    last_data_row = row[0].row
        if last_data_row:
            add("last_row", f"On sheet '{s.title}', what is the row number of the last row that contains any non-empty cell?", last_data_row)
        populated = [c for c in s.columns if c.n_nonempty > 0]
        if populated:
            col = rng.choice(populated)
            header_cell = ws_f.cell(row=s.header_row, column=column_index_from_string(col.letter)).value
            if header_cell not in (None, "") and formula_text(header_cell) is None:
                add("header", f"On sheet '{s.title}', what is the exact text in row {s.header_row} of column {col.letter}?", str(header_cell))
            if s.max_row <= 2000:
                add("nonempty_count", f"On sheet '{s.title}', how many non-empty cells are in column {col.letter} below row {s.header_row} (rows {s.header_row + 1} to {s.max_row})?", col.n_nonempty)
        date_cols = [c.letter for c in s.columns if c.dtype in ("date", "datetime")]
        if date_cols:
            add("date_columns", f"On sheet '{s.title}', which column letters hold dates? Comma-separated.", date_cols)
        # constant cells: shallow (header .. header+8) and deep (anywhere)
        const_shallow, const_deep, formulas = [], [], []
        for r in range(1, s.max_row + 1):
            for c in range(1, min(s.max_col, 40) + 1):
                cell = ws_f.cell(row=r, column=c)
                v = cell.value
                if v is None or v == "":
                    continue
                if formula_text(v):
                    formulas.append((r, c, formula_text(v)))
                elif cell_kind(v, cell.number_format or "General") in ("text", "number", "currency", "date"):
                    (const_shallow if r <= s.header_row + 8 else const_deep).append((r, c, v))
            if r > 3000:
                break
        for (r, c, v) in rng.sample(const_shallow, min(2, len(const_shallow))):
            add("cell_value", f"What is the value in cell {get_column_letter(c)}{r} on sheet '{s.title}'?", v)
        for (r, c, v) in rng.sample(const_deep, min(2, len(const_deep))):
            add("cell_value_deep", f"What is the value in cell {get_column_letter(c)}{r} on sheet '{s.title}'?", v)
        for (r, c, f) in rng.sample(formulas, min(2, len(formulas))):
            add("formula_text", f"What is the exact formula in cell {get_column_letter(c)}{r} on sheet '{s.title}'?", f.split(" {array")[0])
            add("is_formula", f"Is cell {get_column_letter(c)}{r} on sheet '{s.title}' a formula or a constant? Answer 'formula' or 'constant'.", "formula")
        if const_shallow and formulas:
            r, c, _ = const_shallow[0]
            add("is_formula", f"Is cell {get_column_letter(c)}{r} on sheet '{s.title}' a formula or a constant? Answer 'formula' or 'constant'.", "constant")
    qs = [q for q in qs if q["expected"] not in (None, "", [])]
    rng.shuffle(qs)
    return qs[:n]


def build_prompt(view_text: str, qs: list[dict]) -> str:
    numbered = "\n".join(f"{i + 1}. {q['q']}" for i, q in enumerate(qs))
    return (f"## Workbook\n{view_text}\n\n## Questions\n{numbered}\n\n"
            f"Reply with JSON only: {{\"1\": \"answer\", \"2\": \"answer\", ...}}")


def parse_answers(text: str) -> dict:
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.S)
    start, end = text.find("{"), text.rfind("}")
    if start < 0 or end < 0:
        return {}
    try:
        return json.loads(text[start:end + 1])
    except json.JSONDecodeError:
        return {}


async def probe_view(digest: str, tasks: list[dict], model, args, sem: asyncio.Semaphore) -> dict:
    results = []

    async def one(task):
        rng = random.Random(f"{task['id']}-probes")
        try:
            info = load_info(task["init_xlsx"], recalc=True)
            qs = make_questions(task, info, rng, args.questions)
            rendered = render(digest, task["init_xlsx"], task, args.budget)
        except Exception as e:
            return {"id": task["id"], "error": f"prep: {type(e).__name__}: {e}"[:200]}
        prompt = build_prompt(rendered.text, qs)
        async with sem:
            t0 = time.time()
            try:
                text, tok_in, tok_out = await model.complete(SYSTEM, prompt, effort=args.reasoning if args.reasoning != "adaptive" else "low")
            except Exception as e:
                return {"id": task["id"], "error": f"model: {type(e).__name__}: {e}"[:200], "tokens": rendered.meta.get("tokens")}
        answers = parse_answers(text)
        graded = []
        for i, q in enumerate(qs):
            a = answers.get(str(i + 1))
            graded.append({"type": q["type"], "ok": grade(q["type"], q["expected"], a), "q": q["q"],
                           "expected": q["expected"] if not isinstance(q["expected"], datetime.datetime) else q["expected"].isoformat(),
                           "actual": a})
        return {"id": task["id"], "n": len(graded), "correct": sum(g["ok"] for g in graded), "graded": graded,
                "tokens": rendered.meta.get("tokens"), "input_tokens": tok_in, "output_tokens": tok_out,
                "latency_s": round(time.time() - t0, 1), "parsed": bool(answers)}

    results = await asyncio.gather(*(one(t) for t in tasks))
    by_type = defaultdict(lambda: [0, 0])
    for r in results:
        for g in r.get("graded", []):
            by_type[g["type"]][1] += 1
            by_type[g["type"]][0] += int(g["ok"])
    n_q = sum(r.get("n", 0) for r in results)
    n_ok = sum(r.get("correct", 0) for r in results)
    toks = sorted(r["tokens"] for r in results if r.get("tokens"))
    summary = {
        "digest": digest, "model": model.name, "reasoning": args.reasoning, "workbooks": len(tasks),
        "errors": sum(1 for r in results if r.get("error")), "unparsed": sum(1 for r in results if r.get("parsed") is False),
        "questions": n_q, "accuracy": round(n_ok / n_q, 4) if n_q else None,
        "by_type": {k: round(v[0] / v[1], 3) for k, v in sorted(by_type.items())},
        "prompt_tokens_p50": toks[len(toks) // 2] if toks else None,
        "prompt_tokens_p90": toks[int(0.9 * len(toks))] if toks else None,
        "latency_p50_s": sorted(r["latency_s"] for r in results if "latency_s" in r)[len([r for r in results if "latency_s" in r]) // 2] if any("latency_s" in r for r in results) else None,
    }
    return {"summary": summary, "results": results}


def append_table(path: Path, s: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    header = ("| digest | when | model | reasoning | workbooks | accuracy % | " + " | ".join(QTYPES) +
              " | tok p50 | tok p90 | latency p50 s | errors |\n|---|---|---|---|---|---|" + "---|" * len(QTYPES) + "---|---|---|---|\n")
    if not path.exists():
        path.write_text("# Reading probes\n\nCan the model read the view? 20 auto-graded questions per workbook, one call per "
                        "workbook per view. Per-type columns are accuracy in %.\n\n" + header, encoding="utf-8")
    pct = lambda x: "" if x is None else f"{100 * x:.0f}"
    row = [s["digest"], datetime.datetime.now().strftime("%d %H:%M"), s["model"], s["reasoning"], str(s["workbooks"]),
           pct(s["accuracy"])] + [pct(s["by_type"].get(t)) for t in QTYPES] + [str(s["prompt_tokens_p50"]), str(s["prompt_tokens_p90"]),
           str(s["latency_p50_s"]), str(s["errors"])]
    with path.open("a", encoding="utf-8") as f:
        f.write("| " + " | ".join(row) + " |\n")


async def main_async(args):
    load_env()
    ids = [l.strip() for l in Path(args.ids_file).read_text().splitlines() if l.strip()] if args.ids_file else None
    tasks = load_dataset(args.dataset_dir)
    if ids:
        tasks = [t for t in tasks if t["id"] in set(ids)]
    tasks = tasks[: args.limit]
    model = get_model(args.model, project_id=args.project_id, reasoning=args.reasoning if args.reasoning != "adaptive" else "low",
                      max_tokens=args.max_tokens)
    sem = asyncio.Semaphore(args.concurrency)
    out_dir = RESEARCH / "private" / "probes"
    out_dir.mkdir(parents=True, exist_ok=True)
    for digest in args.digest:
        print(f"[{time.strftime('%H:%M:%S')}] probes {digest} on {len(tasks)} workbooks", flush=True)
        res = await probe_view(digest, tasks, model, args, sem)
        (out_dir / f"{digest}-{args.reasoning}.json").write_text(json.dumps(res, indent=1, default=str), encoding="utf-8")
        append_table(RESEARCH / "experiments" / "probes.md", res["summary"])
        print("   ", json.dumps(res["summary"]), flush=True)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--digest", action="append", required=True)
    p.add_argument("--ids-file", default=str(RESEARCH / "eval" / "splits" / "dev100.txt"))
    p.add_argument("--dataset-dir", default=str(RESEARCH / "data" / "spreadsheetbench_verified_400"))
    p.add_argument("--model", default="tinker")
    p.add_argument("--project-id", default=None)
    p.add_argument("--reasoning", default="low")
    p.add_argument("--max-tokens", type=int, default=4096)
    p.add_argument("--budget", type=int, default=None)
    p.add_argument("--questions", type=int, default=20)
    p.add_argument("--limit", type=int, default=100)
    p.add_argument("--concurrency", type=int, default=8)
    asyncio.run(main_async(p.parse_args()))


if __name__ == "__main__":
    main()
