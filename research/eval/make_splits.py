"""Fixed evaluation splits of the Verified 400.

    uv run eval/make_splits.py            # writes eval/splits/{dev100,check50,rest250}.txt + task_features.csv

dev-100  : iterate here (69 cell-level, 31 sheet-level).
check-50 : never tune on it; look at most three times over the weekend.
rest-250 : scored only inside the final 400 run.

Stratified by instruction type x graded-size bucket x multi-sheet, seed 0. Graded size is the
number of cells the scorer compares, computed with sb.answer_cells on the init workbook.
Only init workbooks are opened; nothing here touches a golden file.
"""

import argparse
import csv
import random
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from sb import DEFAULT_DATASET, answer_cells, load_dataset

HERE = Path(__file__).resolve().parent
SPLITS = HERE / "splits"

SIZE_BUCKETS = [(1, 1, "1"), (2, 10, "2-10"), (11, 50, "11-50"), (51, 200, "51-200"), (201, 10**9, ">200")]


def size_bucket(n: int) -> str:
    for lo, hi, name in SIZE_BUCKETS:
        if lo <= n <= hi:
            return name
    return ">200"


def task_features(task: dict) -> dict:
    wb = openpyxl.load_workbook(task["init_xlsx"], read_only=False, data_only=False)
    try:
        cells = answer_cells(task, wb)
    except Exception as e:  # malformed positions still get a row; graded size unknown
        cells, err = [], f"{type(e).__name__}: {e}"[:80]
    else:
        err = ""
    n_formula = 0
    n_nonempty = 0
    max_rows = 0
    max_cols = 0
    for ws in wb.worksheets:
        max_rows = max(max_rows, ws.max_row)
        max_cols = max(max_cols, ws.max_column)
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 2000)):
            for c in row:
                if c.value is None:
                    continue
                n_nonempty += 1
                if isinstance(c.value, str) and c.value.startswith("="):
                    n_formula += 1
    answer_sheets = {s for s, _ in cells}
    return {
        "id": task["id"],
        "type": "cell" if task["instruction_type"].startswith("Cell") else "sheet",
        "n_answer_cells": len(cells),
        "size_bucket": size_bucket(len(cells)),
        "n_sheets": len(wb.worksheets),
        "multi_sheet": int(len(wb.worksheets) > 1),
        "answer_sheet_given": int(bool(task.get("answer_sheet"))),
        "sheet_qualified": int("!" in task["answer_position"]),
        "n_ranges": task["answer_position"].count(",") + 1,
        "max_rows": max_rows,
        "max_cols": max_cols,
        "exceeds_120x30": int(max_rows > 120 or max_cols > 30),
        "n_formula_cells": n_formula,
        "n_nonempty_cells": n_nonempty,
        "instr_len": len(task["instruction"]),
        "answer_cells_error": err,
        "answer_sheets": "|".join(sorted(s or "(active)" for s in answer_sheets)),
    }


def stratified_sample(rows: list[dict], n: int, rng: random.Random, exclude: set[str]) -> list[str]:
    """Proportional allocation across strata, remainder to the largest strata, seeded."""
    pool = [r for r in rows if r["id"] not in exclude]
    strata: dict[tuple, list[dict]] = defaultdict(list)
    for r in pool:
        strata[(r["type"], r["size_bucket"], r["multi_sheet"])].append(r)
    total = len(pool)
    alloc = {k: (len(v) * n) // total for k, v in strata.items()}
    remainder = n - sum(alloc.values())
    for k in sorted(strata, key=lambda k: -len(strata[k]))[:remainder]:
        alloc[k] += 1
    chosen: list[str] = []
    for k, members in sorted(strata.items()):
        take = min(alloc[k], len(members))
        chosen.extend(r["id"] for r in rng.sample(members, take))
    return chosen


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--dataset-dir", default=str(DEFAULT_DATASET))
    p.add_argument("--seed", type=int, default=0)
    p.add_argument("--dev", type=int, default=100)
    p.add_argument("--check", type=int, default=50)
    args = p.parse_args()

    tasks = load_dataset(args.dataset_dir)
    rows = [task_features(t) for t in tasks]
    SPLITS.mkdir(parents=True, exist_ok=True)
    with (HERE / "task_features.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    rng = random.Random(args.seed)
    dev = stratified_sample(rows, args.dev, rng, exclude=set())
    check = stratified_sample(rows, args.check, rng, exclude=set(dev))
    rest = [r["id"] for r in rows if r["id"] not in set(dev) | set(check)]
    for name, ids in (("dev100", dev), ("check50", check), ("rest250", rest)):
        (SPLITS / f"{name}.txt").write_text("\n".join(ids) + "\n", encoding="utf-8")

    def summary(ids):
        sub = [r for r in rows if r["id"] in set(ids)]
        by_type = defaultdict(int)
        by_bucket = defaultdict(int)
        for r in sub:
            by_type[r["type"]] += 1
            by_bucket[r["size_bucket"]] += 1
        return f"n={len(sub)} type={dict(by_type)} size={dict(sorted(by_bucket.items()))}"

    print("dev100  ", summary(dev))
    print("check50 ", summary(check))
    print("rest250 ", summary(rest))
    print("all400  ", summary([r["id"] for r in rows]))
    bad = [r for r in rows if r["answer_cells_error"]]
    if bad:
        print(f"answer_cells failed on {len(bad)} tasks:", [(r["id"], r["answer_cells_error"]) for r in bad][:5])


if __name__ == "__main__":
    main()
