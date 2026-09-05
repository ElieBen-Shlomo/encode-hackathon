import datetime
import json
import os
import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

DEFAULT_DATASET = Path(__file__).resolve().parent / "data" / "spreadsheetbench_verified_400"


def load_dataset(dataset_dir=DEFAULT_DATASET):
    dataset_dir = Path(dataset_dir)
    tasks = json.loads((dataset_dir / "dataset.json").read_text())
    for t in tasks:
        t["id"] = str(t["id"])
        folder = dataset_dir / t["spreadsheet_path"]
        t["init_xlsx"] = str(next(folder.glob("*init*.xlsx")))
        golden = next(folder.glob("*golden*.xlsx"), None)
        t["golden_xlsx"] = str(golden) if golden else None
    return tasks


def parse_answer_position(answer_position):
    cleaned = answer_position.replace("'", "").replace('"', "")
    tokens = [cleaned] if cleaned.count("!") == 1 else cleaned.split(",")
    parsed = []
    for token in tokens:
        token = token.strip()
        if "!" in token:
            sheet, rng = token.rsplit("!", 1)
            parsed.append((sheet, _repair_range(rng)))
        else:
            parsed.append((None, _repair_range(token)))
    return parsed


def _repair_range(rng):
    if ":" not in rng:
        return rng
    start, end = rng.split(":", 1)
    if end.isdigit():
        col = "".join(ch for ch in start if ch.isalpha())
        return f"{start}:{col}{end}"
    return rng


def expand_range(cell_range, max_row=None, max_col=None):
    """Expand A1:B3 to cell coordinates. Whole-column (A:G) and whole-row (3:8) ranges
    need the sheet's max_row / max_col."""
    min_col, min_row, last_col, last_row = range_boundaries(cell_range)
    min_col, min_row = min_col or 1, min_row or 1
    last_row = last_row or max_row or min_row
    last_col = last_col or max_col or min_col
    return [f"{get_column_letter(c)}{r}" for r in range(min_row, last_row + 1) for c in range(min_col, last_col + 1)]


def answer_ranges(task):
    return [(sheet or task.get("answer_sheet"), rng) for sheet, rng in parse_answer_position(task["answer_position"])]


def split_sheet_field(field):
    """Some tasks pack several sheet names into one field: "Consolidated Tracker,Existing Task,..."."""
    if not field:
        return []
    return [part.strip().strip("'\"") for part in str(field).split(",") if part.strip().strip("'\"")]


def _resolve_sheets(sheet, wb):
    """Sheet field -> sheet titles in wb. An exact match wins (sheet names may themselves contain
    commas); otherwise a comma-joined field whose every part is a real sheet means all of them."""
    if wb is None or not sheet or sheet in wb.sheetnames:
        return [sheet]
    parts = split_sheet_field(sheet)
    if len(parts) > 1 and all(part in wb.sheetnames for part in parts):
        return parts
    return [sheet]


def answer_cells(task, wb=None):
    cells = []
    for sheet, rng in answer_ranges(task):
        for resolved in _resolve_sheets(sheet, wb):
            max_row = max_col = None
            if wb is not None:
                ws = wb[resolved] if resolved and resolved in wb.sheetnames else wb.active
                max_row, max_col = ws.max_row, ws.max_column
            cells.extend((resolved, coord) for coord in expand_range(rng, max_row, max_col))
    return cells


def load_answer_values(path, task):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sheet, coord in answer_cells(task, wb):
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        out[(ws.title, coord)] = ws[coord].value
    return out


def load_values_at(path, keys):
    """Read cells at the given (sheet title, coord) keys. Grading passes the golden's keys so
    both workbooks are read at the same cells even for open-ended ranges like A:G, whose
    expansion would otherwise follow each file's own max_row."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sheet, coord in keys:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        out[(sheet, coord)] = ws[coord].value
    return out


def _excel_serial(dt):
    delta = dt - datetime.datetime(1899, 12, 30)
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, datetime.datetime):
        return round(_excel_serial(v), 0)
    if isinstance(v, datetime.time):
        return str(v)[:-3]
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    return v


def values_equal(gold, pred):
    gold, pred = transform_value(gold), transform_value(pred)
    if gold in ("", None) and pred in ("", None):
        return True
    return type(gold) == type(pred) and gold == pred


SOFFICE_CANDIDATES = [
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/lib/libreoffice/program/soffice",
    "/opt/libreoffice/program/soffice",
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]


def soffice_path():
    if os.environ.get("SOFFICE"):
        return os.environ["SOFFICE"]
    for name in ("soffice", "libreoffice", "soffice.exe"):
        p = shutil.which(name)
        if p:
            return p
    for p in SOFFICE_CANDIDATES:
        if os.path.exists(p):
            return p
    return None


def recalculate(xlsx_path, out_dir):
    """Recalculate every formula in xlsx_path with LibreOffice headless, write the result to out_dir."""
    exe = soffice_path()
    if not exe:
        raise RuntimeError("LibreOffice not found. Install it, or set SOFFICE to the soffice executable, or pass --no-recalc")
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as profile:
        subprocess.run(
            [exe, f"-env:UserInstallation={Path(profile).resolve().as_uri()}", "--headless", "--calc",
             "--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", str(out_dir), str(xlsx_path)],
            check=True, capture_output=True, text=True, timeout=180,
        )
    return out_dir / Path(xlsx_path).name


def serialize_workbook(path, max_rows=120, max_cols=30):
    wb = openpyxl.load_workbook(path, data_only=True)
    parts = []
    for ws in wb.worksheets:
        rows, cols = min(ws.max_row, max_rows), min(ws.max_column, max_cols)
        header = "\t".join([""] + [get_column_letter(c) for c in range(1, cols + 1)])
        lines = [f"### Sheet: {ws.title} (showing {rows}x{cols} of {ws.max_row}x{ws.max_column})", header]
        for r in range(1, rows + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, cols + 1)]
            lines.append("\t".join([str(r)] + ["" if v is None else str(v) for v in vals]))
        parts.append("\n".join(lines))
    return "\n\n".join(parts)


def read_jsonl(path):
    with open(path, encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def resolve_output(output, predictions_path):
    if not output:
        return None
    p = Path(str(output).replace("\\", "/"))
    return p if p.is_absolute() else Path(predictions_path).resolve().parent / p
