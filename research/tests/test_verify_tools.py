"""agent/verify.py, agent/workbook_tools.py, agent/skills.py: the deterministic checks the agent relies on."""

import openpyxl
import pytest

import skills
import verify
import workbook_tools


def _edit(path, sheet, coord, value):
    wb = openpyxl.load_workbook(path)
    wb[sheet][coord] = value
    wb.save(path)


# ------------------------------------------------------------------------------------------ verify.py

def test_diff_reports_changed_cells_and_expected_ranges(tasks, init_copy):
    task = tasks["560-12"]  # answer_sheet '工作表1', I2:J7
    before, after = init_copy(task, "before.xlsx"), init_copy(task, "after.xlsx")
    _edit(after, "工作表1", "I2", "ITEM")
    _edit(after, "工作表1", "J2", 3)
    d = verify.diff_workbooks(str(before), str(after), task, expected_changes=["'工作表1'!I2:J7"])
    assert d["changed_count"] == 2 and set(d["changed_sample"]) == {"工作表1!I2", "工作表1!J2"}
    assert d["unexpected_count"] == 0
    d2 = verify.diff_workbooks(str(before), str(after), task, expected_changes=["'工作表1'!A1:A1"])
    assert d2["unexpected_count"] == 2
    d3 = verify.diff_workbooks(str(before), str(after), task, expected_changes=None)
    assert d3["changed_count"] == 2 and d3["unexpected_count"] == 0   # None means "do not police"


def test_diff_distinguishes_types_and_added_sheets(tasks, init_copy):
    task = tasks["560-12"]
    before, after = init_copy(task, "b.xlsx"), init_copy(task, "a.xlsx")
    _edit(after, "工作表1", "D2", "125")          # 125 -> "125": same text, different type, must count as a change
    wb = openpyxl.load_workbook(after); wb.create_sheet("Extra"); wb.save(after)
    d = verify.diff_workbooks(str(before), str(after), task, expected_changes=[])
    assert "工作表1!D2" in d["changed_sample"]
    assert "Extra!<sheet-added-or-removed>" in d["unexpected_sample"]


@pytest.mark.xfail(strict=True, reason="expected_changes without a sheet prefix are keyed under answer_sheet, which is null "
                                      "for cell-level tasks, so every edit is reported as unexpected; should fall back to the active sheet")
def test_unqualified_expected_changes_resolve_to_active_sheet(tasks, init_copy):
    task = tasks["12307"]  # answer_sheet None, graded I12:I13 on active Sheet1
    before, after = init_copy(task, "b.xlsx"), init_copy(task, "a.xlsx")
    _edit(after, "Sheet1", "I12", 2)
    d = verify.diff_workbooks(str(before), str(after), task, expected_changes=["I12:I13"])
    assert d["unexpected_count"] == 0


def test_formula_cells_found_on_named_answer_sheet(tasks, init_copy):
    task = tasks["560-12"]
    p = init_copy(task)
    assert verify.formula_cells(str(p), task) == []
    _edit(p, "工作表1", "I2", "=COUNTA(A2:A10)")
    assert verify.formula_cells(str(p), task) == ["工作表1!I2"]


@pytest.mark.xfail(strict=True, reason="answer_sheet is null for cell-level tasks, so formula_cells skips the graded range "
                                      "and auto-recalculation never runs for 275 of the 400 tasks")
def test_formula_cells_found_when_answer_sheet_is_null(tasks, init_copy):
    task = tasks["12307"]
    p = init_copy(task)
    _edit(p, "Sheet1", "I12", "=COUNTIF(B12:H12,\">0\")")
    assert verify.formula_cells(str(p), task) == ["Sheet1!I12"]


def test_format_verification_text():
    text = verify.format_verification({"changed_count": 2, "changed_sample": ["S!A1", "S!A2"],
                                       "unexpected_count": 1, "unexpected_sample": ["S!A2"]}, ["S!A1"])
    assert "Changed cells: 2; sample: S!A1, S!A2" in text
    assert "Unexpected changed cells: 1; sample: S!A2" in text
    assert "Formula cells in graded range: 1; sample: S!A1" in text


# ---------------------------------------------------------------------------------- workbook_tools.py

def test_inspect_range_shows_coordinates_and_types(tasks):
    ok, text = workbook_tools.inspect_range(tasks["12307"]["init_xlsx"], "Sheet1", "A11:C12")
    assert ok
    assert "A11='Company' [str]" in text and "A12='ABC' [str]" in text and "B12=2 [int]" in text
    ok, text = workbook_tools.inspect_range(tasks["12307"]["init_xlsx"], "Nope", "A1:A2")
    assert not ok and "Sheet not found" in text


def test_inspect_range_with_styles_and_limit(tasks):
    ok, text = workbook_tools.inspect_range(tasks["12307"]["init_xlsx"], "Sheet1", "A1:I13", styles=True, limit=5)
    assert ok and "format=" in text and "... cells omitted ..." in text


def test_assert_blank(tasks, init_copy):
    task = tasks["12307"]
    p = init_copy(task)
    ok, msg = workbook_tools.assert_blank(str(p), "Sheet1", "I12:I13")
    assert ok and msg == "Range is blank"
    _edit(p, "Sheet1", "I12", 2)
    ok, msg = workbook_tools.assert_blank(str(p), "Sheet1", "I12:I13")
    assert not ok and "I12" in msg


def test_assert_sorted(tmp_path):
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    for i, (a, b) in enumerate([("a", 1), ("b", 2), ("c", 3)], start=2):
        ws[f"A{i}"], ws[f"B{i}"] = a, b
    wb.save(p)
    assert workbook_tools.assert_sorted(str(p), "S", "A2:B4", ["A"]) == (True, "Range is sorted ascending")
    ws["A3"] = "z"; wb.save(p)
    ok, msg = workbook_tools.assert_sorted(str(p), "S", "A2:B4", ["A"])
    assert not ok and "not sorted" in msg
    ok, msg = workbook_tools.assert_sorted(str(p), "Missing", "A2:B4", ["A"])
    assert not ok


# ------------------------------------------------------------------------------------------ skills.py

@pytest.mark.parametrize("instruction,expected", [
    ("Please sort the table ascending by name", "Sorting checklist"),
    ("Filter rows between the start date and end date", "Filtering checklist"),
    ("Give me the total per region", "Aggregation checklist"),
    ("Delete duplicates and shift cells up", "Deletion/compaction checklist"),
    ("Mark the row when the value is highlighted", "Conditional-edit checklist"),
])
def test_selected_skills_keywords(instruction, expected):
    assert expected in skills.selected_skills(instruction)


def test_no_skill_for_unrelated_instruction():
    assert skills.selected_skills("Convert the header to upper case") == ""
