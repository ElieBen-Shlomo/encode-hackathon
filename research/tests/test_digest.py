"""agent/digest.py: the prompt view of the workbook and the verification snapshot."""

import openpyxl
import pytest

import digest
from conftest import EDGE_IDS


@pytest.mark.parametrize("tid", EDGE_IDS)
def test_digest_renders_every_edge_case(tasks, tid):
    text = digest.digest(tasks[tid]["init_xlsx"], tasks[tid])
    assert text.strip()
    wb = openpyxl.load_workbook(tasks[tid]["init_xlsx"])   # same loader as the digest (read_only reports different dims)
    for ws in wb.worksheets:
        assert f"### Sheet: {ws.title!r}" in text     # every sheet is announced with its true dimensions
        assert f"{ws.max_row} rows x {ws.max_column} cols" in text


def test_digest_windows_big_sheets_and_says_so(tasks):
    text = digest.digest(tasks["17-35"]["init_xlsx"], tasks["17-35"])   # 315 rows, answer I6:M295
    assert "omitted" in text
    assert "\n1\t" in text and "\n6\t" in text       # head rows and the first graded row are shown
    shown = sum(1 for line in text.splitlines() if line.split("\t")[0].isdigit())
    assert shown <= digest.MAX_SHOWN


def test_digest_is_values_only(tasks):
    """The prompt view never shows formulas, only cached values; a formula without a cache shows empty."""
    text = digest.digest(tasks["17-35"]["init_xlsx"], tasks["17-35"])
    assert "=SUM" not in text and "_xlfn" not in text


def test_verification_snapshot_lists_graded_cells_with_types(tasks):
    task = tasks["560-12"]  # answer_sheet given: '工作表1'!I2:J7, all empty in the init
    text = digest.verification_snapshot(task["init_xlsx"], task)
    assert "## Graded answer cells" in text
    assert "I2=<empty> [NoneType]" in text and "J7=<empty> [NoneType]" in text


def test_verification_snapshot_flags_missing_sheet(tasks):
    task = dict(tasks["560-12"], answer_position="'Nope'!A1:A2", answer_sheet="Nope")
    text = digest.verification_snapshot(task["init_xlsx"], task)
    assert "SHEET MISSING" in text


@pytest.mark.xfail(strict=True, reason="answer_sheet is null for all 275 cell-level tasks; the snapshot reports "
                                      "'None!..: SHEET MISSING' instead of falling back to the active sheet like the grader")
def test_verification_snapshot_falls_back_to_active_sheet_when_answer_sheet_is_null(tasks):
    task = tasks["12307"]  # I12:I13 on the active sheet, answer_sheet None
    text = digest.verification_snapshot(task["init_xlsx"], task)
    assert "SHEET MISSING" not in text
    assert "I12=" in text


def test_verification_snapshot_caps_large_ranges(tasks):
    """17-35 grades 1450 cells and declares A5:N315 as source: both sections are sampled to their limits."""
    text = digest.verification_snapshot(tasks["17-35"]["init_xlsx"], tasks["17-35"], answer_limit=20, source_limit=10)
    answer_part, source_part = text.split("## Relevant declared source cells")
    cells = lambda part: sum(1 for l in part.splitlines() if "=" in l and "[" in l)
    assert "... cells omitted ..." in answer_part and cells(answer_part) == 20
    assert "... cells omitted ..." in source_part and cells(source_part) == 10
