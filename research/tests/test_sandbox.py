"""agent/sandbox.py: model-written code runs in a subprocess inside a per-task workspace.

Two strict expected-failures document safety properties the current sandbox does not provide:
the child receives the real input workbook path (so it can modify the dataset), and it inherits
the full environment including API keys.
"""

import hashlib
from pathlib import Path

import openpyxl
import pytest

import sandbox


def _sha(p):
    return hashlib.sha256(Path(p).read_bytes()).hexdigest()


def _seed(task, init_copy, tmp_path):
    src = init_copy(task)
    out = tmp_path / "out.xlsx"
    out.write_bytes(src.read_bytes())
    work = tmp_path / "work"
    work.mkdir()
    return src, out, work


def test_run_python_edits_output_and_reports_success(tasks, init_copy, tmp_path):
    src, out, work = _seed(tasks["12307"], init_copy, tmp_path)
    code = "import sys, openpyxl\nwb = openpyxl.load_workbook(sys.argv[2]); wb.active['I12'] = 7; wb.save(sys.argv[2]); print('done')"
    ok, log = sandbox.run_python(code, work_dir=work, in_xlsx=str(src), out_xlsx=str(out), turn=1, timeout=60)
    assert ok and "done" in log
    assert openpyxl.load_workbook(out).active["I12"].value == 7
    assert (work / "turn_01.py").exists()                       # scripts are kept in the workspace


def test_env_carries_absolute_in_and_out_paths(tasks, init_copy, tmp_path):
    src, out, work = _seed(tasks["12307"], init_copy, tmp_path)
    code = "import os, pathlib\nprint(os.environ['IN_XLSX'], os.environ['OUT_XLSX'])\nassert pathlib.Path(os.environ['OUT_XLSX']).is_absolute()"
    ok, log = sandbox.run_python(code, work_dir=work, in_xlsx=str(src), out_xlsx=str(out), turn=2, timeout=60)
    assert ok and str(out.resolve()) in log


def test_timeout_is_enforced(tasks, init_copy, tmp_path):
    src, out, work = _seed(tasks["12307"], init_copy, tmp_path)
    ok, log = sandbox.run_python("import time; time.sleep(30)", work_dir=work, in_xlsx=str(src), out_xlsx=str(out), turn=1, timeout=2)
    assert ok is False and "TIMEOUT" in log


def test_nonzero_exit_is_a_failure_with_the_traceback(tasks, init_copy, tmp_path):
    src, out, work = _seed(tasks["12307"], init_copy, tmp_path)
    ok, log = sandbox.run_python("raise ValueError('boom')", work_dir=work, in_xlsx=str(src), out_xlsx=str(out), turn=1, timeout=30)
    assert ok is False and "ValueError: boom" in log


def test_deleting_output_is_a_failure(tasks, init_copy, tmp_path):
    src, out, work = _seed(tasks["12307"], init_copy, tmp_path)
    ok, log = sandbox.run_python("import os, sys; os.remove(sys.argv[2])", work_dir=work, in_xlsx=str(src), out_xlsx=str(out), turn=1, timeout=30)
    assert ok is False and "deleted OUT_XLSX" in log


def test_run_bash_runs_in_the_workspace(tasks, init_copy, tmp_path):
    src, out, work = _seed(tasks["12307"], init_copy, tmp_path)
    ok, log = sandbox.run_bash("pwd; test -f \"$OUT_XLSX\" && echo present", work_dir=work, in_xlsx=str(src), out_xlsx=str(out), timeout=30)
    assert ok and "present" in log and str(work.resolve()) in log


@pytest.mark.xfail(strict=True, reason="the child gets the real input path as argv[1]/IN_XLSX; model code can modify the dataset. "
                                      "Fix: stage a copy of the init into the workspace and pass that (see PR #11)")
def test_script_cannot_modify_the_input_workbook(tasks, init_copy, tmp_path):
    src, out, work = _seed(tasks["12307"], init_copy, tmp_path)
    before = _sha(src)
    code = "import sys, openpyxl\nwb = openpyxl.load_workbook(sys.argv[1]); wb.active['A1'] = 'VANDALISED'; wb.save(sys.argv[1])"
    ok, _ = sandbox.run_python(code, work_dir=work, in_xlsx=str(src), out_xlsx=str(out), turn=1, timeout=60)
    assert ok
    assert _sha(src) == before, "model code modified the input workbook"


@pytest.mark.xfail(strict=True, reason="_env copies os.environ wholesale, so TINKER_API_KEY and friends reach model-written code. "
                                      "Fix: drop *API_KEY*/*TOKEN* variables from the child env (see PR #11)")
def test_api_keys_do_not_reach_model_code(tasks, init_copy, tmp_path, monkeypatch):
    monkeypatch.setenv("TINKER_API_KEY", "secret-should-not-leak")
    src, out, work = _seed(tasks["12307"], init_copy, tmp_path)
    ok, log = sandbox.run_python("import os; print('KEY=' + os.environ.get('TINKER_API_KEY', 'absent'))",
                                 work_dir=work, in_xlsx=str(src), out_xlsx=str(out), turn=1, timeout=30)
    assert ok and "KEY=absent" in log
