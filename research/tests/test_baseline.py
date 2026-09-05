"""research/baseline: prompt/answer plumbing shared by the baselines, and the Tinker script's module-level config."""

import datetime
import os
from pathlib import Path

import openpyxl
import pytest

import common
import tinker_predict


# ------------------------------------------------------------------------------------------ common.py

def test_parse_answer_strips_thinking_and_fences():
    text = '<think>counting rows</think>Here you go:\n```json\n{"cells": [{"cell": "b6", "value": 42}, {"cell": "B7", "value": null}]}\n```'
    ans = common.parse_answer(text)
    assert [(c.cell, c.value) for c in ans.cells] == [("b6", 42), ("B7", None)]


@pytest.mark.parametrize("bad", ["no json here", "<think>only thinking</think>", '{"cells": [{"cell": 5}]}'])
def test_parse_answer_rejects_garbage(bad):
    with pytest.raises(Exception):
        common.parse_answer(bad)


def test_write_output_writes_only_graded_cells_with_python_types(tasks, tmp_path):
    task = tasks["12307"]  # graded I12:I13 on Sheet1
    ans = common.parse_answer('{"cells": [{"cell": "i12", "value": 2}, {"cell": "I13", "value": "2"}, {"cell": "A1", "value": "ignored"}]}')
    out = tmp_path / "o.xlsx"
    common.write_output(task, ans, out)
    ws = openpyxl.load_workbook(out).active
    assert ws["I12"].value == 2 and isinstance(ws["I12"].value, int)
    assert ws["I13"].value == "2"                      # strings are written verbatim (the grader coerces numeric strings)
    assert ws["A1"].value == "Country"                 # cells outside the graded range are untouched


def test_write_output_dates_as_strings_stay_text(tasks, tmp_path):
    """Documents a limitation of the values strategy: a date returned as text is stored as text and can never grade equal."""
    task = tasks["12307"]
    ans = common.parse_answer('{"cells": [{"cell": "I12", "value": "2024-01-31"}]}')
    out = tmp_path / "o.xlsx"
    common.write_output(task, ans, out)
    v = openpyxl.load_workbook(out).active["I12"].value
    assert isinstance(v, str) and not isinstance(v, datetime.datetime)


def test_build_prompt_has_the_three_sections(tasks):
    prompt = common.build_prompt(tasks["12307"])
    assert prompt.startswith("## Instruction\n") and "## Workbook\n" in prompt and "## Answer range\nSheet: active sheet\nCells: I12:I13" in prompt


def test_load_env_sets_but_does_not_override(tmp_path, monkeypatch):
    env = tmp_path / ".env"
    env.write_text('A_TEST_KEY="value1"\n# comment\nB_TEST_KEY=value2\n')
    monkeypatch.delenv("A_TEST_KEY", raising=False)
    monkeypatch.setenv("B_TEST_KEY", "already")
    common.load_env(env)
    assert os.environ["A_TEST_KEY"] == "value1" and os.environ["B_TEST_KEY"] == "already"


# ------------------------------------------------------------------------------------ tinker_predict.py

def test_default_renderer_is_medium_for_qwen38_and_recommended_otherwise():
    assert tinker_predict.default_renderer("Qwen/Qwen3.8-27B") == "qwen3_8_medium_reasoning"
    assert tinker_predict.default_renderer("Qwen/Qwen3-8B") == tinker_predict.get_recommended_renderer_name("Qwen/Qwen3-8B")


def test_ladder_steps_from_most_to_least_reasoning():
    assert tinker_predict.QWEN38_LADDER == ["qwen3_8_xhigh_reasoning", "qwen3_8_medium_reasoning",
                                            "qwen3_8_low_reasoning", "qwen3_8_disable_thinking"]


def test_load_config_reads_yaml_and_rejects_missing_or_non_mapping(tmp_path):
    cfg = tinker_predict.load_config(Path(__file__).resolve().parents[1] / "config" / "qwen.yaml")
    assert cfg["base_model"] == "Qwen/Qwen3.8-27B" and cfg["temperature"] == 0
    assert cfg["renderer"] in tinker_predict.QWEN38_LADDER
    with pytest.raises(FileNotFoundError):
        tinker_predict.load_config(tmp_path / "missing.yaml")
    bad = tmp_path / "list.yaml"
    bad.write_text("- a\n- b\n")
    with pytest.raises(ValueError):
        tinker_predict.load_config(bad)
