"""answer_position parsing on the awkward real cases. All via sb.answer_cells, never hand-parsed."""

import openpyxl
import pytest

import sb


@pytest.mark.parametrize("tid,n_ranges,sheets", [
    ("41-47", 4, {"OUT CAS"}),                      # four ranges, all sheet-qualified
    ("283-32", 2, {"Sheet3", "Sheet4"}),           # whole columns A:G, odd quoting
    ("516-46", 1, {"ورقة1"}),                      # Arabic sheet name
    ("560-12", 1, {"工作表1"}),                     # Chinese sheet name
    ("12307", 1, {None}),                          # unqualified: graded on the active sheet
])
def test_answer_ranges_parse(tasks, tid, n_ranges, sheets):
    ranges = sb.answer_ranges(tasks[tid])
    assert len(ranges) == n_ranges
    assert {s for s, _ in ranges} == sheets


def test_whole_column_expands_to_sheet_height(tasks):
    wb = openpyxl.load_workbook(tasks["283-32"]["init_xlsx"])
    cells = sb.answer_cells(tasks["283-32"], wb)
    assert len(cells) == 14  # both sheets are one row tall in the init, 7 columns each
    assert {s for s, _ in cells} == {"Sheet3", "Sheet4"}


def test_multi_range_cell_count(tasks):
    wb = openpyxl.load_workbook(tasks["41-47"]["init_xlsx"])
    cells = sb.answer_cells(tasks["41-47"], wb)
    assert len(cells) == 3 * 1528 + 3 * 585 + 3 * 12 + 4 * 7


def test_nbsp_prefixed_position_resolves_to_active_sheet(tasks):
    """49300's position starts with a non-breaking space; the grader falls back to the active sheet."""
    task = tasks["49300"]
    assert task["answer_position"].startswith("\xa0")
    wb = openpyxl.load_workbook(task["init_xlsx"])
    cells = sb.answer_cells(task, wb)
    assert [c for _, c in cells] == ["C2", "C3"]
    sheet = cells[0][0]
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    assert ws.title == "Sheet1"


def test_all_400_positions_parse(tasks):
    """Every task's answer_position must expand without error against its init workbook."""
    bad = []
    for tid, task in tasks.items():
        try:
            wb = openpyxl.load_workbook(task["init_xlsx"], read_only=False)
            assert len(sb.answer_cells(task, wb)) > 0
        except Exception as e:
            bad.append((tid, f"{type(e).__name__}: {e}"))
    assert not bad, bad


def test_repair_range_and_expand():
    assert sb._repair_range("A1:3") == "A1:A3"
    assert sb._repair_range("B2:D9") == "B2:D9"
    assert sb.expand_range("A1:B2") == ["A1", "B1", "A2", "B2"]
