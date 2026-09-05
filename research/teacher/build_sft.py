"""Render verified teacher scripts into harness-format SFT trajectories for the Qwen fine-tune.

    uv run teacher/build_sft.py            # -> teacher/sft/train.jsonl

One JSON line per verified train-split script: {"id", "messages": [...]} replaying the exact
conversation the agent harness produces at inference: system prompt, task message, a
run_python action carrying the script, the real tool result (the script is executed again and
the updated digest appended, like harness.solve_task does), an optional recalculate_workbook
turn when the script wrote formulas whose values need LibreOffice, then finish.

Test-split ids are refused outright (DATA_PLAN: the test split is sealed).
"""

import json
import shutil
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent))                    # research: sb
sys.path.insert(0, str(HERE.parent.parent / "agent"))   # agent: harness, digest, sandbox

import openpyxl
from digest import digest
from harness import build_messages, tool_result
from sandbox import run_python
from sb import load_answer_values, load_dataset, recalculate

SCRIPTS = HERE / "scripts"
MANIFEST = HERE / "manifest.jsonl"
SFT = HERE / "sft" / "train.jsonl"
SPLITS = HERE.parent.parent / "datasets" / "splits"


def passed_ids() -> list[str]:
    if not MANIFEST.exists():
        sys.exit("no manifest.jsonl — run `uv run teacher/verify.py --all` first")
    last = {}
    for line in MANIFEST.read_text().splitlines():
        if line.strip():
            record = json.loads(line)
            last[record["id"]] = record
    return sorted(task_id for task_id, record in last.items() if record.get("passed"))


def needs_recalc(out_xlsx: Path, task: dict) -> bool:
    """True when a graded cell is empty pre-recalculation but golden is not: the script wrote
    formulas, so the faithful trajectory includes a recalculate_workbook step."""
    gold = load_answer_values(task["golden_xlsx"], task)
    wb = openpyxl.load_workbook(out_xlsx, data_only=True)
    for (sheet, coord), value in gold.items():
        if value in ("", None):
            continue
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        if ws[coord].value is None:
            return True
    return False


def trajectory(task: dict, script: str) -> list[dict] | None:
    with tempfile.TemporaryDirectory(prefix=f"sft-{task['id']}-") as temp:
        work = Path(temp)
        in_copy = work / Path(task["init_xlsx"]).name
        shutil.copy(task["init_xlsx"], in_copy)
        in_copy.chmod(0o444)
        out = work / "out.xlsx"
        shutil.copy(task["init_xlsx"], out)

        ok, result = run_python(script, work_dir=work, in_xlsx=str(in_copy),
                                out_xlsx=str(out), turn=1, timeout=120)
        if not ok:
            return None  # verified earlier; a replay failure means environment drift — skip loudly
        result += "\n\n## Updated workbook digest\n" + digest(str(out), task)

        messages = build_messages(task)
        messages.append({"role": "assistant",
                         "content": json.dumps({"tool": "run_python", "args": {"code": script}}, ensure_ascii=False)})
        tool_result(messages, "run_python", result)

        if needs_recalc(out, task):
            messages.append({"role": "assistant",
                             "content": json.dumps({"tool": "recalculate_workbook", "args": {}})})
            recalced = recalculate(str(out), work / "recalc")
            shutil.copy(recalced, out)
            tool_result(messages, "recalculate_workbook",
                        "LibreOffice recalculation completed\n\n## Updated workbook digest\n" + digest(str(out), task))

        messages.append({"role": "assistant",
                         "content": json.dumps({"tool": "finish",
                                                "args": {"summary": "instruction applied and OUT_XLSX saved"}})})
    return messages


def main() -> None:
    test_ids = set((SPLITS / "test.txt").read_text().split())
    ids = passed_ids()
    leaked = sorted(set(ids) & test_ids)
    if leaked:
        sys.exit(f"REFUSING: manifest contains passing TEST-split ids {leaked} — the test split is sealed (DATA_PLAN)")

    tasks_by_id = {t["id"]: t for t in load_dataset()}
    SFT.parent.mkdir(parents=True, exist_ok=True)
    written, skipped, chars = 0, [], 0
    with SFT.open("w", encoding="utf-8") as f:
        for task_id in ids:
            task = tasks_by_id.get(task_id)
            script = (SCRIPTS / f"{task_id}.py").read_text(encoding="utf-8")
            messages = trajectory(task, script) if task else None
            if messages is None:
                skipped.append(task_id)
                continue
            line = json.dumps({"id": task_id, "messages": messages}, ensure_ascii=False)
            f.write(line + "\n")
            written += 1
            chars += len(line)
    print(f"{written} trajectories -> {SFT.relative_to(HERE.parent)}  (~{chars // 4:,} tokens)")
    if skipped:
        print(f"skipped (replay failed): {skipped}")


if __name__ == "__main__":
    main()
