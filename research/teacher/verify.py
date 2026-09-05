"""Verify a teacher script: run it in the sandbox, LibreOffice-recalculate the output,
grade against the golden with the shipped scorer semantics (sb.load_answer_values for the
golden cells, sb.values_equal per cell; the prediction is read at exactly the golden's cells).

    uv run teacher/verify.py 13-1                  # one id -> verdict JSON, exit 0 pass / 1 fail
    uv run teacher/verify.py 13-1 --script alt.py  # verify an alternative file for that id
    uv run teacher/verify.py --all                 # sweep scripts/, rewrite manifest.jsonl

A script passes only if every graded cell matches the golden after recalculation — the same
bar evaluate.py applies to pipeline outputs (DATA_PLAN: never trust an unverified label).
hardcode_hits is advisory: distinctive golden strings found verbatim in the script; they are
only acceptable when the instruction text itself contains those values.
"""

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent           # research/teacher
sys.path.insert(0, str(HERE.parent))             # research: sb
sys.path.insert(0, str(HERE.parent.parent / "agent"))  # agent: sandbox

import openpyxl
from sandbox import run_python
from sb import DEFAULT_DATASET, load_answer_values, load_dataset, recalculate, values_equal

SCRIPTS = HERE / "scripts"
MANIFEST = HERE / "manifest.jsonl"
SPLITS = HERE.parent.parent / "datasets" / "splits"


def read_at(path, keys):
    """Read cells at the golden's (sheet title, coord) keys so both files are graded at the
    same cells even for open-ended ranges like A:G."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sheet, coord in keys:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        out[(sheet, coord)] = ws[coord].value
    return out


def hardcode_hits(script_text: str, gold: dict) -> list[str]:
    distinctive = {str(v) for v in gold.values() if isinstance(v, str) and len(str(v)) >= 4}
    return sorted(v for v in distinctive if v in script_text)[:10]


def verify(task: dict, script_path: Path, timeout: int) -> dict:
    script = script_path.read_text(encoding="utf-8")
    verdict = {"id": task["id"], "passed": False, "run_ok": False}
    with tempfile.TemporaryDirectory(prefix=f"teacher-{task['id']}-") as temp:
        work = Path(temp)
        in_copy = work / Path(task["init_xlsx"]).name
        shutil.copy(task["init_xlsx"], in_copy)
        in_copy.chmod(0o444)  # the script must never write the input
        out = work / "out.xlsx"
        shutil.copy(task["init_xlsx"], out)

        ok, output = run_python(script, work_dir=work, in_xlsx=str(in_copy),
                                out_xlsx=str(out), turn=1, timeout=timeout)
        verdict["run_ok"] = ok
        verdict["output"] = output[-2000:]
        if not ok:
            return verdict
        try:
            recalced = recalculate(str(out), work / "recalc")
            gold = load_answer_values(task["golden_xlsx"], task)
            pred = read_at(recalced, gold.keys())
        except Exception as exc:
            verdict["error"] = f"{type(exc).__name__}: {exc}"[:300]
            return verdict
        mismatches = [{"cell": f"{sheet}!{coord}", "expected": g, "actual": pred.get((sheet, coord))}
                      for (sheet, coord), g in gold.items() if not values_equal(g, pred.get((sheet, coord)))]
        verdict.update(cells=len(gold), correct=len(gold) - len(mismatches), passed=not mismatches,
                       mismatches=mismatches[:8], hardcode_hits=hardcode_hits(script, gold))
    return verdict


def split_of(task_id: str) -> str:
    for name in ("train", "test"):
        if task_id in (SPLITS / f"{name}.txt").read_text().split():
            return name
    return "unknown"


def sweep(tasks_by_id: dict, timeout: int) -> None:
    lines, passed = [], 0
    for script_path in sorted(SCRIPTS.glob("*.py")):
        task = tasks_by_id.get(script_path.stem)
        if task is None:
            print(f"{script_path.stem:<8} not in dataset, skipped", file=sys.stderr)
            continue
        verdict = verify(task, script_path, timeout)
        record = {"id": task["id"], "split": split_of(task["id"]), "type": task["instruction_type"],
                  "passed": verdict["passed"], "run_ok": verdict["run_ok"],
                  "cells": verdict.get("cells"), "correct": verdict.get("correct"),
                  "hardcode_hits": len(verdict.get("hardcode_hits", [])),
                  "error": verdict.get("error")}
        lines.append(record)
        passed += verdict["passed"]
        detail = "" if verdict["passed"] else (verdict.get("error") or
                 (verdict.get("mismatches") or [{}])[0] if verdict.get("cells") else verdict["output"][-120:])
        print(f"{task['id']:<8} {'PASS' if verdict['passed'] else 'FAIL'} "
              f"{verdict.get('correct', 0)}/{verdict.get('cells', '?')}  {str(detail)[:100]}")
    MANIFEST.write_text("".join(json.dumps(r, ensure_ascii=False, default=str) + "\n" for r in lines))
    n_test = sum(r["split"] == "test" for r in lines)
    print(f"\n{passed}/{len(lines)} scripts pass -> {MANIFEST.relative_to(HERE.parent)}")
    if n_test:
        print(f"WARNING: {n_test} scripts are for TEST-split ids; they must never enter SFT data (DATA_PLAN).")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("task_id", nargs="?", help="task id; omit with --all")
    p.add_argument("--script", help="script path (default: teacher/scripts/<id>.py)")
    p.add_argument("--all", action="store_true", help="verify every script in scripts/, rewrite manifest.jsonl")
    p.add_argument("--dataset-dir", default=str(DEFAULT_DATASET))
    p.add_argument("--timeout", type=int, default=120, help="seconds for the script to run")
    args = p.parse_args()

    tasks_by_id = {t["id"]: t for t in load_dataset(args.dataset_dir)}
    if args.all:
        sweep(tasks_by_id, args.timeout)
        return
    if not args.task_id:
        p.error("give a task id or --all")
    task = tasks_by_id.get(args.task_id)
    if task is None:
        sys.exit(f"no task {args.task_id!r}")
    script_path = Path(args.script) if args.script else SCRIPTS / f"{args.task_id}.py"
    if not script_path.exists():
        sys.exit(f"no script {script_path}")
    verdict = verify(task, script_path, args.timeout)
    print(json.dumps(verdict, indent=2, ensure_ascii=False, default=str))
    sys.exit(0 if verdict["passed"] else 1)


if __name__ == "__main__":
    main()
