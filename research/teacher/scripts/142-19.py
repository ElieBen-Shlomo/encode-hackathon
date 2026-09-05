import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

def blank(r):
    return all(ws.cell(row=r, column=c).value in (None, "") for c in (1, 2))

# Split the sheet into ranges separated by blank rows.
blocks = []
current = []
for r in range(1, ws.max_row + 1):
    if blank(r):
        if current:
            blocks.append(current)
            current = []
    else:
        current.append(r)
if current:
    blocks.append(current)

# Within each range the first row is a header; sort the remaining rows A->Z by
# column A as text, carrying the paired name in column B along with it.
for rows in blocks:
    header, data = rows[0], rows[1:]
    pairs = [(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value) for r in data]
    pairs.sort(key=lambda p: str(p[0]))
    for r, (a, b) in zip(data, pairs):
        ws.cell(row=r, column=1).value = a
        ws.cell(row=r, column=2).value = b

wb.save(OUT)
