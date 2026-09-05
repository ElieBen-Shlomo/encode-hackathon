import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
pl = wb["PL Recon Items"]
st = wb["Statement Recon Items"]


def key(inv, amt):
    inv = str(inv).strip() if inv is not None else ""
    amt = round(float(amt), 2) if isinstance(amt, (int, float)) else str(amt)
    return (inv, amt)


# (invoice, amount) pairs: PL uses columns C and D, Statement uses F and I.
pl_pairs = {key(pl.cell(r, 3).value, pl.cell(r, 4).value) for r in range(2, pl.max_row + 1)}
st_pairs = {key(st.cell(r, 6).value, st.cell(r, 9).value) for r in range(2, st.max_row + 1)}
matched = pl_pairs & st_pairs

# Delete matching rows from both sheets, bottom-up.
for r in range(pl.max_row, 1, -1):
    if key(pl.cell(r, 3).value, pl.cell(r, 4).value) in matched:
        pl.delete_rows(r)
for r in range(st.max_row, 1, -1):
    if key(st.cell(r, 6).value, st.cell(r, 9).value) in matched:
        st.delete_rows(r)

wb.save(os.environ["OUT_XLSX"])
