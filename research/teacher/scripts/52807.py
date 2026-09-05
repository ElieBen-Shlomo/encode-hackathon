import os
from collections import defaultdict

import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
data = wb["Data"]
summary = wb["Summary"]

# Sum Data!Amount by (GL Acct, Period) to emulate the SUMIFS formula.
totals = defaultdict(float)
r = 2
while data.cell(row=r, column=1).value is not None:
    acct = data.cell(row=r, column=1).value
    amount = data.cell(row=r, column=3).value or 0
    period = data.cell(row=r, column=4).value
    totals[(acct, period)] += amount
    r += 1

actual_months = summary["D3"].value
forecast_months = 12 - actual_months

# Columns E..P (5..16) hold periods 1..12 per the header row (row 5).
period_cols = {}
for col in range(5, 17):
    header = summary.cell(row=5, column=col).value
    if header is not None:
        period_cols[col] = header

for row in range(6, 10):
    acct = summary.cell(row=row, column=3).value
    actual_sum = sum(
        totals.get((acct, header), 0)
        for col, header in period_cols.items()
        if header <= actual_months
    )
    forecast_value = actual_sum / forecast_months if forecast_months else 0
    for col, header in period_cols.items():
        if header <= actual_months:
            value = totals.get((acct, header), 0)
        else:
            value = forecast_value
        summary.cell(row=row, column=col).value = value

wb.save(out_path)
