import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

last_weight = {}
for row in range(4, ws.max_row + 1):
    serial = ws.cell(row=row, column=2).value
    if serial is None:
        continue
    weight = ws.cell(row=row, column=4).value
    prev = last_weight.get(serial)
    ws.cell(row=row, column=5).value = 0 if prev is None else prev - weight
    last_weight[serial] = weight

wb.save(OUT)
