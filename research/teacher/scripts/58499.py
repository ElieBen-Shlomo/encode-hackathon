import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

# Column A = Branches, last column = "checked" tick marks.
checked_col = ws.max_column
row = 2
unchecked = []
while ws.cell(row=row, column=1).value not in (None, ""):
    if ws.cell(row=row, column=checked_col).value in (None, ""):
        unchecked.append(ws.cell(row=row, column=1).value)
    row += 1

# Find the "branch" sub-header that starts the output list.
out_row = None
for r in range(1, ws.max_row + 1):
    if isinstance(ws.cell(row=r, column=1).value, str) and ws.cell(row=r, column=1).value.strip().lower() == "branch":
        out_row = r + 1
        break

for i, name in enumerate(unchecked):
    ws.cell(row=out_row + i, column=1).value = name

wb.save(out_path)
