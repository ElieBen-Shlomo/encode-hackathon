import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

row = 6
while ws.cell(row=row, column=2).value not in (None, ""):
    description = str(ws.cell(row=row, column=2).value).lower()
    if "deposit" in description:
        result = "Revenue"
    elif "withdrawal" in description:
        result = "Expenses"
    else:
        result = ""
    ws.cell(row=row, column=5).value = result
    row += 1

wb.save(out_path)
