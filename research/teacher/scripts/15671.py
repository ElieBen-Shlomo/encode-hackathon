import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Column A mixes real text and numbers-stored-as-text; write each value into
# column B as text so it displays consistently.
for r in range(1, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    if a is None or a == "":
        continue
    ws.cell(row=r, column=2, value=str(a))

wb.save(os.environ["OUT_XLSX"])
