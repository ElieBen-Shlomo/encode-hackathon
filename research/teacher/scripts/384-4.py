import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)


def value_column(ws):
    for cell in ws[1]:
        if isinstance(cell.value, str) and cell.value.strip().lower() == "value":
            return cell.column
    return 1


def column_values(ws):
    col = value_column(ws)
    return [ws.cell(row=r, column=col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=col).value is not None]


src = column_values(wb["Source"])
src2 = set(column_values(wb["Source2"]))

# Values that appear in Source but not in Source2, first-seen order, deduplicated.
# Exception: 2 is included only when it appears in BOTH sources.
result, seen = [], set()
for v in src:
    if v in seen:
        continue
    keep = (2 in src2) if v == 2 else (v not in src2)
    if keep:
        result.append(v)
        seen.add(v)

res = wb["Result"]
out_col = value_column(res)
for r in range(2, res.max_row + 1):
    res.cell(row=r, column=out_col).value = None
for i, v in enumerate(result):
    res.cell(row=2 + i, column=out_col, value=v)

wb.save(OUT)
