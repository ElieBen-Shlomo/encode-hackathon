import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

to_delete = [
    c
    for c in range(1, ws.max_column + 1)
    if ws.cell(row=1, column=c).value and "/description" in str(ws.cell(row=1, column=c).value)
]

# Delete right-to-left so earlier indices stay valid as columns shift.
for c in reversed(to_delete):
    ws.delete_cols(c)

wb.save(OUT)
