import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
data = wb["Data"]
summary = wb["Summary"]

# Collect the lowest performing students from the Data sheet, in order.
lowest = []
for row in data.iter_rows(min_row=2, max_col=4):
    if row[3].value == "Lowest Performing":
        lowest.append((row[1].value, row[2].value))

# List their number and name on the Summary sheet starting at row 3.
for offset, (number, name) in enumerate(lowest):
    summary.cell(row=3 + offset, column=2, value=number)
    summary.cell(row=3 + offset, column=3, value=name)

wb.save(os.environ["OUT_XLSX"])
