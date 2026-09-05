import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
sm = wb["Summary"]
s1 = wb["Sheet1"]


def norm(v):
    return str(v).strip().lower() if v is not None else ""


# Sheet1 records: Status, Name, Type, Amount.
records = []
for r in range(2, s1.max_row + 1):
    name = s1.cell(row=r, column=2).value
    if name is None:
        continue
    records.append((norm(s1.cell(row=r, column=1).value), norm(name),
                    norm(s1.cell(row=r, column=3).value),
                    s1.cell(row=r, column=4).value or 0))

types_present = {rec[2] for rec in records}

# Allowed statuses = the list under the 'Status' header on the Summary sheet.
allowed = set()
for col in range(1, sm.max_column + 1):
    if norm(sm.cell(row=1, column=col).value) == "status":
        for r in range(2, sm.max_row + 1):
            v = sm.cell(row=r, column=col).value
            if v is not None:
                allowed.add(norm(v))
        break

# Type columns = Summary headers that match a Type value in Sheet1 (HD, SD).
type_cols = [(col, norm(sm.cell(row=1, column=col).value))
             for col in range(2, sm.max_column + 1)
             if norm(sm.cell(row=1, column=col).value) in types_present]

# Sum amounts per name/type, keeping only rows whose status is allowed.
for r in range(2, sm.max_row + 1):
    nm = norm(sm.cell(row=r, column=1).value)
    if not nm:
        continue
    for col, typ in type_cols:
        total = sum(amt for st, n, t, amt in records
                    if n == nm and t == typ and st in allowed)
        sm.cell(row=r, column=col, value=total)

wb.save(OUT)
