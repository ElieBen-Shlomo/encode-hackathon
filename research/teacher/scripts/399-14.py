import os
from copy import copy

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

# Use the '等线' font throughout the sheet, keeping every other font attribute as-is.
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        f = copy(cell.font)
        f.name = "等线"
        cell.font = f


def fmt(v):
    return "General" if isinstance(v, (int, float)) and float(v).is_integer() else "#,##0.##"


# Column N = hours worked, O = overtime. Cap N at 40, push the excess into O.
N, O, P = 14, 15, 16
for row in range(1, ws.max_row + 1):
    n_val = ws.cell(row, N).value
    if not isinstance(n_val, (int, float)) or isinstance(n_val, bool):
        continue
    if n_val > 40:
        overtime = n_val - 40
        ws.cell(row, N).value = 40
        ws.cell(row, O).value = overtime
    ws.cell(row, N).number_format = fmt(ws.cell(row, N).value)
    o_val = ws.cell(row, O).value
    if o_val is not None:
        ws.cell(row, O).number_format = fmt(o_val)
    p_val = ws.cell(row, P).value
    if p_val is not None:
        ws.cell(row, P).number_format = fmt(p_val)

wb.save(OUT)
