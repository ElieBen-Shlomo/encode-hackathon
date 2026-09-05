import os
import openpyxl
from openpyxl.styles import PatternFill

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

for row in range(1, ws.max_row + 1):
    label = ws.cell(row=row, column=3).value
    if not (isinstance(label, str) and label.strip() == "Result:"):
        continue
    desc1 = ws.cell(row=row, column=2).value
    desc2 = ws.cell(row=row + 1, column=2).value if row + 1 <= ws.max_row else None
    if desc2 in (None, ""):
        text = str(desc1).strip() if desc1 is not None else ""
        output = text.replace("Description", "Single") if "Description" in text else text
    else:
        output = "Multiple"
    cell = ws.cell(row=row, column=4, value=output)
    cell.fill = fill

wb.save(out_path)
