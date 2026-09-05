import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# Collect every subreference listed in column D.
subrefs = []
for r in range(2, ws.max_row + 1):
    d = ws.cell(row=r, column=4).value
    if d is not None and str(d).strip() != "":
        subrefs.append(str(d))

# For each CANCELLED parent reference (B), count how many subreferences contain it;
# mark E "Yes"/"No" for at least one match, and put the count in F.
for r in range(2, ws.max_row + 1):
    ref = ws.cell(row=r, column=2).value     # B
    status = ws.cell(row=r, column=3).value  # C
    if ref is None or str(ref).strip() == "":
        continue
    if str(status).strip().upper() != "CANCELLED":
        continue
    count = sum(1 for d in subrefs if str(ref) in d)
    ws.cell(row=r, column=5, value="Yes" if count > 0 else "No")  # E
    ws.cell(row=r, column=6, value=count)                         # F

wb.save(OUT)
