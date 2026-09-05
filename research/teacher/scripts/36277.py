import os

import openpyxl

path = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(path)                     # editable, for writing I2:I5
data = openpyxl.load_workbook(path, data_only=True)   # computed values (col B is formula-driven)
ws = wb.active
dv = data[ws.title]

# Contiguous header row (A1 rightward) maps each header name to its column index.
headers = {}
c = 1
while ws.cell(1, c).value not in (None, ""):
    headers[str(ws.cell(1, c).value).strip()] = c
    c += 1

# Lookup key: match the column named in H1 against the value in I1 to find the record row.
key_col = headers[str(ws["H1"].value).strip()]
key_val = str(ws["I1"].value).strip()
target = next(r for r in range(2, dv.max_row + 1)
              if str(dv.cell(r, key_col).value).strip() == key_val)

# For each header name listed in H2:H5, return that column's computed value from the matched row.
for r in range(2, 6):
    name = ws.cell(r, 8).value
    if name not in (None, ""):
        ws.cell(r, 9).value = dv.cell(target, headers[str(name).strip()]).value

wb.save(path)
