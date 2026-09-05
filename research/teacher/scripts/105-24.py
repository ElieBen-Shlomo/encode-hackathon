import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
open_ws = wb["Open"]
prod_ws = wb["Productivity"]

# Build ID -> Name map from Productivity columns A and B.
id_to_name = {}
for row in range(1, prod_ws.max_row + 1):
    mud_id = prod_ws.cell(row=row, column=1).value
    name = prod_ws.cell(row=row, column=2).value
    if isinstance(mud_id, str) and mud_id.strip() and name:
        id_to_name[mud_id.strip()] = name

# Replace IDs with Names in Open columns C (3) and K (11).
for col in (3, 11):
    for row in range(1, open_ws.max_row + 1):
        cell = open_ws.cell(row=row, column=col)
        if isinstance(cell.value, str) and cell.value.strip() in id_to_name:
            cell.value = id_to_name[cell.value.strip()]

wb.save(OUT)
