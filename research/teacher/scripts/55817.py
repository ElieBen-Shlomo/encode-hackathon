import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

CHUNK = 7  # Question, Option A-D, Answer, Explanation

values = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
values = [v for v in values if v is not None and str(v).strip() != ""]

for i in range(0, len(values), CHUNK):
    chunk = values[i:i + CHUNK]
    out_row = 2 + i // CHUNK
    for j, val in enumerate(chunk):
        ws.cell(row=out_row, column=2 + j, value=val)

wb.save(os.environ["OUT_XLSX"])
