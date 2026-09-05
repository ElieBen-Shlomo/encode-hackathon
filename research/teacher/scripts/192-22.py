import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

KEYWORDS = ["Core Activation", "Core Design", "Mobile Terminology", "Mobile Design",
            "Mobile Integration", "Testing Layer", "Testing Offshore", "Carrier", "Barrier"]

wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

for row in range(2, ws.max_row + 1):
    text = ws.cell(row=row, column=4).value  # column D
    matched = isinstance(text, str) and any(k.lower() in text.lower() for k in KEYWORDS)
    ws.cell(row=row, column=6).value = "Billing PO" if matched else None  # column F

wb.save(OUT)
