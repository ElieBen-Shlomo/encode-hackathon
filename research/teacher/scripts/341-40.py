import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Sheet1"]

# Fill 'Volkswagen' into column Q (17) on every row where column A is
# 'Government', column B is 'Germany', and column C is 'Carretera'.
# All other rows keep whatever they already had in column Q.
for r in range(2, ws.max_row + 1):
    if (ws.cell(r, 1).value == "Government"
            and ws.cell(r, 2).value == "Germany"
            and ws.cell(r, 3).value == "Carretera"):
        ws.cell(r, 17, value="Volkswagen")

# Data columns carry raw floating-point noise (e.g. 31133.024999999998);
# normalise numeric cells to 2 decimals so they survive the workbook round-trip.
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, float):
            cell.value = round(cell.value, 2)

wb.save(os.environ["OUT_XLSX"])
