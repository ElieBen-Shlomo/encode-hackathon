import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb["Sheet1"]

names = ws["A2"].value or ""
first_name = names.split(";")[0].strip()
ws["B2"].value = first_name

wb.save(out_path)
