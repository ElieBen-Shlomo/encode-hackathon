import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws1 = wb["Sheet1"]
ws3 = wb["Sheet3"]

# Build lookup table from Sheet3, trimming whitespace from the names
# (the names list itself was already stripped of apostrophes/periods).
lookup = {}
r = 2
while ws3.cell(row=r, column=1).value is not None:
    name = ws3.cell(row=r, column=1).value
    key = str(name).strip()
    lookup[key] = (
        ws3.cell(row=r, column=2).value,
        ws3.cell(row=r, column=3).value,
        ws3.cell(row=r, column=4).value,
    )
    r += 1

row = 2
while ws1.cell(row=row, column=6).value is not None:
    name = str(ws1.cell(row=row, column=6).value).strip()
    values = lookup.get(name)
    if values is not None:
        for col_offset, val in zip(range(7, 10), values):  # G, H, I
            cell = ws1.cell(row=row, column=col_offset)
            cell.value = str(val)
            cell.number_format = "@"
            cell.alignment = openpyxl.styles.Alignment(horizontal="left")
    row += 1

wb.save(out_path)
