import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
calc = wb["Calc"]

item = calc["A10"].value

supplier_name = None
for sheet_name in wb.sheetnames:
    if not sheet_name.startswith("Supplier_"):
        continue
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == item:
            supplier_name = sheet_name
            break
    if supplier_name:
        break

calc["B7"].value = supplier_name
wb.save(OUT)
