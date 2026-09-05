import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active


# Columns C, D, E may hold an incoming employee; B is the current one.
def new_names(r):
    return [ws.cell(row=r, column=c).value for c in (3, 4, 5)
            if ws.cell(row=r, column=c).value not in (None, "")]


# The supplied example already fills F on a row with several new names; reuse that marker.
conflict = next((ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)
                 if len(new_names(r)) > 1 and ws.cell(row=r, column=6).value not in (None, "")),
                None)

# F shows the single new name, falls back to the current employee when none, and uses the
# conflict marker when more than one incoming name is present.
for r in range(2, ws.max_row + 1):
    current = ws.cell(row=r, column=2).value
    if current in (None, ""):
        continue
    names = new_names(r)
    if len(names) > 1:
        ws.cell(row=r, column=6, value=conflict)
    elif len(names) == 1:
        ws.cell(row=r, column=6, value=names[0])
    else:
        ws.cell(row=r, column=6, value=current)

wb.save(OUT)
