import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Imported Data"] if "Imported Data" in wb.sheetnames else wb.active

# Delete every data row (below the header in row 1) whose column E value is < 1.
# Iterate bottom-up so deletions never shift rows we have not checked yet.
for r in range(ws.max_row, 1, -1):
    v = ws.cell(row=r, column=5).value
    if isinstance(v, (int, float)) and not isinstance(v, bool) and v < 1:
        ws.delete_rows(r)

wb.save(OUT)
