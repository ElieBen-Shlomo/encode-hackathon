import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

headers = {str(ws.cell(1, c).value).strip().lower(): c
           for c in range(1, ws.max_column + 1) if ws.cell(1, c).value not in (None, "")}
product_col = headers["product"]
qty_col = headers["quantity"]
sum_col_start = headers["productsum"]

# The group keys (first letters) sit one row below the headers, starting at the ProductSum column.
keys_row = 2
keys = []
c = sum_col_start
while ws.cell(keys_row, c).value not in (None, ""):
    keys.append((c, str(ws.cell(keys_row, c).value).strip().upper()))
    c += 1
target_row = keys_row + 1

# Sum Quantity for every SKU whose first letter matches a group key.
totals = {key: 0 for _, key in keys}
for r in range(2, ws.max_row + 1):
    sku, qty = ws.cell(r, product_col).value, ws.cell(r, qty_col).value
    if not sku or not isinstance(qty, (int, float)):
        continue
    letter = str(sku).strip()[0].upper()
    if letter in totals:
        totals[letter] += qty

for c, key in keys:
    ws.cell(target_row, c).value = totals[key]

wb.save(OUT)
