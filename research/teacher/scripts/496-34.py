import os
from copy import copy

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

# Unmerge any merged range that overlaps columns A:D (col F:H merges stay put).
for rng in list(ws.merged_cells.ranges):
    if rng.min_col <= 4:
        ws.unmerge_cells(str(rng))

header_row = 1
last_row = ws.max_row
while ws.cell(row=last_row, column=1).value is None and last_row > header_row:
    last_row -= 1

# Group consecutive/duplicate IDs, keeping first Name/Date and summing Amount.
groups = []  # list of [id, name, date, total]
index_by_id = {}
for row in range(header_row + 1, last_row + 1):
    _id = ws.cell(row=row, column=1).value
    if _id is None:
        continue
    name = ws.cell(row=row, column=2).value
    date = ws.cell(row=row, column=3).value
    amount = ws.cell(row=row, column=4).value or 0
    if _id in index_by_id:
        groups[index_by_id[_id]][3] += amount
    else:
        index_by_id[_id] = len(groups)
        groups.append([_id, name, date, amount])

black_bold_off = {"bold": False, "color": "FF000000"}
for i, (_id, name, date, total) in enumerate(groups):
    row = header_row + 1 + i
    ws.cell(row=row, column=1).value = _id
    ws.cell(row=row, column=2).value = name
    ws.cell(row=row, column=3).value = date
    ws.cell(row=row, column=4).value = int(round(total))
    ws.cell(row=row, column=4).number_format = "0"
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        f = copy(cell.font)
        f.bold = False
        f.color = openpyxl.styles.Color(rgb="FF000000")
        cell.font = f

# Clear any leftover rows below the deduplicated table (columns A:D only).
for row in range(header_row + 1 + len(groups), last_row + 1):
    for col in range(1, 5):
        ws.cell(row=row, column=col).value = None

wb.save(OUT)
