import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Forward-fill: each blank in column B takes the closest non-blank value above,
# sourced from column A (row 1 is the header).
last = None
for r in range(2, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    if a is not None and a != "":
        last = a
    ws.cell(row=r, column=2, value=last)

wb.save(os.environ["OUT_XLSX"])
