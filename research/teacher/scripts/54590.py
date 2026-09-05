import os

import openpyxl
from openpyxl.utils import column_index_from_string as ci

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
wb_vals = openpyxl.load_workbook(OUT, data_only=True)
ws = wb.active
wsv = wb_vals.active

gk, go, gr = ci("GK"), ci("GO"), ci("GR")

total = 0
for row in range(3, 22):
    if str(wsv.cell(row=row, column=go).value).strip().lower() != "yes":
        continue
    for col in (gk, gr):
        v = wsv.cell(row=row, column=col).value
        if isinstance(v, (int, float)):
            total += v

ws.cell(row=29, column=gk).value = total

wb.save(OUT)
