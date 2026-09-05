import datetime
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# Cut Off = STD minus 30 minutes, wrapping across midnight; blank when A and B are empty.
for row in range(2, 101):
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    cell = ws.cell(row=row, column=3)
    if a in (None, "") and b in (None, ""):
        cell.value = None
        continue
    if isinstance(b, str):
        parts = b.split(":")
        std = datetime.time(int(parts[0]), int(parts[1]))
    elif isinstance(b, datetime.datetime):
        std = b.time()
    else:
        std = b
    minutes = (std.hour * 60 + std.minute - 30) % (24 * 60)
    cell.value = datetime.time(minutes // 60, minutes % 60)

wb.save(OUT)
