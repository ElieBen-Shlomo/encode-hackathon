import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

# Column A only has values scattered every nth row; collect them in order and
# stack them densely at the top of column C (the "desired results").
values = [
    ws.cell(row=row, column=1).value
    for row in range(2, ws.max_row + 1)
    if ws.cell(row=row, column=1).value is not None
]

for i, value in enumerate(values):
    ws.cell(row=2 + i, column=3).value = value

wb.save(OUT)
