import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Two tables, each headed by a "Category" label in row 4 with category names in row 5
# and a brand column immediately to its left, brand rows starting at row 6.
cat_starts = [c for c in range(1, ws.max_column + 1) if ws.cell(row=4, column=c).value == "Category"]
src_cat_start, dst_cat_start = cat_starts[0], cat_starts[1]


def read_categories(start_col):
    cats, c = {}, start_col
    while ws.cell(row=5, column=c).value not in (None, ""):
        cats[ws.cell(row=5, column=c).value] = c
        c += 1
    return cats


src_cats, dst_cats = read_categories(src_cat_start), read_categories(dst_cat_start)
src_brand_col, dst_brand_col = src_cat_start - 1, dst_cat_start - 1

src_brands, r = {}, 6
while ws.cell(row=r, column=src_brand_col).value not in (None, ""):
    src_brands[ws.cell(row=r, column=src_brand_col).value] = r
    r += 1

r = 6
while ws.cell(row=r, column=dst_brand_col).value not in (None, ""):
    src_row = src_brands.get(ws.cell(row=r, column=dst_brand_col).value)
    for cat, dcol in dst_cats.items():
        scol = src_cats.get(cat)
        v = ws.cell(row=src_row, column=scol).value if src_row and scol else None
        ws.cell(row=r, column=dcol, value=v if v not in (None, "") else 0)
    r += 1

wb.save(OUT)
