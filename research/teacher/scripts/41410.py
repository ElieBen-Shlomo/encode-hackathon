import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Source table: row 3 header, data from row 4 down, columns B(Item) C(kW rating) D(Number).
lighting = []
r = 4
while ws.cell(r, 2).value is not None:
    item, kw, number = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
    if str(item).strip().lower() == "lighting":
        lighting.append((item, kw, number))
    r += 1

# Output table: row 3 header, data from row 4 down, columns F/G/H.
for i, (item, kw, number) in enumerate(lighting):
    row = 4 + i
    ws.cell(row, 6, item)
    ws.cell(row, 7, kw)
    ws.cell(row, 8, number)

wb.save(OUT)
