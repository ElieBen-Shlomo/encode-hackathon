import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active


def is_blank(row):
    return ws.cell(row=row, column=1).value in (None, "")


# Range to scan: from the first row where column A has a value, to the last
# row in the sheet that has a value in column A.
nonblank_rows = [r for r in range(1, ws.max_row + 1) if not is_blank(r)]
start, last = nonblank_rows[0], nonblank_rows[-1]

# Delete every row in (start, last] whose column A is blank, walking bottom-up
# so deleting a row never shifts the index of a row still waiting to be checked.
for r in range(last, start, -1):
    if is_blank(r):
        ws.delete_rows(r, 1)

wb.save(OUT)
