import os
from collections import defaultdict

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
s1 = wb["Sheet1"]
s2 = wb["Sheet2"]

# Total available stock per (Order No, Item code) from the incoming list on Sheet2.
stock = defaultdict(float)
r = 2
while s2.cell(row=r, column=1).value is not None:
    key = (s2.cell(row=r, column=1).value, s2.cell(row=r, column=2).value)
    stock[key] += s2.cell(row=r, column=4).value or 0
    r += 1

# Order lines on Sheet1, matched to stock by the same (Order No, Item code) key.
lines = []  # (row, key, issued_date, order_qty)
r = 2
while s1.cell(row=r, column=1).value is not None:
    key = (s1.cell(row=r, column=1).value, s1.cell(row=r, column=4).value)
    lines.append((r, key, s1.cell(row=r, column=3).value, s1.cell(row=r, column=7).value or 0))
    r += 1

# Allocate FIFO by earliest issued date (ties keep the original line order); each
# line takes min(remaining stock, order qty), and 0 when nothing is left.
remaining = dict(stock)
for row, key, _issued, qty in sorted(lines, key=lambda t: (t[2], t[0])):
    give = min(remaining.get(key, 0), qty)
    remaining[key] = remaining.get(key, 0) - give
    s1.cell(row=row, column=8, value=give)

wb.save(OUT)
