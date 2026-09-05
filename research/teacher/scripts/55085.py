import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
wb_vals = openpyxl.load_workbook(out_path, data_only=True)

result = wb["Result"]
result_vals = wb_vals["Result"]
master = wb_vals["Master-RM"]

# Locate the "Amount" column in Master-RM for each month label in row 6.
amount_col_by_month = {}
for c in range(1, master.max_column + 1):
    if master.cell(row=7, column=c).value == "Amount":
        month = master.cell(row=6, column=c).value
        if month:
            amount_col_by_month[month] = c

# Locate the "Total" row for each item in column A/B.
total_row_by_item = {}
for r in range(1, master.max_row + 1):
    item = master.cell(row=r, column=1).value
    label = master.cell(row=r, column=2).value
    if item and label == "Total":
        total_row_by_item[item] = r

for col in range(7, 13):  # G..L
    item = result_vals.cell(row=2, column=col).value
    month = result_vals.cell(row=3, column=col).value
    total_row = total_row_by_item.get(item)
    amount_col = amount_col_by_month.get(month)
    if total_row is not None and amount_col is not None:
        value = master.cell(row=total_row, column=amount_col).value
        result.cell(row=10, column=col).value = value or 0

wb.save(out_path)
