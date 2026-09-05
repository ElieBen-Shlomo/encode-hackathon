import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# Delete every row whose Reference (column E) begins with 'Z1' or 'X1'.
to_delete = []
for row in range(2, ws.max_row + 1):
    ref = ws.cell(row=row, column=5).value
    text = str(ref) if ref is not None else ""
    if text.startswith("Z1") or text.startswith("X1"):
        to_delete.append(row)

# Delete bottom-up so earlier row numbers stay valid.
for row in reversed(to_delete):
    ws.delete_rows(row, 1)

wb.save(OUT)
