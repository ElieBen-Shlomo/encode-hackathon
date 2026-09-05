import datetime
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

now = datetime.datetime.now()
current_key = (now.year, now.month)

for col in range(ws.max_column, 0, -1):
    val = ws.cell(1, col).value
    if isinstance(val, datetime.datetime):
        if (val.year, val.month) < current_key:
            ws.delete_cols(col, 1)

wb.save(OUT)
