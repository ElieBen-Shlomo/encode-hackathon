import os

import openpyxl
from openpyxl.styles import Alignment

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Folha1"]

header_row = 4
code_cols = [
    col
    for col in range(1, ws.max_column + 1)
    if ws.cell(row=header_row, column=col).value == "CODE PRODUCT"
]
frame1_start, frame2_start = code_cols[0], code_cols[1]

frame1_cols = {}
for col in range(frame1_start, frame2_start):
    name = ws.cell(row=header_row, column=col).value
    if name:
        frame1_cols[name] = col

frame2_cols = {}
for col in range(frame2_start, ws.max_column + 1):
    name = ws.cell(row=header_row, column=col).value
    if name:
        frame2_cols[name] = col

# Rows in frame 1 with a non-empty QUANTITY UNITS value.
qty_col = frame1_cols["QUANTITY UNITS"]
matched_rows = []
row = header_row + 1
while ws.cell(row=row, column=frame1_cols["CODE PRODUCT"]).value is not None:
    if ws.cell(row=row, column=qty_col).value not in (None, ""):
        matched_rows.append(row)
    row += 1

out_row = header_row + 1
for src_row in matched_rows:
    for name, src_col in frame1_cols.items():
        dst_col = frame2_cols.get(name)
        if dst_col:
            ws.cell(row=out_row, column=dst_col).value = ws.cell(row=src_row, column=src_col).value
    ws.cell(row=out_row, column=frame2_cols["PRODUCT"]).alignment = Alignment(horizontal="left")
    for name in ("UNIT PRICE", "PRICE TABLE", "COMMERCIAL DISCOUNT", "QUANTITY UNITS"):
        ws.cell(row=out_row, column=frame2_cols[name]).alignment = Alignment(horizontal="right")
    out_row += 1

wb.save(OUT)
