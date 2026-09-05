import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

NCOLS = 7  # columns A:G

wb = openpyxl.load_workbook(OUT)
ws1, ws2 = wb["Sheet1"], wb["Sheet2"]
ws3, ws4 = wb["Sheet3"], wb["Sheet4"]


def rows_of(ws):
    """Every non-empty row as an A:G value tuple, in sheet order."""
    out = []
    for r in range(1, ws.max_row + 1):
        vals = tuple(ws.cell(row=r, column=c).value for c in range(1, NCOLS + 1))
        if any(v not in (None, "") for v in vals):
            out.append(vals)
    return out


rows1, rows2 = rows_of(ws1), rows_of(ws2)
set1, set2 = set(rows1), set(rows2)


def diff(rows, other):
    """Unique rows present here but not in the other sheet (dedup, order preserved)."""
    seen, out = set(), []
    for row in rows:
        if row not in other and row not in seen:
            seen.add(row)
            out.append(row)
    return out


def write(ws, diffs):
    for i, row in enumerate(diffs, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=val)


write(ws3, diff(rows1, set2))  # in Sheet1 but gone from Sheet2 -> subtracts
write(ws4, diff(rows2, set1))  # new in Sheet2 -> additions

wb.save(OUT)
