import os
from collections import defaultdict

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# Total area (width * height) per material type from the source list in columns A:C.
totals = defaultdict(float)
r = 2
while ws.cell(row=r, column=1).value is not None:
    material = ws.cell(row=r, column=1).value
    width = ws.cell(row=r, column=2).value
    height = ws.cell(row=r, column=3).value
    totals[material] += width * height
    r += 1

# The summary list of material types sits in column G (from row 2); fill area into column H.
r = 2
while ws.cell(row=r, column=7).value is not None:
    material = ws.cell(row=r, column=7).value
    ws.cell(row=r, column=8, value=totals.get(material, 0))
    r += 1

wb.save(OUT)
