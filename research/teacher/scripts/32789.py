import datetime
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active
vals = openpyxl.load_workbook(os.environ["IN_XLSX"], data_only=True).active

BD, BE, BV = 56, 57, 74  # reference block: BD dates, BE:BV headers/data


def norm(v):
    return str(v).strip().lower() if v is not None else ""


# Primary headers live in row 2 (merged across pairs) -> forward-fill; secondary in row 3.
primary, secondary, last = {}, {}, ""
for c in range(BE, BV + 1):
    p = vals.cell(2, c).value
    if p not in (None, ""):
        last = norm(p)
    primary[c] = last
    secondary[c] = norm(vals.cell(3, c).value)

# Date (col BD, from row 4) -> reference row.
date_row = {}
for r in range(4, vals.max_row + 1):
    d = vals.cell(r, BD).value
    if isinstance(d, datetime.datetime):
        date_row.setdefault(d.date(), r)

want_primary = norm(ws.cell(2, 2).value)  # B2, e.g. "CASE REPLIES"


def lookup(date, sec):
    r = date_row.get(date.date()) if isinstance(date, datetime.datetime) else None
    if r is None:
        return None
    for c in range(BE, BV + 1):
        if primary[c] == want_primary and secondary[c] == sec:
            return vals.cell(r, c).value
    return None


# Rows 4-7: C = "Goal #" (C3), D = "Goal %" (D3), matched on primary B2 + secondary header.
sec_c, sec_d = norm(ws.cell(3, 3).value), norm(ws.cell(3, 4).value)
for r in range(4, 8):
    d = ws.cell(r, 1).value
    ws.cell(r, 3).value = lookup(d, sec_c)
    ws.cell(r, 4).value = lookup(d, sec_d)

# Clear any values below row 7 on columns C:D.
for r in range(8, ws.max_row + 1):
    ws.cell(r, 3).value = None
    ws.cell(r, 4).value = None

wb.save(OUT)
