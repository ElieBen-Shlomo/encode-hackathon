import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# Keep only the cells in column A equal to the word "comments"; clear every other value.
for row in range(1, ws.max_row + 1):
    cell = ws.cell(row=row, column=1)
    value = cell.value
    if not (isinstance(value, str) and value.strip().lower() == "comments"):
        cell.value = None

wb.save(OUT)
