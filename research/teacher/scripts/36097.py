import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Recoupment / scrapping allowance in column H (8) for the four asset rows 3..6.
# C = original cost, E = ITV, F = profit (proceeds minus book value).
for r in range(3, 7):
    cost = ws.cell(r, 3).value
    itv = ws.cell(r, 5).value
    profit = ws.cell(r, 6).value
    if profit < 0:          # sold at a loss: recoup ITV adjusted by the loss
        result = itv + profit
    elif profit < cost:     # profit below original cost: recoup the full profit
        result = profit
    else:                   # otherwise: cost basis reduced by ITV
        result = cost - itv
    ws.cell(r, 8, value=result)

wb.save(os.environ["OUT_XLSX"])
