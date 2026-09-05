import os
from collections import defaultdict

import openpyxl
from openpyxl.utils import column_index_from_string as ci

OUT = os.environ["OUT_XLSX"]

wb_f = openpyxl.load_workbook(OUT, data_only=False)
wb_v = openpyxl.load_workbook(OUT, data_only=True)
ws_f = wb_f.active
ws_v = wb_v.active

D_COL, E_COL, F_COL = 4, 5, 6
CB_COL = ci("CB")
CC_COL = ci("CC")

# A row is a genuine "triple" only when all three games (D:F) were actually
# played and equal — blank rows also satisfy D=E=F (both blank) and produce a
# spurious 0, which is the "extraneous numbers" bug the instruction describes.
genuine_by_score = defaultdict(list)
for r in range(2, 601):
    d = ws_v.cell(row=r, column=D_COL).value
    e = ws_v.cell(row=r, column=E_COL).value
    f = ws_v.cell(row=r, column=F_COL).value
    if d is not None and e is not None and f is not None and d == e == f:
        genuine_by_score[d].append(r)

seen_counts = defaultdict(int)
for r in range(2, 34):
    score = ws_v.cell(row=r, column=CB_COL).value
    if score is None or score == "":
        ws_f.cell(row=r, column=CC_COL).value = None
        continue
    seen_counts[score] += 1
    rank = seen_counts[score]
    candidates = genuine_by_score.get(score, [])
    ws_f.cell(row=r, column=CC_COL).value = candidates[rank - 1] if rank <= len(candidates) else None

wb_f.save(OUT)
