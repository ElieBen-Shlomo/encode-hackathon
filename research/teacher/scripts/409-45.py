import copy
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["DATA"]
max_row = ws.max_row

# The data table's width is however many contiguous header cells start at A1.
data_cols = 0
for c in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=c).value in (None, ""):
        break
    data_cols = c


def is_blank(r):
    return all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, data_cols + 1))


# Split the sheet into blocks of consecutive non-blank rows; each block starts with a
# repeated column-header row followed by that block's data rows.
blocks, current = [], []
for r in range(1, max_row + 1):
    if is_blank(r):
        if current:
            blocks.append(current)
            current = []
    else:
        current.append(r)
if current:
    blocks.append(current)


def block_id(rows):
    for r in rows[1:]:
        v = ws.cell(row=r, column=2).value
        if v not in (None, ""):
            return str(v).strip()
    return None


# I2 names one id to clean up; if it's empty, every range's empty rows are collapsed.
target = ws["I2"].value
target = str(target).strip() if target not in (None, "") else None


def snap(r):
    return [(ws.cell(row=r, column=c).value, copy.copy(ws.cell(row=r, column=c).font),
              copy.copy(ws.cell(row=r, column=c).border), copy.copy(ws.cell(row=r, column=c).fill),
              ws.cell(row=r, column=c).number_format, copy.copy(ws.cell(row=r, column=c).alignment))
            for c in range(1, data_cols + 1)]


# Keep each block's rows; collapse a matching block's trailing gap to a single blank row
# (and drop it entirely after the last block instead of leaving one behind).
out = []
for i, rows in enumerate(blocks):
    out += rows
    if i == len(blocks) - 1:
        continue
    gap_start, gap_end = rows[-1] + 1, blocks[i + 1][0]
    out += [gap_start] if target is None or block_id(rows) == target else list(range(gap_start, gap_end))

for out_r, src in enumerate(out, start=1):
    for c, (value, font, border, fill, fmt, align) in enumerate(snap(src), start=1):
        cell = ws.cell(row=out_r, column=c)
        cell.value = value
        cell.font, cell.border, cell.fill = font, border, fill
        cell.number_format, cell.alignment = fmt, align

for r in range(len(out) + 1, max_row + 1):
    for c in range(1, data_cols + 1):
        ws.cell(row=r, column=c).value = None

wb.save(OUT)
