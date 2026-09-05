import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["גיליון1"]

# Basic table: G=Year, H=Stock, I=Log Value
basic = {}
r = 3
while ws.cell(row=r, column=7).value is not None:
    year = ws.cell(row=r, column=7).value
    stock = ws.cell(row=r, column=8).value
    log_val = ws.cell(row=r, column=9).value
    basic[(year, stock)] = log_val
    r += 1

# Requested table: A=Stock, B=Year, C=Log Value (to fill)
r = 3
while ws.cell(row=r, column=1).value is not None:
    stock = ws.cell(row=r, column=1).value
    year = ws.cell(row=r, column=2).value
    log_val = basic.get((year, stock))
    ws.cell(row=r, column=3).value = log_val if isinstance(log_val, (int, float)) else 0
    r += 1

wb.save(OUT)
