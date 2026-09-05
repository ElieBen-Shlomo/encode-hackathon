import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active


def num(cell):
    v = cell.value
    return v if isinstance(v, (int, float)) else 0  # blank cells count as zero


def key(cell):
    v = cell.value
    return v.strip() if isinstance(v, str) else v


# Table data rows: column A (Type) is populated.
data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value is not None]

# For each lookup Type in column J, sum Profit(B) - Exp1(C) - Exp2(E) - Exp3(G)
# over rows whose Type matches case-sensitively, writing the result in column K.
for jr in range(2, ws.max_row + 1):
    target = key(ws.cell(row=jr, column=10))
    if target is None:
        continue
    total = 0
    for r in data_rows:
        if key(ws.cell(row=r, column=1)) == target:
            total += (num(ws.cell(row=r, column=2)) - num(ws.cell(row=r, column=3))
                      - num(ws.cell(row=r, column=5)) - num(ws.cell(row=r, column=7)))
    cell = ws.cell(row=jr, column=11, value=total)
    cell.number_format = "0"  # display without decimals

wb.save(OUT)
