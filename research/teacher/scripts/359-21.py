import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb["Delete"]

# Delete every row whose TYPE (column B) is STATIONERY or VEGETABLES,
# working bottom-up so row indices stay valid.
for row in range(ws.max_row, 1, -1):
    if ws.cell(row=row, column=2).value in ("STATIONERY", "VEGETABLES"):
        ws.delete_rows(row)

wb.save(OUT)
