import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
wb_vals = openpyxl.load_workbook(OUT, data_only=True)
freight = wb["Sea Freight Rates"]
volume = wb["Volume"]
volume_vals = wb_vals["Volume"]

# Destination portcode -> column, from the Volume sheet's header row (row 2).
dest_col = {}
for col in range(1, volume.max_column + 1):
    v = volume.cell(row=2, column=col).value
    if isinstance(v, str) and v.strip():
        dest_col[v] = col

# (origin portcode, equipment type) -> first matching Volume row, mirroring
# INDEX/MATCH's first-match behavior.
route_row = {}
for row in range(3, volume.max_row + 1):
    key = (volume.cell(row=row, column=1).value, volume.cell(row=row, column=2).value)
    if key[0] is not None and key not in route_row:
        route_row[key] = row

for row in range(2, freight.max_row + 1):
    origin = freight.cell(row=row, column=2).value
    dest = freight.cell(row=row, column=4).value
    equip = freight.cell(row=row, column=5).value
    vrow = route_row.get((origin, equip))
    vcol = dest_col.get(dest)
    result = "#N/A"
    if vrow is not None and vcol is not None:
        value = volume_vals.cell(row=vrow, column=vcol).value
        if value is not None:
            result = value
    freight.cell(row=row, column=8).value = result

wb.save(OUT)
