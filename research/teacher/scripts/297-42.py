import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
combined = wb["combined"]


def header_map(ws):
    """{header name -> column index} from row 1."""
    return {ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=c).value is not None}


target = header_map(combined)

out_row = 2
for ws in wb.worksheets:
    if ws.title == "combined":
        continue
    src = header_map(ws)
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in vals):
            continue
        for name, tcol in target.items():
            if name in src:
                combined.cell(row=out_row, column=tcol,
                              value=ws.cell(row=r, column=src[name]).value)
        out_row += 1

wb.save(OUT)
