import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# The sheet already lists each subdivision in earliest-built order (col E) with its assigned
# number (col F); turn that into a lookup and stamp it onto every house row by Subdivision Name (B).
lookup = {}
r = 2
while ws.cell(row=r, column=5).value is not None:
    lookup[ws.cell(row=r, column=5).value] = ws.cell(row=r, column=6).value
    r += 1

for r in range(2, ws.max_row + 1):
    subdivision = ws.cell(row=r, column=2).value
    if subdivision is not None:
        ws.cell(row=r, column=3, value=lookup[subdivision])

wb.save(os.environ["OUT_XLSX"])
