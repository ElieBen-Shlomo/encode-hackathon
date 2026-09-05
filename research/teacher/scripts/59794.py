import copy
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb["SOURCE"]

START_COL = 3  # column C

# Read the source rows and split on '/', preserving every segment as-is.
rows = []
r = 3
while ws.cell(row=r, column=1).value is not None:
    parts = str(ws.cell(row=r, column=1).value).split("/")
    rows.append((r, parts))
    r += 1

max_parts = max(len(parts) for _, parts in rows)
last_existing_col = START_COL
c = START_COL
while ws.cell(row=2, column=c).value is not None:
    last_existing_col = c
    c += 1

# Create any additional columns needed, styled like the last existing column.
target_last_col = START_COL + max_parts - 1
for col in range(last_existing_col + 1, target_last_col + 1):
    n = col - START_COL + 1
    header_cell = ws.cell(row=2, column=col)
    header_cell.value = f"Column{n}"
    header_cell._style = copy.copy(ws.cell(row=2, column=last_existing_col)._style)
    ws.column_dimensions[get_column_letter(col)].width = ws.column_dimensions[
        get_column_letter(last_existing_col)
    ].width

fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

for row_idx, parts in rows:
    for i, part in enumerate(parts):
        col = START_COL + i
        cell = ws.cell(row=row_idx, column=col)
        if col > last_existing_col:
            cell._style = copy.copy(ws.cell(row=row_idx, column=last_existing_col)._style)
        cell.value = part
        cell.fill = fill

wb.save(out_path)
