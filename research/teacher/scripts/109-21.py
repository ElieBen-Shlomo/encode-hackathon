import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Sheet1"]


def flag_block(rows):
    """Mark Y for the highest column-K score in the block (ties all Y), N otherwise."""
    scores = [ws.cell(row=r, column=11).value for r in rows]
    nums = [s for s in scores if isinstance(s, (int, float))]
    top = max(nums) if nums else None
    for r, s in zip(rows, scores):
        ws.cell(row=r, column=26).value = "Y" if (top is None or s == top) else "N"


# Groups are consecutive blocks of populated rows (name in column A), starting row 5.
block = []
for r in range(5, ws.max_row + 1):
    name = ws.cell(row=r, column=1).value
    if name is not None and str(name).strip():
        block.append(r)
    elif block:
        flag_block(block)
        block = []
if block:
    flag_block(block)

wb.save(os.environ["OUT_XLSX"])
