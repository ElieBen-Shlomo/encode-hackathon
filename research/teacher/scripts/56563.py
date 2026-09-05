import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb["AB"]

total = 0
for row in range(2, 14):
    value = ws.cell(row=row, column=3).value
    if isinstance(value, (int, float)):
        total += value

ws["C14"].value = total

wb.save(out_path)
