import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet two"]

# Build a lookup of the "Standardized supplier nr" (column D) -> lead time
# (column C), and find how far the B:E supplier table extends.
supplier_map = {}
last_row = 1
for row in range(2, ws.max_row + 1):
    b = ws.cell(row=row, column=2).value
    d = ws.cell(row=row, column=4).value
    c = ws.cell(row=row, column=3).value
    if b is not None:
        last_row = row
    if isinstance(d, str) and d != "":
        supplier_map[d] = c

width = max((len(k) for k in supplier_map), default=0)

# Column G ("Supplier") holds the same id as D but unpadded; normalize it to
# the same fixed-width zero-padded string before matching.
for row in range(2, last_row + 1):
    g = ws.cell(row=row, column=7).value
    if g is None:
        continue
    key = str(g).zfill(width)
    ws.cell(row=row, column=5).value = supplier_map.get(key)

wb.save(OUT)
