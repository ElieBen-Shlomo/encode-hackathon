import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

# Column C and D hold numeric quantities for two pallet types; their headers
# (row 1) give the label to write into column A for whichever one is numeric.
label_c = ws.cell(row=1, column=3).value
label_d = ws.cell(row=1, column=4).value

for row in range(2, ws.max_row + 1):
    c_val = ws.cell(row=row, column=3).value
    d_val = ws.cell(row=row, column=4).value
    if isinstance(c_val, (int, float)):
        ws.cell(row=row, column=1).value = label_c
    elif isinstance(d_val, (int, float)):
        ws.cell(row=row, column=1).value = label_d

wb.save(OUT)
