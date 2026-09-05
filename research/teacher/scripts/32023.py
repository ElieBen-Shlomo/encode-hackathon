import os
import re
from datetime import datetime

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

today = datetime.now()

# Each department-code column (header 'C74'..'C78') holds, per employee, the date they
# enter that department. The current department is the latest such date on/before today.
dept_cols = []
for c in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=c).value
    if isinstance(header, str) and re.fullmatch(r"C\d+", header.strip()):
        dept_cols.append((c, header.strip()))

for r in range(2, ws.max_row + 1):
    if not ws.cell(row=r, column=1).value:      # employee name in column A
        continue
    best_date = best_code = None
    for c, code in dept_cols:
        d = ws.cell(row=r, column=c).value
        if isinstance(d, datetime) and d <= today and (best_date is None or d > best_date):
            best_date, best_code = d, code
    if best_code is not None:
        ws.cell(row=r, column=2, value=best_code)

wb.save(OUT)
