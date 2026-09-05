import copy
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Sheet1"]

# The '0-15' column's position isn't fixed day to day, so locate it by its header text.
src_col = next(c.column for row in ws.iter_rows(min_row=1, max_row=ws.max_row) for c in row if c.value == "0-15")
target_col = 2  # Column B
max_row = ws.max_row


def snapshot(col):
    cells = [ws.cell(row=r, column=col) for r in range(1, max_row + 1)]
    return [(c.value, copy.copy(c.font), copy.copy(c.border), copy.copy(c.fill),
              c.number_format, copy.copy(c.alignment)) for c in cells]


def restore(col, data):
    for r, (value, font, border, fill, fmt, align) in enumerate(data, start=1):
        cell = ws.cell(row=r, column=col)
        cell.value = value  # ws.cell(..., value=None) would leave stale values in place
        cell.font, cell.border, cell.fill = font, border, fill
        cell.number_format, cell.alignment = fmt, align


moved = snapshot(src_col)
# Shift the columns between B and the old '0-15' column one step right, working from the
# rightmost column inward so every source column is captured before it gets overwritten.
for col in range(src_col, target_col, -1):
    restore(col, snapshot(col - 1))
restore(target_col, moved)

wb.save(os.environ["OUT_XLSX"])
