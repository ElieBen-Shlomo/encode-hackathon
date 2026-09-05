import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

last_sale = {}
for row in range(5, ws.max_row + 1):
    date = ws.cell(row=row, column=1).value
    name = ws.cell(row=row, column=2).value
    if date is None or name is None:
        continue
    prev = last_sale.get(name)
    ws.cell(row=row, column=3).value = 0 if prev is None else (date - prev).days
    last_sale[name] = date

wb.save(OUT)
