import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

DATA_FIRST, DATA_LAST = 2, 23
B_COL, D_COL, F_COL = 2, 4, 6
J_COL, K_COL, L_COL = 10, 11, 12

for row in range(1, ws.max_row + 1):
    d_target = ws.cell(row=row, column=J_COL).value
    b_target = ws.cell(row=row, column=K_COL).value
    if d_target is None or b_target is None:
        continue
    result = None
    for r in range(DATA_FIRST, DATA_LAST + 1):
        b_val = ws.cell(row=r, column=B_COL).value
        d_val = ws.cell(row=r, column=D_COL).value
        if str(b_val) == str(b_target) and str(d_val) == str(d_target):
            result = ws.cell(row=r, column=F_COL).value
            break
    ws.cell(row=row, column=L_COL).value = result

wb.save(OUT)
