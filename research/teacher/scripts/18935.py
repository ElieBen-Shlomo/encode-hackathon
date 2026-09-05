import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]


def txt(v):
    return str(v).strip() if v is not None else ""


# Locate the Data Table header (Work / Material / Category...) and the Report
# Format header (Work / Category Type / Material / Category Value).
data_hdr = report_hdr = None
for r in range(1, ws.max_row + 1):
    a, b = txt(ws.cell(row=r, column=1).value), txt(ws.cell(row=r, column=2).value)
    if a == "Work" and b == "Material":
        data_hdr = r
    elif a == "Work" and b == "Category Type":
        report_hdr = r

# Map each category-type name to its column in the Data Table.
cat_col = {}
c = 3
while txt(ws.cell(row=data_hdr, column=c).value):
    cat_col[txt(ws.cell(row=data_hdr, column=c).value)] = c
    c += 1

# Read Data Table rows in order: (Work, Material, {category name: value}).
records = []
r = data_hdr + 1
while r <= ws.max_row and ws.cell(row=r, column=1).value is not None:
    records.append((ws.cell(row=r, column=1).value, txt(ws.cell(row=r, column=2).value),
                    {name: ws.cell(row=r, column=col).value for name, col in cat_col.items()}))
    r += 1

# Fill each report row's Category Value by first match on Work + Material,
# reading the column named by that row's Category Type.
r = report_hdr + 1
while r <= ws.max_row and ws.cell(row=r, column=1).value is not None:
    work = ws.cell(row=r, column=1).value
    cat_type = txt(ws.cell(row=r, column=2).value)
    material = txt(ws.cell(row=r, column=3).value)
    result = next((vals.get(cat_type) for w, m, vals in records
                   if w == work and m == material), None)
    ws.cell(row=r, column=4, value=result)
    r += 1

wb.save(OUT)
