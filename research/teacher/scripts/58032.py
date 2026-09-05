import os

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

# Lookup table: H=Name, I=Title, J=Department Code. For each data row, find
# the person whose Title+Department Code (I,J) matches this row's B,C.
lookup = {}
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=8).value
    title = ws.cell(row=r, column=9).value
    dept = ws.cell(row=r, column=10).value
    if name is not None and title is not None and dept is not None:
        lookup[(title, dept)] = name

orange_fill = PatternFill(fill_type="solid", fgColor="FCD5B4")
gray_fill = PatternFill(fill_type="solid", fgColor="CCCCCC")
thin_border = Border(*(Side(style="thin"),) * 4)

for r in range(2, ws.max_row + 1):
    title = ws.cell(row=r, column=2).value
    dept = ws.cell(row=r, column=3).value
    match = lookup.get((title, dept)) if title is not None and dept is not None else None
    cell_a = ws.cell(row=r, column=1)
    if match is not None:
        cell_a.value = match
        cell_a.fill = orange_fill
    ws.cell(row=r, column=4).fill = gray_fill

# Formatting per instruction: Arial everywhere, bold+italic+bordered header row.
for row in ws.iter_rows():
    for cell in row:
        cell.font = Font(name="Arial", bold=cell.font.bold, italic=cell.font.italic)

for c in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(name="Arial", bold=True, italic=True)
    cell.border = thin_border

ws.sheet_view.showGridLines = False

wb.save(OUT)
