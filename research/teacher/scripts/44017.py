import os
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Data"]
ws_vals = openpyxl.load_workbook(OUT, data_only=True)["Data"]

OLD_START = column_index_from_string("Q")   # P01..P12 old rates
NEW_START = column_index_from_string("AD")  # P01..P12 new rates
N_MONTHS = 12


def add_months(date, n):
    total = date.month - 1 + n
    year = date.year + total // 12
    month = total % 12 + 1
    return date.replace(year=year, month=month)


new_dates = [ws_vals.cell(row=9, column=NEW_START + i).value for i in range(N_MONTHS)]

for row in range(14, ws.max_row + 1):
    freq = ws_vals.cell(row=row, column=column_index_from_string("J")).value
    eff_date = ws.cell(row=row, column=column_index_from_string("L")).value
    increases = [
        ws.cell(row=row, column=column_index_from_string(col)).value
        for col in ("M", "N", "O", "P")
    ]

    valid = isinstance(freq, (int, float)) and isinstance(eff_date, datetime)
    thresholds = None
    if valid:
        thresholds = [add_months(eff_date, int(freq) * k) for k in range(4)]

    for i in range(N_MONTHS):
        cell = ws.cell(row=row, column=NEW_START + i)
        cell.fill = PatternFill(fill_type=None)
        if not valid:
            cell.value = 0
            continue
        old_val = ws.cell(row=row, column=OLD_START + i).value
        old_val = old_val if isinstance(old_val, (int, float)) else 0
        col_date = new_dates[i]
        factor = 1.0
        for k, (thresh, pct) in enumerate(zip(thresholds, increases)):
            if pct is not None and col_date is not None and col_date >= thresh:
                factor += pct
        cell.value = old_val * factor

wb.save(OUT)
