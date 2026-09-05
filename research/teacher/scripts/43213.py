import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

# Instruction: "B12 = average of B11:B12, B13 = average of B11:B13, ..." — a
# growing window whose start is fixed at row 11 (explicit in the instruction)
# and whose source values live in column A (column B is the empty output).
start_row = 11
value_col = 1
target_col = 2

last_row = start_row
for r in range(start_row, ws.max_row + 1):
    if ws.cell(row=r, column=value_col).value is not None:
        last_row = r

for r in range(start_row + 1, last_row + 1):
    vals = [ws.cell(row=rr, column=value_col).value for rr in range(start_row, r + 1)]
    vals = [v for v in vals if v is not None]
    if vals:
        ws.cell(row=r, column=target_col).value = sum(vals) / len(vals)

wb.save(OUT)
