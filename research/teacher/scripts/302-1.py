import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
items = wb["ITEMS"]

# Master value of column B for each data row of the ITEMS sheet.
master = {r: items.cell(row=r, column=2).value for r in range(2, items.max_row + 1)}

for name in ("SHEET1", "SHEET2"):
    ws = wb[name]
    for r in range(2, ws.max_row + 1):
        # skip blank rows in the target sheet; replace B with the ITEMS value at the same row
        if ws.cell(row=r, column=1).value in (None, "") and ws.cell(row=r, column=2).value in (None, ""):
            continue
        if master.get(r) is not None:
            ws.cell(row=r, column=2, value=master[r])

wb.save(OUT)
