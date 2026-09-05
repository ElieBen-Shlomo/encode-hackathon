import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws1 = wb["book1"]
ws2 = wb["book2"]

# Sum columns H:J in book2, grouped by the Place key in column A.
sums = {}
for row in ws2.iter_rows(min_row=2):
    place = row[0].value  # column A
    if place in (None, ""):
        continue
    total = sum(c.value for c in row[7:10] if isinstance(c.value, (int, float)))  # H, I, J
    sums[place] = sums.get(place, 0) + total

row = 3
while ws1.cell(row=row, column=1).value not in (None, ""):
    place = ws1.cell(row=row, column=1).value
    ws1.cell(row=row, column=2).value = sums.get(place, 0)
    row += 1

wb.save(out_path)
