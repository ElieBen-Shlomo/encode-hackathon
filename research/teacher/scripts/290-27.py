import os
import re

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Sheet1"]

# In column B, strip the leading uppercase letter code (e.g. 'PID1' -> 1,
# 'GG 1' -> 1) leaving only the number. Dates, numbers and empty cells stay.
for row in range(1, ws.max_row + 1):
    cell = ws.cell(row=row, column=2)
    value = cell.value
    if isinstance(value, str):
        stripped = re.sub(r"[A-Za-z]", "", value).strip()
        if stripped.isdigit():
            cell.value = int(stripped)

wb.save(os.environ["OUT_XLSX"])
