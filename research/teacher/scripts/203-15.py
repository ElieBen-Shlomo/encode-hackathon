import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
main = wb["Main"]
out = wb["Output Required"]

# Main: A-G = staff info, H = allowance name (8), I = amount (9).
# Group rows per staff (col A), collecting each allowance amount, and record the
# global order in which allowance names first appear (becomes the output columns).
allowances = []
staff_order = []
records = {}
for r in range(2, main.max_row + 1):
    sid = main.cell(r, 1).value
    if sid is None or str(sid).strip() == "":
        continue
    if sid not in records:
        records[sid] = {"info": [main.cell(r, c).value for c in range(1, 8)], "amts": {}}
        staff_order.append(sid)
    name = main.cell(r, 8).value
    records[sid]["amts"][name] = main.cell(r, 9).value
    if name not in allowances:
        allowances.append(name)

# Rebuild the output sheet from scratch.
for row in out.iter_rows():
    for cell in row:
        cell.value = None

header = [main.cell(1, c).value for c in range(1, 8)] + allowances + ["total".capitalize()]
for c, h in enumerate(header, start=1):
    out.cell(1, c, h)

for i, sid in enumerate(staff_order):
    rec = records[sid]
    row = i + 2
    for c, v in enumerate(rec["info"], start=1):
        out.cell(row, c, v)
    total = 0
    for j, name in enumerate(allowances):
        v = rec["amts"].get(name)
        out.cell(row, 8 + j, v)
        if isinstance(v, (int, float)):
            total += v
    out.cell(row, 8 + len(allowances), total)

wb.save(os.environ["OUT_XLSX"])
