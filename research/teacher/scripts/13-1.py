import os
from datetime import datetime

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
src = wb["RANGES"]
dst = wb["LISTS"]


def as_date(v):
    if isinstance(v, datetime):
        return v
    return datetime.strptime(str(v).strip(), "%m/%d/%Y")


# Parse RANGES into sections: title row (col C), header row (A == 'S.N'), then data rows.
sections = []  # (title, [(date, ref, amount), ...])
r = 1
while r <= src.max_row:
    if src.cell(row=r, column=1).value == "S.N":
        title = src.cell(row=r - 1, column=3).value
        rows = []
        r += 1
        while r <= src.max_row and src.cell(row=r, column=1).value is not None:
            rows.append((as_date(src.cell(row=r, column=2).value),
                         src.cell(row=r, column=4).value,
                         src.cell(row=r, column=5).value))
            r += 1
        sections.append((title, rows))
    else:
        r += 1

# Delete the old ranges in LISTS before populating.
for row in dst.iter_rows(min_row=1, max_row=dst.max_row, min_col=1, max_col=dst.max_column):
    for cell in row:
        cell.value = None

# Rebuild each section: combine duplicates on (DATE, REF), sort by date then REF, add TOTAL.
row = 1
for title, rows in sections:
    dst.cell(row=row, column=3, value=title)
    for col, header in enumerate(("SN", "DATE", "REF", "AMOUNTS"), start=1):
        dst.cell(row=row + 1, column=col, value=header)
    groups = {}
    for date, ref, amount in rows:
        groups[(date, ref)] = groups.get((date, ref), 0) + amount
    row += 2
    for sn, (date, ref) in enumerate(sorted(groups), start=1):
        dst.cell(row=row, column=1, value=sn)
        dst.cell(row=row, column=2, value=date)
        dst.cell(row=row, column=3, value=ref)
        dst.cell(row=row, column=4, value=groups[(date, ref)])
        row += 1
    dst.cell(row=row, column=1, value="TOTAL")
    dst.cell(row=row, column=4, value=sum(groups.values()))
    row += 3  # TOTAL row plus two blank separator rows

wb.save(OUT)
