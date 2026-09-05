import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])


def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


# Every time "Line No" appears in column A and the cell directly below it is blank,
# delete both the "Line No" row and the blank row beneath it. Applies to every sheet.
for ws in wb.worksheets:
    hits = []
    for r in range(1, ws.max_row):  # need row r+1 to exist
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip() == "Line No" and is_blank(ws.cell(row=r + 1, column=1).value):
            hits.append(r)
    # Delete the marked pairs from the bottom up so earlier indices stay valid.
    for r in sorted({x for h in hits for x in (h, h + 1)}, reverse=True):
        ws.delete_rows(r, 1)

wb.save(os.environ["OUT_XLSX"])
