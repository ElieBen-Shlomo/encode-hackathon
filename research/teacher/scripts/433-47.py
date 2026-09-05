import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

headers = {c: ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}

# Unit columns in ascending numeric-suffix order (left-to-right add order; reversed for subtract).
unit_cols = []
for c, h in headers.items():
    if isinstance(h, str) and h.strip().lower().startswith("unit"):
        try:
            unit_cols.append((int(h.strip()[4:]), c))
        except ValueError:
            continue
unit_cols = [c for _, c in sorted(unit_cols)]

residual_col = next(c for c, h in headers.items() if isinstance(h, str) and "residual" in h.lower())

min_row = max_row = None
for r in range(1, ws.max_row + 1):
    label = ws.cell(row=r, column=1).value
    if isinstance(label, str) and label.strip().lower() == "min":
        min_row = r
    elif isinstance(label, str) and label.strip().lower() == "max":
        max_row = r

mins = {c: ws.cell(row=min_row, column=c).value for c in unit_cols}
maxs = {c: ws.cell(row=max_row, column=c).value for c in unit_cols}

for r in range(2, ws.max_row + 1):
    residual = ws.cell(row=r, column=residual_col).value
    if not isinstance(residual, (int, float)):
        continue
    remaining = abs(residual)
    order = unit_cols if residual > 0 else list(reversed(unit_cols))
    for c in order:
        if remaining <= 0:
            break
        cur = ws.cell(row=r, column=c).value or 0
        if cur == 0:
            continue
        limit = maxs[c] - cur if residual > 0 else cur - mins[c]
        delta = min(max(limit, 0), remaining)
        if delta <= 0:
            continue
        ws.cell(row=r, column=c).value = cur + delta if residual > 0 else cur - delta
        remaining -= delta
    ws.cell(row=r, column=residual_col).value = 0

wb.save(OUT)
