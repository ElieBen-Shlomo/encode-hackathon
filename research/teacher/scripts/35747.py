import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

start = ws["M1"].value   # inclusive start date
end = ws["N1"].value     # exclusive end date
target = ws["L2"].value  # Asset ID to match

# Columns B..J (2..10) whose header date in row 1 falls in [start, end).
date_cols = [c for c in range(2, 11) if start <= ws.cell(1, c).value < end]
# The row in A2:A11 whose Asset ID matches L2.
asset_row = next(r for r in range(2, 12) if ws.cell(r, 1).value == target)

count = sum(1 for c in date_cols
            if isinstance(ws.cell(asset_row, c).value, (int, float))
            and not isinstance(ws.cell(asset_row, c).value, bool)
            and ws.cell(asset_row, c).value > 100)

ws["M2"] = count
wb.save(os.environ["OUT_XLSX"])
