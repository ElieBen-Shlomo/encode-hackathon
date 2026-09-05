import os

import openpyxl
from openpyxl.styles import Font, PatternFill

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Read the input table with computed values: header "Country | INR | USD", then the
# contiguous rows of user input below it (col C = USD = INR/rate). Stop at the first
# blank so an unrelated table further down the column is not picked up.
src = openpyxl.load_workbook(os.environ["IN_XLSX"], data_only=True).active
hdr = next(r for r in range(1, src.max_row + 1)
           if str(src.cell(row=r, column=1).value).strip() == "Country"
           and str(src.cell(row=r, column=3).value).strip() == "USD")
usd = {}
r = hdr + 1
while r <= src.max_row and src.cell(row=r, column=1).value not in (None, ""):
    v = src.cell(row=r, column=3).value
    key = str(src.cell(row=r, column=1).value).strip()
    usd[key] = usd.get(key, 0) + (v if isinstance(v, (int, float)) else 0)
    r += 1

# Revenue column F: return each country's USD amount (0 when absent/zero), and flag the
# revised cells with the "bad" style (fill #FFC7CE, font #9C0006).
fill = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
font = Font(color="FF9C0006")
for row in range(4, 245):
    country = ws.cell(row=row, column=5).value
    key = str(country).strip() if country is not None else ""
    cell = ws.cell(row=row, column=6, value=usd.get(key, 0))
    cell.fill = fill
    cell.font = font

wb.save(OUT)
