import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["merge"]

# Merge duplicate (A, B, C) combinations, summing D ("Total"), keeping first-seen order.
groups = {}
for r in range(2, ws.max_row + 1):
    a, b, c, d = (ws.cell(row=r, column=col).value for col in (1, 2, 3, 4))
    if a is None and b is None and c is None:
        continue
    key = (a, b, c)
    groups[key] = groups.get(key, 0) + (d or 0)

# Write merged results starting at G2, mirroring the A/B/C/D column names already in G1:J1.
for i, ((a, b, c), total) in enumerate(groups.items(), start=2):
    ws.cell(row=i, column=7, value=a)
    ws.cell(row=i, column=8, value=b)
    ws.cell(row=i, column=9, value=c)
    ws.cell(row=i, column=10, value=total)

wb.save(OUT)
