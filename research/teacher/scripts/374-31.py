import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Find the row whose column A holds the header text "Code".
code_row = None
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v is not None and str(v).strip() == "Code":
        code_row = r
        break

# Delete the blank rows (empty column A) sitting above that row; keep rows below intact.
if code_row:
    for r in range(code_row - 1, 0, -1):
        a = ws.cell(row=r, column=1).value
        if a is None or str(a).strip() == "":
            ws.delete_rows(r)

wb.save(OUT)
