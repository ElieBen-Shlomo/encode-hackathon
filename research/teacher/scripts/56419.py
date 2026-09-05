import os
import openpyxl
from openpyxl.styles import PatternFill

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")

# Column A = Type, column B = Quantity, data starts at row 2.
row = 2
types, quantities = [], []
while ws.cell(row=row, column=1).value not in (None, ""):
    types.append(ws.cell(row=row, column=1).value)
    quantities.append(ws.cell(row=row, column=2).value)
    row += 1
last_row = row - 1

# Ordered list of Types whose Quantity is non-zero.
non_zero_types = [t for t, q in zip(types, quantities) if isinstance(q, (int, float)) and q != 0]

for i, r in enumerate(range(2, last_row + 1)):
    ws.cell(row=r, column=8).value = non_zero_types[i] if i < len(non_zero_types) else None
    ws.cell(row=r, column=8).fill = fill

wb.save(out_path)
