import datetime
import os
import re

import openpyxl
from openpyxl.utils import column_index_from_string

EXCEL_EPOCH = datetime.datetime(1899, 12, 30)


def as_serial(v):
    """Normalize a date-like cell value (datetime or numeric serial) to a number."""
    if isinstance(v, datetime.datetime):
        return (v - EXCEL_EPOCH).days
    return v

OUT = os.environ["OUT_XLSX"]

wb_f = openpyxl.load_workbook(OUT, data_only=False)
wb_v = openpyxl.load_workbook(OUT, data_only=True)
ws_f = wb_f["Total (2)"]
ws_v = wb_v["Total (2)"]

FORMULA_RE = re.compile(
    r'SUMIF\(\$?([A-Z]+)\$?(\d+):\$?[A-Z]+\$?(\d+),\s*"\*([A-Za-z]+)\*"\s*,'
    r'\$?([A-Z]+)\$?(\d+):\$?[A-Z]+\$?(\d+)\)'
)

# Locate each "Yearly Totals - <agent>" block: the 3 rows below the label hold
# per-category (Rentals/Sales/Commercial) SUMIF formulas with no date filter,
# and the 4th row below holds the year-boundary dates (same pattern already
# used correctly in the "Revenue" table earlier in the sheet).
for row in range(1, ws_f.max_row + 1):
    label = ws_f.cell(row=row, column=10).value
    if not label or "Yearly" not in str(label) or "Totals" not in str(label):
        continue
    bounds_row = row + 4
    raw_bounds = [ws_v.cell(row=bounds_row, column=c).value for c in range(11, 16)]
    if not all(isinstance(b, (datetime.datetime, int, float)) for b in raw_bounds):
        continue
    bounds = [as_serial(b) for b in raw_bounds]
    for cat_row in range(row + 1, row + 4):
        formula = ws_f.cell(row=cat_row, column=11).value
        if not isinstance(formula, str):
            continue
        m = FORMULA_RE.search(formula)
        if not m:
            continue
        h_col = column_index_from_string(m.group(1))
        r0, r1 = int(m.group(2)), int(m.group(3))
        agent_code = m.group(4)
        c_col = column_index_from_string(m.group(5))

        for i, col in enumerate(range(11, 16)):
            lo = bounds[i]
            hi = bounds[i + 1] if i + 1 < len(bounds) else None
            total = 0
            for r in range(r0, r1 + 1):
                h = ws_v.cell(row=r, column=h_col).value
                f = ws_v.cell(row=r, column=6).value
                c = ws_v.cell(row=r, column=c_col).value
                if h is None or f is None or c is None:
                    continue
                if agent_code.upper() not in str(h).upper():
                    continue
                f = as_serial(f)
                if f > lo and (hi is None or f < hi):
                    total += c
            ws_f.cell(row=cat_row, column=col).value = total

wb_f.save(OUT)
