import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Column D flags a break-even row (A = 0). Modify it to ignore empty A cells:
# leave D blank until profit data is entered, otherwise 1 if A == 0 else 0.
for row in range(2, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    if a is None or (isinstance(a, str) and a.strip() == ""):
        ws.cell(row=row, column=4).value = None
    else:
        ws.cell(row=row, column=4).value = 1 if a == 0 else 0

wb.save(OUT)
