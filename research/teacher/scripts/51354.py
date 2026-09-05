import calendar
import os

import openpyxl
from openpyxl.styles import PatternFill

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

row = 2
while ws.cell(row=row, column=1).value not in (None, ""):
    text = str(ws.cell(row=row, column=1).value)
    tail = text[-6:].strip()  # e.g. "Oct 21"
    month_str, year_str = tail.split()
    month_num = list(calendar.month_abbr).index(month_str.title())
    year_num = int(year_str)

    new_month = month_num % 12 + 1
    new_year = year_num + (1 if month_num == 12 else 0)

    ws.cell(row=row, column=5).value = f"{calendar.month_abbr[new_month]} {new_year:02d}"
    ws.cell(row=row, column=4).value = year_num
    ws.cell(row=row, column=4).fill = fill
    row += 1

wb.save(out_path)
