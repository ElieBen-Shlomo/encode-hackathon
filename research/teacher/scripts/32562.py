import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Column B = the running order of occurrence of each value in column A:
# the Nth time a number appears in A (top to bottom), B holds N.
seen = {}
for r in range(2, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    if a is None:
        continue
    seen[a] = seen.get(a, 0) + 1
    ws.cell(row=r, column=2).value = seen[a]

wb.save(OUT)
