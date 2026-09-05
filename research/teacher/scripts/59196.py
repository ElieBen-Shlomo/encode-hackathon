import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

# Header row: has text headings across the data columns, with "MAX" at the end.
header_row = None
for r in range(1, ws.max_row + 1):
    if any(str(c.value).strip().upper() == "MAX" for c in ws[r] if c.value is not None):
        header_row = r
        break

max_col = ws.max_column
data_cols = [c for c in range(1, max_col) if ws.cell(row=header_row, column=c).value not in (None, "")]

for row in range(header_row + 1, ws.max_row + 1):
    values = [ws.cell(row=row, column=c).value for c in data_cols]
    if all(v is None for v in values):
        continue
    best_col = max(data_cols, key=lambda c: ws.cell(row=row, column=c).value or 0)
    heading = ws.cell(row=header_row, column=best_col).value
    ws.cell(row=row, column=max_col, value=heading)

wb.save(out_path)
