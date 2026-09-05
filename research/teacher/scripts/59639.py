"""Split each string in column A into its individual characters, one per cell, starting at
column D on the same row (matching the row-4 example)."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

for r in range(5, ws.max_row + 1):
    text = ws.cell(row=r, column=1).value
    if not isinstance(text, str):
        continue
    for i, ch in enumerate(text):
        ws.cell(row=r, column=4 + i).value = ch

wb.save(os.environ["OUT_XLSX"])
