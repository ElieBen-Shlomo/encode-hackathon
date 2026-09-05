import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws1 = wb["worksheet 1"]
ws2 = wb["worksheet2"]

statuses = {}
for row in range(3, ws1.max_row + 1):
    name = ws1.cell(row=row, column=1).value
    status = ws1.cell(row=row, column=2).value
    if name is not None:
        statuses[str(name).strip()] = str(status).strip() if status is not None else None

for row in range(3, ws2.max_row + 1):
    name = ws2.cell(row=row, column=1).value
    status = ws2.cell(row=row, column=2).value
    if name is None:
        continue
    other_status = statuses.get(str(name).strip())
    this_status = str(status).strip() if status is not None else None
    result = "Matched" if other_status == this_status else "Not Matched"
    ws2.cell(row=row, column=3, value=result)

wb.save(os.environ["OUT_XLSX"])
