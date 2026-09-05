import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

mrn = ws.cell(row=2, column=1).value  # A2

result = ""
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=17).value == mrn:  # column Q
        u = ws.cell(row=row, column=21).value  # column U
        y = ws.cell(row=row, column=25).value  # column Y
        if u is not None and y is not None:
            result = (y - u).days
        break

ws.cell(row=2, column=3).value = result  # C2

wb.save(OUT)
