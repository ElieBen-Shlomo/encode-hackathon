import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
sheet1 = wb["Sheet1"]
data = wb["Data"]

# Two-criteria lookup replicating INDEX/MATCH(1, (title=B)*(months=header), 0):
# map the Data header row (months, columns C+) to columns, and the job-title
# column (B, rows 3+) to rows, then intersect for each Sheet1 row.
header_col = {}
for col in range(3, data.max_column + 1):
    v = data.cell(row=2, column=col).value
    if v is not None:
        header_col[v] = col

title_row = {}
for row in range(3, data.max_row + 1):
    v = data.cell(row=row, column=2).value
    if v is not None:
        title_row[v] = row

for row in range(2, sheet1.max_row + 1):
    title = sheet1.cell(row=row, column=1).value
    months = sheet1.cell(row=row, column=2).value
    r = title_row.get(title)
    c = header_col.get(months)
    if r is not None and c is not None:
        sheet1.cell(row=row, column=3).value = data.cell(row=r, column=c).value

wb.save(OUT)
