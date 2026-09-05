import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["master"]
ws_vals = openpyxl.load_workbook(OUT, data_only=True)["master"]

for row in range(2, ws.max_row + 1):
    hours = ws_vals.cell(row=row, column=2).value
    rate = ws_vals.cell(row=row, column=4).value
    if not isinstance(hours, (int, float)) or not isinstance(rate, (int, float)):
        continue
    multiplier = 2 if hours < 37.5 else 1.5
    ws.cell(row=row, column=5).value = rate * multiplier

wb.save(OUT)
