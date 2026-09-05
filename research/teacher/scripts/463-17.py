import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(in_path)
ws = wb["Sheet1"]

row = 2
names = []
while ws.cell(row=row, column=1).value is not None:
    name = ws.cell(row=row, column=1).value
    count = ws.cell(row=row, column=2).value
    names.extend([name] * int(count))
    row += 1

wb_out = openpyxl.load_workbook(out_path)
ws_out = wb_out["Sheet1"]
for i, name in enumerate(names):
    ws_out.cell(row=2 + i, column=4).value = name

wb_out.save(out_path)
