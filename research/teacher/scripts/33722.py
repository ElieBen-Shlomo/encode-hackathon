import os
import statistics

import openpyxl
from openpyxl.styles import Alignment, Border, Side

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Last data row (marketcap in col B).
last = 1
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 2).value not in (None, ""):
        last = r

# E2:E6: for output row k take companies from row k downward whose percentile (col C) is in
# 0..1% (0..0.01), and write the median of their market caps (col B). Each next row excludes
# the rows above it from the selection.
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for k in range(2, 7):
    caps = [ws.cell(r, 2).value for r in range(k, last + 1)
            if isinstance(ws.cell(r, 3).value, (int, float)) and 0 <= ws.cell(r, 3).value <= 0.01]
    cell = ws.cell(k, 5)
    cell.value = statistics.median(caps) if caps else None
    cell.number_format = "0"
    cell.alignment = Alignment(horizontal="left")
    cell.border = border

wb.save(OUT)
