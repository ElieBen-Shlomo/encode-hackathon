import os

import openpyxl
from openpyxl.styles import PatternFill

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Two tables share the same layout; both header rows start with "Pay Item Number".
header_rows = [r for r in range(1, ws.max_row + 1)
               if ws.cell(row=r, column=1).value == "Pay Item Number"]
ref_header, entry_header = header_rows[0], header_rows[1]

# Reference lookup: Description -> (Pay Item Number, UOM) from the first table.
desc_map = {}
for r in range(ref_header + 1, entry_header):
    desc = ws.cell(row=r, column=3).value
    if desc:
        desc_map[str(desc).strip()] = (ws.cell(row=r, column=1).value,
                                       ws.cell(row=r, column=2).value)

# Entry table below the black row spans rows 20-26 (per the instruction).
fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
for r in range(entry_header + 1, 27):
    desc = ws.cell(row=r, column=3).value
    if desc and str(desc).strip() in desc_map:
        item, uom = desc_map[str(desc).strip()]
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=uom)
    else:
        ws.cell(row=r, column=1, value="")
        ws.cell(row=r, column=2, value="")
    ws.cell(row=r, column=1).fill = fill
    ws.cell(row=r, column=2).fill = fill

wb.save(OUT)
