import os
from datetime import datetime

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# The last date is the final entry in column A (not the maximum); B1 gets its
# day-of-month as a static number.
last = None
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if isinstance(v, datetime):
        last = v
ws["B1"] = last.day

wb.save(OUT)
