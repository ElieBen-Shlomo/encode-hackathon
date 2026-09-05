import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

wb_vals = openpyxl.load_workbook(OUT, data_only=True)
ws_vals = wb_vals.active

slip = ws_vals["M5"].value

NOT_FOUND = "not found - please re-enter"
result = NOT_FOUND

if slip not in (None, ""):
    for row in range(6, ws_vals.max_row + 1):
        fm = ws_vals.cell(row=row, column=7).value  # column G
        to = ws_vals.cell(row=row, column=8).value  # column H
        if fm is not None and to is not None and fm <= slip <= to:
            result = ws_vals.cell(row=row, column=10).value  # column J
            break

ws["B2"].value = result

wb.save(OUT)
