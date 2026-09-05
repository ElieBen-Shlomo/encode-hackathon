import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
test = wb["Test"]
master = wb["master"]

YELLOW = "FFFFFF00"
fill = PatternFill(fill_type="solid", fgColor=YELLOW)
border = Border(*(Side(style="thin"),) * 4)

# Collect non-blank, yellow-shaded cells from column A of 'Test'.
values = []
for row in range(1, test.max_row + 1):
    cell = test.cell(row=row, column=1)
    if cell.value is None or str(cell.value).strip() == "":
        continue
    cell_fill = cell.fill
    if cell_fill and cell_fill.fgColor and cell_fill.fgColor.rgb == YELLOW:
        values.append(cell.value)

# Transpose into row 1 of 'master', formatting each with yellow shading + borders.
for i, val in enumerate(values, start=1):
    c = master.cell(row=1, column=i, value=val)
    c.fill = fill
    c.border = border

wb.save(os.environ["OUT_XLSX"])
