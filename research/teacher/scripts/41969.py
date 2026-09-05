import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

SOURCE_ROW, TARGET_ROW, BLOCK = 3, 6, 3

# COUNTBLANK over successive 3-column blocks of row 3 (A:C, D:F, G:I, ...),
# written left-to-right starting at column A of row 6.
n_blocks = -(-ws.max_column // BLOCK)  # ceil division
for i in range(n_blocks):
    start_col = i * BLOCK + 1
    end_col = min(start_col + BLOCK - 1, ws.max_column)
    blanks = sum(1 for c in range(start_col, end_col + 1)
                 if ws.cell(row=SOURCE_ROW, column=c).value in (None, ""))
    ws.cell(row=TARGET_ROW, column=i + 1, value=blanks)

wb.save(OUT)
