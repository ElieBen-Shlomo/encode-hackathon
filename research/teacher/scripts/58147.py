import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb["Here"]
ws_vals = openpyxl.load_workbook(out_path, data_only=True)["Here"]

row = 2
while ws_vals.cell(row=row, column=9).value not in (None, ""):
    text = str(ws_vals.cell(row=row, column=9).value)
    parts = text.split(" - ")
    if len(parts) >= 3:
        name = parts[-2]
    elif len(parts) == 2:
        name = parts[0]
    else:
        name = text
    ws.cell(row=row, column=10, value=name)
    row += 1

wb.save(out_path)
