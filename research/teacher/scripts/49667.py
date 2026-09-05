import os
from datetime import datetime, time

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

HEADER_COL_START = 6  # column F
DEST_COLS = [2, 3, 4, 5]  # B, C, D, E: Start1, Finish1, Start2, Finish2


def to_time(val):
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        return datetime.strptime(val.strip(), "%H:%M").time()
    return None


headers = {}
for col in range(HEADER_COL_START, ws.max_column + 1):
    t = to_time(ws.cell(row=1, column=col).value)
    if t is not None:
        headers[col] = t

for row in range(2, ws.max_row + 1):
    # Find contiguous runs of 'm' marks across the time-block columns.
    blocks = []
    run_start = None
    for col in sorted(headers):
        val = ws.cell(row=row, column=col).value
        is_m = isinstance(val, str) and val.strip().lower() == "m"
        if is_m and run_start is None:
            run_start = col
        elif not is_m and run_start is not None:
            blocks.append((run_start, col - 1))
            run_start = None
    if run_start is not None:
        blocks.append((run_start, max(headers)))

    results = []
    for start_col, end_col in blocks[:2]:
        results.append(headers[start_col])
        results.append(headers[end_col])
    results += [None] * (4 - len(results))

    for dest_col, value in zip(DEST_COLS, results):
        ws.cell(row=row, column=dest_col, value=value)

wb.save(os.environ["OUT_XLSX"])
