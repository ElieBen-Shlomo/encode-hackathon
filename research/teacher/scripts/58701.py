import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
entry = wb["Entry Tab"]
table = wb["Table Tab"]

code_by_location = {
    table.cell(row=r, column=1).value: table.cell(row=r, column=2).value
    for r in range(2, table.max_row + 1)
    if table.cell(row=r, column=1).value is not None
}

for row in range(2, entry.max_row + 1):
    location = entry.cell(row=row, column=2).value  # column B
    if location in code_by_location:
        entry.cell(row=row, column=5).value = code_by_location[location]  # column E

wb.save(OUT)
