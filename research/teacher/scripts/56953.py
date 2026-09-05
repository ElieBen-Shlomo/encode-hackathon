import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet2"]

target = ws["F2"].value

# Column A:D is a series of blocks, each starting with a "Power (kW)" header
# row followed by data rows until the next header. For the requested power
# value, emit each block whose data rows match it (header + matching rows)
# into I:L, stacked in order.
header_rows = [r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Power (kW)"]

blocks = []
for idx, hr in enumerate(header_rows):
    end = header_rows[idx + 1] - 1 if idx + 1 < len(header_rows) else ws.max_row
    blocks.append((hr, list(range(hr + 1, end + 1))))

out_row = 2
for hr, data_rows in blocks:
    matches = [r for r in data_rows if ws.cell(row=r, column=1).value == target]
    if not matches:
        continue
    for c in range(1, 5):
        ws.cell(row=out_row, column=8 + c).value = ws.cell(row=hr, column=c).value
    out_row += 1
    for r in matches:
        for c in range(1, 5):
            ws.cell(row=out_row, column=8 + c).value = ws.cell(row=r, column=c).value
        out_row += 1

wb.save(OUT)
