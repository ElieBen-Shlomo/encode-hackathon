import os

import openpyxl


def norm(v):
    if isinstance(v, (int, float)) and float(v).is_integer():
        return int(v)
    return v


wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
s1 = wb["Sheet1"]
s2 = wb["Sheet2"]

# Build a lookup from Sheet2 keyed by (meet, race#), keeping the FIRST occurrence.
# Sheet2: D=track(4), H=race#(8), I=top marker(9), J..M=finishing order(10..13).
lookup = {}
for r in range(2, s2.max_row + 1):
    track = s2.cell(r, 4).value
    race = s2.cell(r, 8).value
    if track is None or race is None:
        continue
    key = (str(track).strip().upper(), norm(race))
    if key in lookup:
        continue
    positions = {}  # tab number -> finishing rank (J=1, K=2, L=3, M=4)
    for rank, col in enumerate((10, 11, 12, 13), start=1):
        val = s2.cell(r, col).value
        if val is not None and val != "":
            positions[norm(val)] = rank
    lookup[key] = (s2.cell(r, 9).value, positions)

# Fill Sheet1: H = the meet/race top marker, I = the finishing rank of this row's Tab.
# Sheet1: D=meet(4), G=race#(7), H=top(8), I=rank(9), J=tab(10).
for r in range(2, s1.max_row + 1):
    meet = s1.cell(r, 4).value
    race = s1.cell(r, 7).value
    if meet is None or race is None:
        continue
    key = (str(meet).strip().upper(), norm(race))
    if key not in lookup:
        continue
    top, positions = lookup[key]
    s1.cell(r, 8).value = top
    tab = norm(s1.cell(r, 10).value)
    if tab in positions:
        s1.cell(r, 9).value = positions[tab]

wb.save(os.environ["OUT_XLSX"])
