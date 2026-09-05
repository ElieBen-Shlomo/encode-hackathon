import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active


def status(upto_row):
    """Priority order over the expanding A2:A{row}/B2:B{row} range: any Potato paired
    with an actual boolean FALSE beats Tomato-FALSE beats Pickle-FALSE beats "Good"."""
    false_fruits = {ws.cell(r, 1).value for r in range(2, upto_row + 1) if ws.cell(r, 2).value is False}
    if "Potato" in false_fruits:
        return "Worst"
    if "Tomato" in false_fruits:
        return "Ignore"
    if "Pickle" in false_fruits:
        return "Bad"
    return "Good"


for row in range(2, ws.max_row + 1):
    ws.cell(row, 3, status(row))

wb.save(os.environ["OUT_XLSX"])
