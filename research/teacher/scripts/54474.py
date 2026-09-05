"""For each site listed in WHP!C (from row 7 down), look up the matching site row in the
'WHP DATA' sheet (by name in col B) and copy its EEG/Dis/LTU numbers (cols C/D/E) into
WHP!E:G on the same row."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws_whp = wb["WHP"]
ws_data = wb["WHP DATA"]

lookup = {}
for r in range(1, ws_data.max_row + 1):
    name = ws_data.cell(row=r, column=2).value
    eeg, dis, ltu = (ws_data.cell(row=r, column=c).value for c in (3, 4, 5))
    if isinstance(name, str) and isinstance(eeg, (int, float)):
        lookup[name] = (eeg, dis, ltu)

row = 7
while True:
    site = ws_whp.cell(row=row, column=3).value
    if site is None:
        break
    if site in lookup:
        eeg, dis, ltu = lookup[site]
        ws_whp.cell(row=row, column=5).value = eeg
        ws_whp.cell(row=row, column=6).value = dis
        ws_whp.cell(row=row, column=7).value = ltu
    row += 1

wb.save(os.environ["OUT_XLSX"])
