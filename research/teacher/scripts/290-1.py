import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# Collect (number, alphagram) pairs from columns I and H, skipping the header row.
pairs = []
for row in range(2, ws.max_row + 1):
    num = ws.cell(row=row, column=9).value
    text = ws.cell(row=row, column=8).value
    if num is not None and text is not None:
        pairs.append((num, text))

# Lay out four pairs per row into column groups K/L, N/O, Q/R, T/U starting at row 2.
group_cols = (11, 14, 17, 20)
for idx, (num, text) in enumerate(pairs):
    row = 2 + idx // 4
    col = group_cols[idx % 4]
    ws.cell(row=row, column=col, value=num)
    ws.cell(row=row, column=col + 1, value=text)

wb.save(OUT)
