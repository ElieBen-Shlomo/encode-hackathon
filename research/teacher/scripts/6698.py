import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["raw 1"]

# Count rows where B or C is nonzero, restricted to rows whose A matches the
# class in the corresponding F cell.
counts = {}
for r in range(2, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    b = ws.cell(row=r, column=2).value
    c = ws.cell(row=r, column=3).value
    if a is None:
        continue
    if (b not in (0, None)) or (c not in (0, None)):
        counts[a] = counts.get(a, 0) + 1

for r in range(4, 9):
    target = ws.cell(row=r, column=6).value
    ws.cell(row=r, column=7).value = counts.get(target, 0)

ws.auto_filter.ref = "A1:C1"

wb.save(OUT)
