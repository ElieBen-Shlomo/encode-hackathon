import os
from collections import defaultdict

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

rows = range(2, ws.max_row + 1)

statuses_by_mrn = defaultdict(set)
for r in rows:
    mrn = ws.cell(row=r, column=1).value
    status = str(ws.cell(row=r, column=4).value or "").strip().lower()
    statuses_by_mrn[mrn].add(status)

for r in rows:
    mrn = ws.cell(row=r, column=1).value
    statuses = statuses_by_mrn[mrn]
    has_sch = "sch" in statuses
    has_no_show = "no show" in statuses

    if has_sch and has_no_show:
        result = "NS/SCHED"
    elif has_sch:
        result = "FUTURE"
    elif has_no_show:
        next_date = ws.cell(row=r, column=5).value
        result = "NO ACTION NEEDED" if next_date not in (None, "") else "CALL PT"
    else:
        result = ""

    ws.cell(row=r, column=6).value = result

wb.save(OUT)
