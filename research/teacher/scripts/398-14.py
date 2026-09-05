import os

import openpyxl
from openpyxl.styles import Alignment, Font

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
coll = next(wb[s] for s in wb.sheetnames if s.upper() == "COLLECTION")
sources = [wb[s] for s in wb.sheetnames if wb[s] is not coll]


def cols(ws):
    return {str(ws.cell(1, c).value).strip().upper(): c
            for c in range(1, ws.max_column + 1) if ws.cell(1, c).value not in (None, "")}


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


# Aggregate SALE/RET by (TY, OR) across every non-COLLECTION sheet, whatever their count/columns.
totals, br_of, order = {}, {}, []
for ws in sources:
    c = cols(ws)
    ty_c, or_c = c.get("TY"), c.get("OR")
    if not ty_c or not or_c:
        continue
    for r in range(2, ws.max_row + 1):
        ty, orr = ws.cell(r, ty_c).value, ws.cell(r, or_c).value
        if ty in (None, "") or orr in (None, ""):
            continue
        key = (ty, orr)
        if key not in totals:
            totals[key] = [0.0, 0.0]
            br_of[key] = ws.cell(r, c["BR"]).value if "BR" in c else None
            order.append(key)
        if "SALE" in c:
            totals[key][0] += num(ws.cell(r, c["SALE"]).value)
        if "RET" in c:
            totals[key][1] += num(ws.cell(r, c["RET"]).value)

cc = cols(coll)
item_c, br_c, ty_c, or_c, sale_c, ret_c, bal_c = (cc["ITEM"], cc["BR"], cc["TY"], cc["OR"],
                                                   cc["SALE"], cc["RET"], cc["BALANCE"])

# Keep existing rows/order as-is (they're already the reference), just turn BALANCE into a value.
existing, last_row, max_item = set(), 1, 0
for r in range(2, coll.max_row + 1):
    ty, orr = coll.cell(r, ty_c).value, coll.cell(r, or_c).value
    if ty in (None, ""):
        continue
    existing.add((ty, orr))
    last_row = r
    max_item = max(max_item, num(coll.cell(r, item_c).value))
    coll.cell(r, bal_c).value = num(coll.cell(r, sale_c).value) - num(coll.cell(r, ret_c).value)

# Append combos present in the source sheets but missing from COLLECTION, to the bottom.
r = last_row
for key in order:
    if key in existing:
        continue
    r += 1
    max_item += 1
    sale, ret = totals[key]
    coll.cell(r, item_c).value = max_item
    coll.cell(r, br_c).value = br_of[key]
    coll.cell(r, ty_c).value = key[0]
    coll.cell(r, or_c).value = key[1]
    coll.cell(r, sale_c).value = sale
    coll.cell(r, ret_c).value = ret
    coll.cell(r, bal_c).value = sale - ret
last_row = r

# Formatting per the instruction: un-bold left-aligned headers, unbold Calibri 11 data with
# A/E/F/G right and B/C/D left aligned, zero shown as a hyphen, negatives in red.
for c in range(1, coll.max_column + 1):
    coll.cell(1, c).font = Font(name="Calibri", size=11, bold=False)
    coll.cell(1, c).alignment = Alignment(horizontal="left")
for row in range(2, last_row + 1):
    for c in range(1, coll.max_column + 1):
        cell = coll.cell(row, c)
        red = c == bal_c and isinstance(cell.value, (int, float)) and cell.value < 0
        cell.font = Font(name="Calibri", size=11, bold=False, color="FFFF0000" if red else None)
        cell.alignment = Alignment(horizontal="left" if c in (br_c, ty_c, or_c) else "right")
        if c in (sale_c, ret_c, bal_c):
            cell.number_format = r'0;\-0;\-'

wb.save(OUT)
