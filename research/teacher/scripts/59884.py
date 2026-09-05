import os
from collections import defaultdict

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

# Build the map: first-value -> ordered list of second-values (columns A, B).
values_by_first = defaultdict(list)
for row in range(2, ws.max_row + 1):
    first = ws.cell(row=row, column=1).value
    second = ws.cell(row=row, column=2).value
    if first is not None and second is not None:
        values_by_first[first].append(second)

# Column G holds the unique first-values already; transpose the matching
# second-values into columns I onward on the same row.
for row in range(2, ws.max_row + 1):
    first = ws.cell(row=row, column=7).value  # column G
    if first is None:
        continue
    for offset, value in enumerate(values_by_first.get(first, [])):
        ws.cell(row=row, column=9 + offset).value = value  # column I onward

wb.save(OUT)
