"""Delete rows where the first character of column B is not 'H' or 'A' (keep the header row).
Compacts kept rows to the top, then trims the leftover tail in one delete_rows call --
avoids doing thousands of single-row deletes on a 10k+ row sheet."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active


def keep(row: int) -> bool:
    v = ws.cell(row=row, column=2).value
    if v in (None, ""):
        return False
    return str(v).strip()[:1].upper() in ("H", "A")


max_row, max_col = ws.max_row, ws.max_column
kept_rows = [r for r in range(2, max_row + 1) if keep(r)]

for new_r, old_r in enumerate(kept_rows, start=2):
    if new_r != old_r:
        for c in range(1, max_col + 1):
            ws.cell(row=new_r, column=c).value = ws.cell(row=old_r, column=c).value

first_excess = 2 + len(kept_rows)
if first_excess <= max_row:
    ws.delete_rows(first_excess, max_row - first_excess + 1)

wb.save(os.environ["OUT_XLSX"])
