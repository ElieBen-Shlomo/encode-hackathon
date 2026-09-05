import datetime
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

ABBR = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# Build the public-holiday map: date -> Type, from the reference list (col I date, col J type).
holidays = {}
for r in range(2, ws.max_row + 1):
    d = ws.cell(r, 9).value
    ty = ws.cell(r, 10).value
    if isinstance(d, datetime.datetime) and ty not in (None, ""):
        holidays[d.date()] = ty

# WORKDAY (col E): holiday Type when the date is a public holiday, else the weekday abbreviation.
for r in range(2, ws.max_row + 1):
    d = ws.cell(r, 1).value
    if not isinstance(d, datetime.datetime):
        continue
    key = d.date()
    ws.cell(r, 5).value = holidays[key] if key in holidays else ABBR[d.weekday()]

wb.save(OUT)
