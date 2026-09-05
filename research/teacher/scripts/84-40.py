"""Rebuild the LIST sheet from scratch: for every other sheet in the workbook, sum its
IMPORT (col C) and EXPORT (col D) columns, list the sheet name with those totals and their
balance (C-D), then add a TOTAL row summing across all sheets. Rebuilding from scratch
keeps LIST correct if sheets are added, removed, or reordered."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
list_ws = wb["LIST"]
sheet_names = [name for name in wb.sheetnames if name != "LIST"]


def col_sum(ws, col):
    total = 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=col).value
        if isinstance(val, (int, float)):
            total += val
    return total


# Clear any previously generated rows below the header.
for r in range(2, list_ws.max_row + 1):
    for c in range(1, 6):
        list_ws.cell(row=r, column=c).value = None

total_c = total_d = 0
row = 2
for idx, name in enumerate(sheet_names, start=1):
    ws = wb[name]
    c_sum = col_sum(ws, 3)
    d_sum = col_sum(ws, 4)
    total_c += c_sum
    total_d += d_sum
    list_ws.cell(row=row, column=1).value = idx
    list_ws.cell(row=row, column=2).value = name
    list_ws.cell(row=row, column=3).value = c_sum
    list_ws.cell(row=row, column=4).value = d_sum
    list_ws.cell(row=row, column=5).value = c_sum - d_sum
    row += 1

list_ws.cell(row=row, column=2).value = "TOTAL"
list_ws.cell(row=row, column=3).value = total_c
list_ws.cell(row=row, column=4).value = total_d
list_ws.cell(row=row, column=5).value = total_c - total_d

wb.save(os.environ["OUT_XLSX"])
