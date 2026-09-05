import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Sequentially number each non-empty code in column B (skipping the random gaps),
# starting at 010 and adding +1 each time, then combine it with the prefix in B2
# as <B2>-02-<3-digit number> (matching the example already shown in column E).
prefix = ws["B2"].value
count = 0
for row in range(3, ws.max_row + 1):
    code = ws.cell(row=row, column=2).value
    if code is None or str(code).strip() == "":
        continue
    count += 1
    ws.cell(row=row, column=1, value=f"{prefix}-02-{9 + count:03d}")

wb.save(OUT)
