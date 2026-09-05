"""For each data column, sum the lengths of runs of consecutive 1's that are 2 or longer
(runs of a single isolated 1 don't count), writing the totals into the 'Ans' row."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

ans_row = None
for r in range(1, ws.max_row + 1):
    if str(ws.cell(row=r, column=1).value).strip().lower() == "ans":
        ans_row = r
        break

cols = range(2, ws.max_column + 1)

data_rows = []
r = ans_row - 1
while r >= 1:
    row_vals = [ws.cell(row=r, column=c).value for c in cols]
    if all(v is None for v in row_vals):
        break
    data_rows.append(r)
    r -= 1
data_rows.reverse()

for c in cols:
    values = [ws.cell(row=r, column=c).value for r in data_rows]
    total, run = 0, 0
    for v in values:
        if v == 1:
            run += 1
        else:
            if run >= 2:
                total += run
            run = 0
    if run >= 2:
        total += run
    ws.cell(row=ans_row, column=c).value = total

wb.save(os.environ["OUT_XLSX"])
