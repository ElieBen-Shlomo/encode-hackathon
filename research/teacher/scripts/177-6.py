import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
data = wb["DATA"]
comb = wb["combined"]

KEY_COL = 8   # column H is the reference column to merge on
NCOLS = 18    # columns A:R

# Group DATA rows by the reference column, preserving first-seen order.
order, groups = [], {}
for r in range(2, data.max_row + 1):
    key = data.cell(row=r, column=KEY_COL).value
    if key is None or (isinstance(key, str) and not key.strip()):
        continue
    if key not in groups:
        groups[key] = []
        order.append(key)
    groups[key].append(r)

# Clear existing data rows in 'combined' (keep row 1 for the header).
for row in comb.iter_rows(min_row=2, max_row=comb.max_row, max_col=NCOLS):
    for cell in row:
        cell.value = None

# Header row copied from DATA row 1.
for c in range(1, NCOLS + 1):
    src = data.cell(row=1, column=c)
    dst = comb.cell(row=1, column=c, value=src.value)
    dst.number_format = src.number_format

# One consolidated row per group.
for i, key in enumerate(order):
    out_row = 2 + i
    rows = groups[key]
    # Columns A:H are identical within a group -> copy from its first row.
    for c in range(1, KEY_COL + 1):
        src = data.cell(row=rows[0], column=c)
        dst = comb.cell(row=out_row, column=c, value=src.value)
        dst.number_format = src.number_format
    # Columns I:R -> sum numeric values; a 0 total is blank; two decimals.
    for c in range(KEY_COL + 1, NCOLS + 1):
        total = sum(data.cell(row=r, column=c).value for r in rows
                    if isinstance(data.cell(row=r, column=c).value, (int, float)))
        dst = comb.cell(row=out_row, column=c, value=total if total != 0 else None)
        dst.number_format = "0.00"

wb.save(OUT)
