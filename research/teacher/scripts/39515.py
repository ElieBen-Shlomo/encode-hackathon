import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Month header (JAN..DEC) -> column index, from row 1 across C:N.
month_col = {}
for col in range(3, 15):
    h = ws.cell(row=1, column=col).value
    if h:
        month_col[str(h).strip().upper()] = col

# The latest year is the "current" one; earlier years fall back to January.
years = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
current_year = max(y for y in years if isinstance(y, (int, float)))

for row in range(2, ws.max_row + 1):
    month = ws.cell(row=row, column=1).value
    year = ws.cell(row=row, column=2).value
    if year == current_year:
        col = month_col.get(str(month).strip().upper())
    else:
        col = month_col.get("JAN")
    if col:
        ws.cell(row=row, column=15, value=ws.cell(row=row, column=col).value)

wb.save(OUT)
