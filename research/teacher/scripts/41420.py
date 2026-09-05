import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

OUT_COLS = 5  # C:G

r = 2
while ws.cell(r, 1).value is not None:
    digits = sorted(set(ch for ch in str(ws.cell(r, 1).value) if ch.isdigit()))
    for i in range(OUT_COLS):
        ws.cell(r, 3 + i, int(digits[i]) if i < len(digits) else None)
    r += 1

wb.save(OUT)
