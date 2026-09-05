import os
import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Build model -> qty total from columns C (model) and D (qty), rows 4..last data row.
totals = {}
for row in range(4, ws.max_row + 1):
    model = ws.cell(row, 3).value
    qty = ws.cell(row, 4).value
    if model is None:
        continue
    totals[model] = totals.get(model, 0) + (qty or 0)

# Fill column G for each model listed in column F.
for row in range(4, ws.max_row + 1):
    model = ws.cell(row, 6).value
    if model is None:
        continue
    ws.cell(row, 7).value = totals.get(model, 0)

wb.save(OUT)
