import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Sheet1"]

ws["H1"] = "Debits"
ws["I1"] = "Credits"

for row in range(2, ws.max_row + 1):
    amount = ws.cell(row=row, column=3).value
    if not isinstance(amount, (int, float)):
        continue
    if amount > 0:
        ws.cell(row=row, column=8, value=abs(amount))
    elif amount < 0:
        ws.cell(row=row, column=9, value=abs(amount))

wb.save(os.environ["OUT_XLSX"])
