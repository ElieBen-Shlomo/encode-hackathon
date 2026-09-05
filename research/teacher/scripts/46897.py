import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

target_header = ws["I1"].value
target_value = ws["I2"].value

# Find which column (row 1, before the summary block) has this header.
target_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == target_header:
        target_col = col
        break

count = 0
if target_col is not None:
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=target_col).value == target_value:
            count += 1

ws["I3"] = count

wb.save(OUT)
