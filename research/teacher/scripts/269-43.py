import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]


def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


# Delete any row that has a value in column I but a blank column H.
# Iterate bottom-up so earlier row indices stay valid after each deletion.
for r in range(ws.max_row, 0, -1):
    has_i = not is_blank(ws.cell(row=r, column=9).value)
    blank_h = is_blank(ws.cell(row=r, column=8).value)
    if has_i and blank_h:
        ws.delete_rows(r, 1)

wb.save(OUT)
