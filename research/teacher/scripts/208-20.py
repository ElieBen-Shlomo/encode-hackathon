import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Q"]

FIRST = 3          # first data row of the left-hand table
NCOL = 5           # columns A-E hold the reorderable values; column F is a per-row
LAST = ws.max_row  # formula (=IF(...A{r})) that must stay in place and recalc itself

# Group contiguous rows by ticker (col A). The days number (col E) appears once per group.
groups = []
for r in range(FIRST, LAST + 1):
    vals = [ws.cell(r, c).value for c in range(1, NCOL + 1)]
    ticker = vals[0]
    if ticker is None:
        continue
    if groups and groups[-1][0] == ticker:
        groups[-1][1].append(vals)
    else:
        groups.append((ticker, [vals]))


def days(group):
    nums = [row[4] for row in group[1] if isinstance(row[4], (int, float))]
    return max(nums) if nums else float("-inf")


# Order groups from highest to lowest; stable sort keeps ties in their original order.
groups.sort(key=days, reverse=True)

# Write the reordered A-E values back, leaving the column F formulas untouched.
r = FIRST
for _, rows in groups:
    for vals in rows:
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c, v)
        r += 1

wb.save(os.environ["OUT_XLSX"])
