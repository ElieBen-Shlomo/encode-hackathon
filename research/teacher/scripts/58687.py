import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb["Sheet1"]
ws_vals = openpyxl.load_workbook(out_path, data_only=True)["Sheet1"]

# Row 4 holds month numbers, row 6 holds cumulative apples-sold totals,
# starting at column C. Both are formulas, so read the cached values.
months = []
col = 3
while ws_vals.cell(row=4, column=col).value is not None:
    month = ws_vals.cell(row=4, column=col).value
    accum = ws_vals.cell(row=6, column=col).value
    months.append((month, accum))
    col += 1

row = 9
while ws_vals.cell(row=row, column=1).value is not None:
    apple_index = ws_vals.cell(row=row, column=1).value
    # First month whose cumulative total reaches this apple's index.
    for month, accum in months:
        if accum >= apple_index:
            ws.cell(row=row, column=2).value = month
            break
    row += 1

wb.save(out_path)
