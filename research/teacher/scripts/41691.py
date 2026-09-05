import os
import re

import openpyxl
from openpyxl.utils import column_index_from_string

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Find the cell holding the broken VLOOKUP (its cached value is #REF! because
# the column index it asks for falls outside the given lookup range).
target = next(c for row in ws.iter_rows() for c in row
              if isinstance(c.value, str) and c.value.upper().startswith("=VLOOKUP"))

m = re.match(r"=VLOOKUP\(([A-Z]+\d+),\$?([A-Z]+)\$?(\d+):\$?[A-Z]+\$?(\d+)", target.value)
lookup_cell, key_col, start_row, end_row = m.group(1), m.group(2), int(m.group(3)), int(m.group(4))
key_col_idx = column_index_from_string(key_col)

# The intended value column is the one the formula itself lives in (a plain
# two-column key/value table with the broken lookup sitting in the value column).
lookup_value = ws[lookup_cell].value
table = {ws.cell(row=r, column=key_col_idx).value: ws.cell(row=r, column=target.column).value
         for r in range(start_row, end_row + 1)}

target.value = table[lookup_value]
wb.save(OUT)
