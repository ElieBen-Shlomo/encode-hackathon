import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb["Sheet1"]

# Determine the holidays range in column J (starts at J5, runs until blank).
r = 5
while ws.cell(row=r, column=10).value is not None:
    r += 1
last_holiday_row = r - 1

# Fill H7:H38 with a WORKDAY formula counting forward from B3, skipping
# weekends and the holidays listed in column J. ROWS($B$7:Bn) grows by one
# each row down, giving the correct number of workdays to add (0 on the
# first row so it returns the start date itself).
row = 7
while ws.cell(row=row, column=8).value is not None:
    formula = (
        f"=WORKDAY($B$3,ROWS($B$7:B{row})-1,"
        f"$J$5:$J${last_holiday_row})"
    )
    ws.cell(row=row, column=8).value = formula
    row += 1

wb.save(out_path)
