import os
import re

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Before"]

# In each column A cell, the substring starting at the first alphabetic letter is the
# "event_type=..." text: move it to column B and keep the leading part (timestamp) in A.
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if not isinstance(v, str):
        continue
    m = re.search(r"[A-Za-z]", v)
    if not m or m.start() == 0:
        continue
    idx = m.start()
    ws.cell(r, 1).value = v[:idx].rstrip()
    ws.cell(r, 2).value = v[idx:]

wb.save(os.environ["OUT_XLSX"])
