import os
import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

START_COL = 3  # column C
STEP = 4

# Table rows: column A holds a code, used as the lookup key.
table = {}
for row in range(1, ws.max_row + 1):
    code = ws.cell(row, 1).value
    if code is not None:
        table[code] = row

# Query rows: column A blank, column B holds the code to look up.
for row in range(1, ws.max_row + 1):
    if ws.cell(row, 1).value is not None:
        continue
    key = ws.cell(row, 2).value
    if key is None or key not in table:
        continue
    src_row = table[key]
    total = 0
    for col in range(START_COL, ws.max_column + 1, STEP):
        val = ws.cell(src_row, col).value
        if val is not None:
            total += val
    ws.cell(row, 3).value = total

wb.save(OUT)
