import os
import openpyxl
from openpyxl.utils import column_index_from_string

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

current_week = ws.cell(2, 2).value
first_col = column_index_from_string("D")
target_col = column_index_from_string("BE")

for row in range(3, ws.max_row + 1):
    if ws.cell(row, 3).value is None:
        continue
    total = 0
    for col in range(first_col, target_col):
        week = ws.cell(2, col).value
        if week is None:
            continue
        val = ws.cell(row, col).value
        if val is not None and week >= current_week:
            total += val
    ws.cell(row, target_col).value = total

wb.save(OUT)
