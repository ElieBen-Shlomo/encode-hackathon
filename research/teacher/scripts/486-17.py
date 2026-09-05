import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Blad1"]

# Column A holds '0yyyymmdd' (9 chars, leading digit dropped). Reformat into
# 'yyyy mm dd' by taking fixed-width slices, equivalent to
# =MID(A,2,4)&" "&MID(A,6,2)&" "&MID(A,8,2). B1 stays blank (no header).
for row in range(2, ws.max_row + 1):
    raw = ws.cell(row=row, column=1).value
    if raw is None:
        continue
    s = str(raw)
    ws.cell(row=row, column=2).value = f"{s[1:5]} {s[5:7]} {s[7:9]}"

wb.save(os.environ["OUT_XLSX"])
