import os
import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
extract = wb["Extract"]
result = wb["Desired Result"]

# Build (person, car) -> value from the Extract sheet.
values = {}
for row in range(2, extract.max_row + 1):
    person = extract.cell(row, 1).value
    car = extract.cell(row, 2).value
    if person is None or car is None:
        continue
    values[(person, car)] = extract.cell(row, 3).value

headers = {col: result.cell(1, col).value for col in range(2, result.max_column + 1)}

for row in range(2, result.max_row + 1):
    person = result.cell(row, 1).value
    if person is None:
        continue
    for col, car in headers.items():
        result.cell(row, col).value = values.get((person, car))

wb.save(OUT)
