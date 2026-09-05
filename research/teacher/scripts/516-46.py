import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

# Source table: A:E starting row 2 (date, brand, batch, origin, quantity).
rows = []
r = 2
while ws.cell(row=r, column=1).value is not None:
    rows.append([ws.cell(row=r, column=c).value for c in range(1, 6)])
    r += 1

order = []
last_date = {}
for date, brand, _, _, _ in rows:
    if brand not in order:
        order.append(brand)
    if brand not in last_date or date > last_date[brand]:
        last_date[brand] = date

# For each brand, sum quantity across every entry on that brand's last date;
# batch/origin come from the first such matching entry.
last_row = {}
for date, brand, batch, origin, qty in rows:
    if date != last_date[brand]:
        continue
    if brand not in last_row:
        last_row[brand] = [date, brand, batch, origin, qty]
    else:
        last_row[brand][4] += qty

out_row = 2
for brand in order:
    date, b, batch, origin, qty = last_row[brand]
    ws.cell(row=out_row, column=8).value = date
    ws.cell(row=out_row, column=9).value = b
    ws.cell(row=out_row, column=10).value = batch
    ws.cell(row=out_row, column=11).value = origin
    ws.cell(row=out_row, column=12).value = qty
    out_row += 1

wb.save(OUT)
