import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
min_col = next(c for c, h in enumerate(headers, start=1)
               if isinstance(h, str) and h.strip().lower().startswith("min price"))
seller_col = next(c for c, h in enumerate(headers, start=1)
                   if isinstance(h, str) and h.strip().lower() == "seller")
price_cols = range(2, min_col)  # seller-price columns between product name and the min column

# For each row, find the smallest value greater than zero among the seller-price columns
# and write the comma-separated header(s) of the column(s) achieving that minimum.
for r in range(2, ws.max_row + 1):
    candidates = [(c, ws.cell(r, c).value) for c in price_cols]
    positive = [(c, v) for c, v in candidates if isinstance(v, (int, float)) and v > 0]
    if not positive:
        continue
    lowest = min(v for _, v in positive)
    winners = [headers[c - 1] for c, v in positive if v == lowest]
    ws.cell(r, seller_col).value = ",".join(str(w) for w in winners)

wb.save(OUT)
