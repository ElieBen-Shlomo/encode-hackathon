import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Find the first row whose column A contains "Invoice No." (case-insensitive),
# then delete every row above it, leaving it and everything below untouched.
target = "invoice no."
first_row = None
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None and target in str(v).strip().lower():
        first_row = r
        break

if first_row and first_row > 1:
    ws.delete_rows(1, first_row - 1)

wb.save(OUT)
