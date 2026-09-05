import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Sheet1"]

# Collect the word list from column A
words = []
for row in range(1, ws.max_row + 1):
    v = ws.cell(row=row, column=1).value
    if isinstance(v, str) and v.strip():
        words.append(v.strip())

# Sort the names in column A alphabetically
words.sort()
for i, w in enumerate(words, start=1):
    ws.cell(row=i, column=1).value = w

word_set = set(words)

# A word matches when moving its fifth letter to the front (EARTHPEA -> HEARTPEA)
# produces another word that is also in the list; last three letters stay the same.
pairs = []  # (transformation, original)
for w in words:
    if len(w) >= 5:
        t = w[4] + w[:4] + w[5:]
        if t != w and t in word_set:
            pairs.append((t, w))

# Group results by their endings in the requested order, alphabetical within group
ending_order = ["ING", "ERS", "ATE", "EST", "ONE", "IER", "ILY"]

def sort_key(pair):
    end = pair[1][-3:]
    idx = ending_order.index(end) if end in ending_order else len(ending_order)
    return (idx, end, pair[1])

pairs.sort(key=sort_key)

# Paste pairs at columns C (transformation) and D (original) starting below the example
r = 2
for t, w in pairs:
    ws.cell(row=r, column=3).value = t
    ws.cell(row=r, column=4).value = w
    r += 1

wb.save(os.environ["OUT_XLSX"])
