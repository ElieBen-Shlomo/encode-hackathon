import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)

main_ws = wb["Main unique ID"]
getting_ws = wb["Result what i am getting"]
NCOLS = 15  # columns A:O

# Master list of IDs in the order given by 'Main unique ID'.
main_ids = [main_ws.cell(row=r, column=1).value
            for r in range(1, main_ws.max_row + 1)
            if main_ws.cell(row=r, column=1).value is not None]

# Existing rows keyed by their ID (column A), keeping the full A:O payload.
header = [getting_ws.cell(row=1, column=c).value for c in range(1, NCOLS + 1)]
existing = {}
for r in range(2, getting_ws.max_row + 1):
    uid = getting_ws.cell(row=r, column=1).value
    if uid is not None:
        existing[uid] = [getting_ws.cell(row=r, column=c).value for c in range(1, NCOLS + 1)]

# Build MyResult: header, then every master ID. IDs missing from 'getting' get an
# otherwise-empty new row carrying only the ID in column A.
if "MyResult" in wb.sheetnames:
    del wb["MyResult"]
result = wb.create_sheet("MyResult", index=1)
result.append(header)
for uid in main_ids:
    if uid in existing:
        result.append(existing[uid])
    else:
        row = [None] * NCOLS
        row[0] = uid
        result.append(row)

wb.active = wb.sheetnames.index("MyResult")
wb.save(OUT)
