import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

first_row, last_row = 2, ws.max_row
wins = {row: 0 for row in range(first_row, last_row + 1)}

# Each game is a column (B onward, up to but excluding the win-count column);
# the player with the highest score in that column gets a win.
for col in range(2, ws.max_column):
    best_row, best_score = None, None
    for row in range(first_row, last_row + 1):
        score = ws.cell(row=row, column=col).value
        if isinstance(score, (int, float)) and (best_score is None or score > best_score):
            best_row, best_score = row, score
    if best_row is not None:
        wins[best_row] += 1

for row in range(first_row, last_row + 1):
    ws.cell(row=row, column=ws.max_column).value = wins[row]

wb.save(OUT)
