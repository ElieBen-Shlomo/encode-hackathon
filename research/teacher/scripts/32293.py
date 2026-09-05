import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb.active

# The "OUTPUT" marker in column A separates the reference table (above) from the
# rows to populate (below).
output_row = next(r for r in range(1, ws.max_row + 1)
                  if str(ws.cell(row=r, column=1).value).strip().upper() == "OUTPUT")

# Build key -> value map from the reference table above the marker.
lookup = {}
for r in range(1, output_row):
    key = ws.cell(row=r, column=1).value
    val = ws.cell(row=r, column=2).value
    if key is not None and val is not None:
        lookup[str(key).strip()] = val

# Populate column B for each labelled row below the marker.
for r in range(output_row + 1, ws.max_row + 1):
    key = ws.cell(row=r, column=1).value
    if key is not None and str(key).strip() in lookup:
        ws.cell(row=r, column=2).value = lookup[str(key).strip()]

wb.save(OUT)
