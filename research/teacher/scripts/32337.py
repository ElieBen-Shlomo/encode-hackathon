import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Column O ('age (year)') holds ROUND formulas; read their cached values.
data = openpyxl.load_workbook(os.environ["IN_XLSX"], data_only=True).active

# Build the age -> CATEGORY lookup from the YEAR(age) / CATEGORY table (cols H, I).
lookup = {}
for r in range(1, data.max_row + 1):
    age = data.cell(row=r, column=8).value
    cat = data.cell(row=r, column=9).value
    if isinstance(age, (int, float)) and cat is not None:
        lookup[round(age)] = cat

# Fill 'Expected Result' (col E): look up each row's rounded age (col O) in the table.
for r in range(3, ws.max_row + 1):
    age = data.cell(row=r, column=15).value
    if isinstance(age, (int, float)) and round(age) in lookup:
        ws.cell(row=r, column=5).value = lookup[round(age)]

wb.save(OUT)
