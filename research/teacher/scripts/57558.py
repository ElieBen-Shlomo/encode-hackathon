import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
hurdles = wb["RateHurdles"]
deposits = wb["Deposits"]

rules = []
r = 2
while hurdles.cell(row=r, column=1).value is not None:
    start = hurdles.cell(row=r, column=1).value
    end = hurdles.cell(row=r, column=2).value
    sale_type = hurdles.cell(row=r, column=3).value
    person = hurdles.cell(row=r, column=4).value
    rate = hurdles.cell(row=r, column=5).value
    rules.append((start, end, sale_type, person, rate))
    r += 1

row = 2
while deposits.cell(row=row, column=2).value is not None:
    sale_type = deposits.cell(row=row, column=2).value
    person = deposits.cell(row=row, column=3).value
    date = deposits.cell(row=row, column=4).value
    for start, end, r_type, r_person, rate in rules:
        if r_type == sale_type and r_person == person and start <= date <= end:
            deposits.cell(row=row, column=1).value = rate
            break
    row += 1

wb.save(out_path)
