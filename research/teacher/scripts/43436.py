import calendar
import datetime
import os
import re

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

headers = {c: ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}


def find_col(name):
    return next(c for c, h in headers.items() if isinstance(h, str) and h.strip().lower() == name)


created_col = find_col("date created")
closed_col = find_col("date closed")
dept_col = find_col("department")  # record-level department (first occurrence)
# summary table repeats the "Department" header further right; take the last occurrence.
summary_dept_col = max(c for c, h in headers.items() if isinstance(h, str) and h.strip().lower() == "department")

months = {name.lower(): i for i, name in enumerate(calendar.month_abbr) if name}
months.update({name.lower(): i for i, name in enumerate(calendar.month_name) if name})

# Columns to the right of the summary table whose header names a "<Month> <Year>" target.
month_cols = {}
for c, h in headers.items():
    if not isinstance(h, str) or c <= summary_dept_col:
        continue
    m = re.search(r"([A-Za-z]+)\s+(\d{4})", h)
    if m and m.group(1).lower() in months:
        month_cols[c] = (int(m.group(2)), months[m.group(1).lower()])

records = []
for r in range(2, ws.max_row + 1):
    dept = ws.cell(row=r, column=dept_col).value
    created = ws.cell(row=r, column=created_col).value
    if not dept or not isinstance(created, datetime.datetime):
        continue
    closed = ws.cell(row=r, column=closed_col).value
    records.append((dept, created, closed if isinstance(closed, datetime.datetime) else None))

for r in range(2, ws.max_row + 1):
    dept = ws.cell(row=r, column=summary_dept_col).value
    if not dept:
        continue
    for c, (year, month) in month_cols.items():
        month_end = datetime.datetime(year, month, calendar.monthrange(year, month)[1])
        count = sum(1 for d, created, closed in records
                    if d == dept and created <= month_end and (closed is None or closed > month_end))
        ws.cell(row=r, column=c).value = count

wb.save(OUT)
