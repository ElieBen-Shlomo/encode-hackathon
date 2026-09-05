import os
from datetime import datetime

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Find the Date/Amount header, then collect (month name, amount) for the data rows below it.
header_row = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(row=r, column=1).value == "Date" and ws.cell(row=r, column=2).value == "Amount")
data = []
r = header_row + 1
while r <= ws.max_row and ws.cell(row=r, column=1).value not in (None, ""):
    date_val = ws.cell(row=r, column=1).value
    if not isinstance(date_val, datetime):
        date_val = datetime.strptime(str(date_val).strip(), "%Y-%m-%d")
    amount = ws.cell(row=r, column=2).value
    data.append((date_val.strftime("%B"), amount or 0))
    r += 1

# Find "Result" header, then sum matching amounts for each month listed below it.
# Summing naturally ignores zero-valued rows (adding zero never changes the total).
result_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=2).value == "Result")
r = result_row + 1
while r <= ws.max_row and ws.cell(row=r, column=1).value not in (None, ""):
    month = str(ws.cell(row=r, column=1).value).strip()
    total = sum(amount for name, amount in data if name == month)
    ws.cell(row=r, column=2, value=total)
    r += 1

wb.save(os.environ["OUT_XLSX"])
