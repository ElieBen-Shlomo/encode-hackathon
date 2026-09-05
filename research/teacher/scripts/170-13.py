import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
s1 = wb["Sheet1"]
s2 = wb["Sheet2"]
s3 = wb["Sheet3"]

# Sheet2: each column header (row 1) is a form name; the values below it are the
# associated data items for that header.
headers = {}  # form name -> list of associated data values
for col in range(1, s2.max_column + 1):
    name = s2.cell(row=1, column=col).value
    if name is None:
        continue
    headers[name] = [s2.cell(row=r, column=col).value
                     for r in range(2, s2.max_row + 1)
                     if s2.cell(row=r, column=col).value is not None]

# For each Sheet1 entry (in order), find the Sheet2 header it contains, then emit
# one output row per associated data value (entry string + value). Entries whose
# header has no column in Sheet2 are skipped.
outputs = []
for r in range(2, s1.max_row + 1):
    entry = s1.cell(row=r, column=1).value
    if entry is None:
        continue
    for name, vals in headers.items():
        if name in entry:
            outputs.extend(f"{entry}{v}" for v in vals)
            break

# Rewrite Sheet3 column A: keep the header on row 1, replace the rest.
for r in range(2, s3.max_row + 1):
    s3.cell(row=r, column=1).value = None
for i, val in enumerate(outputs, start=2):
    s3.cell(row=i, column=1, value=val)

wb.save(OUT)
