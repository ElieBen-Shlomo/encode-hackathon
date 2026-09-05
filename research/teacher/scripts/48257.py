import os
import re

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active


def parse_amount(text):
    """Extract a numeric value from a currency string, auto-detecting which of
    '.' or ',' is the decimal sign and treating the other (plus spaces) as
    thousands separators."""
    s = str(text)
    negative = "-" in s
    s = re.sub(r"[A-Za-z]+", "", s)  # drop currency codes
    s = re.sub(r"[\s-]", "", s)  # drop spaces and the sign, kept separately
    m = re.search(r"[.,](\d{1,2})$", s)
    if m:
        decimals = m.group(1)
        integer = re.sub(r"[.,]", "", s[: m.start()])
    else:
        decimals = ""
        integer = re.sub(r"[.,]", "", s)
    integer = integer or "0"
    value = float(f"{integer}.{decimals}" if decimals else integer)
    return round(-value if negative else value, 2)


for row in range(2, ws.max_row + 1):
    raw = ws.cell(row=row, column=2).value
    if raw is None or not str(raw).strip():
        continue
    ws.cell(row=row, column=3).value = parse_amount(raw)

wb.save(os.environ["OUT_XLSX"])
