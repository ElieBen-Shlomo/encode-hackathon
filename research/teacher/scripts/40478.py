import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active


def extract(text):
    """Text after the second space, up to (not including) a dash character."""
    if not isinstance(text, str):
        return None
    parts = text.split(" ", 2)
    if len(parts) < 3:
        return None
    rest = parts[2]
    dash = rest.find("-")
    return (rest[:dash] if dash != -1 else rest).strip()


for row in range(1, ws.max_row + 1):
    value = extract(ws.cell(row, 1).value)
    if value is not None:
        ws.cell(row, 2).value = value

wb.save(OUT)
