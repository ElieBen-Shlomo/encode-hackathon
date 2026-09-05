import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

running = 0        # count of A>0 rows since the last 1 in column B
seen_one = False   # the first 1 has no preceding segment -> stays blank
for r in range(1, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    b = ws.cell(row=r, column=2).value
    if b == 1:
        ws.cell(row=r, column=3, value=running if seen_one else None)
        running = 0
        seen_one = True
    else:
        ws.cell(row=r, column=3, value=None)  # non-1 rows are left blank
        if isinstance(a, (int, float)) and not isinstance(a, bool) and a > 0:
            running += 1

wb.save(OUT)
