import os
import re

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

date_pat = re.compile(r"\d{1,2}/\d{1,2}/\d{4}")

# Data columns are the ones whose header row looks like a date; find "last 4 trend" col.
data_cols = [c for c in range(2, ws.max_column + 1) if isinstance(ws.cell(row=1, column=c).value, str)
             and date_pat.fullmatch(ws.cell(row=1, column=c).value.strip())]
trend_col = next(c for c in range(1, ws.max_column + 1)
                  if isinstance(ws.cell(row=1, column=c).value, str)
                  and "trend" in ws.cell(row=1, column=c).value.lower())

for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=1).value is None:
        continue
    scores = [ws.cell(row=row, column=c).value for c in data_cols if ws.cell(row=row, column=c).value is not None]
    last4 = scores[-4:]
    if last4:
        ws.cell(row=row, column=trend_col).value = sum(last4) / len(last4)

wb.save(OUT)
