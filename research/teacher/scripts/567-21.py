import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

first_row, last_row = 3, ws.max_row

records = []
for row in range(first_row, last_row + 1):
    if ws.cell(row=row, column=1).value is None:
        continue
    vals = [ws.cell(row=row, column=c).value for c in range(1, 8)]
    if vals[4] is None:  # column E: no date -> delete this row entirely
        continue
    records.append(vals)

# Keep only rows at the most recent date E for each (A, B) account/number pair;
# ties at the max date are all kept.
max_date = {}
for vals in records:
    key = (vals[0], vals[1])
    if key not in max_date or vals[4] > max_date[key]:
        max_date[key] = vals[4]

kept = [vals for vals in records if vals[4] == max_date[(vals[0], vals[1])]]

for row in range(first_row, last_row + 1):
    for col in range(1, 8):
        ws.cell(row=row, column=col).value = None

for i, vals in enumerate(kept):
    row = first_row + i
    for col, v in enumerate(vals, start=1):
        ws.cell(row=row, column=col).value = v

wb.save(OUT)
