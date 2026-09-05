import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
data = openpyxl.load_workbook(OUT, data_only=True)  # cached formula results

TOTALS = "Totals"
data_titles = [t for t in wb.sheetnames if t != TOTALS]


def total_column(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Total":
                return cell.column
    return None


# Per data sheet, map each code (column A) to its "Total" value.
sheet_maps = []
for title in data_titles:
    ws = data[title]
    tcol = total_column(ws)
    mapping = {}
    if tcol is not None:
        for r in range(1, ws.max_row + 1):
            code = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=tcol).value
            if code is not None and isinstance(val, (int, float)):
                mapping[code] = val
    sheet_maps.append(mapping)

# Sum each code's total across every data sheet into the Totals sheet (column B).
totals = wb[TOTALS]
for r in range(2, totals.max_row + 1):
    code = totals.cell(row=r, column=1).value
    if code is None:
        continue
    totals.cell(row=r, column=2, value=sum(m.get(code, 0) for m in sheet_maps))

wb.save(OUT)
