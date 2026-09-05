import datetime
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Column J should show only the time-of-day from the datetime in column I,
# formatted as a proper time value (h:mm:ss AM/PM), not a RIGHT() text string.
TIME_FMT = "[$-F400]h:mm:ss\\ AM/PM"

for r in range(2, ws.max_row + 1):
    dt = ws.cell(row=r, column=9).value
    if isinstance(dt, datetime.datetime):
        cell = ws.cell(row=r, column=10)
        cell.value = dt.time()
        cell.number_format = TIME_FMT

wb.save(OUT)
