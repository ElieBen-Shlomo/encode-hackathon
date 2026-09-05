import os

import openpyxl
from openpyxl.styles import PatternFill

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Lookup set: every non-empty value in column F.
f_values = set()
for row in range(1, ws.max_row + 1):
    v = ws.cell(row=row, column=6).value
    if v is not None and v != "":
        f_values.add(v)

fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

# For each data row: "Delete" where A is found in F; highlight C where A and B are numbers.
for row in range(2, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    c = ws.cell(row=row, column=3)
    if a is not None and a in f_values:
        c.value = "Delete"
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        c.fill = fill

wb.save(OUT)
