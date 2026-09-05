import os
from collections import Counter

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

header_row = 4
first, last = header_row + 1, ws.max_row

# The operator's own extension is whichever value dominates column D: each
# incoming call lands on it once, then it forwards the call (possibly to
# several staff in a row) before the next incoming call arrives.
counts = Counter(ws.cell(row=r, column=4).value for r in range(first, last + 1))
operator = counts.most_common(1)[0][0]

run_len = 0
for row in range(first, last + 2):
    is_transfer = row <= last and ws.cell(row=row, column=4).value == operator
    if is_transfer:
        run_len += 1
    elif run_len:
        ws.cell(row=row - 1, column=8).value = run_len
        run_len = 0

wb.save(OUT)
