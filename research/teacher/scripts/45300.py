import os
import re

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Match 'FT' at a word start (so it triggers on "... TAJALI FT2206..." but not mid-word),
# even though 'FT' runs directly into digits with no boundary on its right side.
pattern = re.compile(r"\bFT")

for r in range(1, ws.max_row + 1):
    val = ws.cell(r, 1).value
    if not isinstance(val, str):
        continue
    m = pattern.search(val)
    ws.cell(r, 2).value = val[:m.start()] if m else val

wb.save(OUT)
