import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
src = wb["BZTB"]
dst = wb["Tema UV COS"]

# Rows whose description (column B) starts with COS Pass / COS Comm / COS W/S
# and contains "Used", excluding account 398726.
prefixes = ("COS Pass", "COS Comm", "COS W/S")
matches = []
for row in src.iter_rows(min_row=1, max_col=4):
    desc = row[1].value
    if (isinstance(desc, str) and desc.startswith(prefixes) and "Used" in desc
            and str(row[0].value) != "398726"):
        matches.append([cell.value for cell in row])

# Paste columns A-D into 'Tema UV COS' starting at row 1.
for r, values in enumerate(matches, start=1):
    for c, value in enumerate(values, start=1):
        dst.cell(row=r, column=c, value=value)

wb.save(os.environ["OUT_XLSX"])
