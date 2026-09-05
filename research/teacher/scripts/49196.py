import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

# Column C rebuilds column K's "DD.MM.YY HH:MM" text with slashes instead of dots,
# mirroring the user's formula LEFT/MID pattern -- but the year must come from K
# itself (MID(K,7,2)) rather than TEXT(TODAY(),"yy"), which was stamping today's
# year on every row instead of the year already encoded in the source string.
for row in range(2, ws.max_row + 1):
    k = ws.cell(row=row, column=11).value
    if not isinstance(k, str):
        continue
    day, month, year = k[0:2], k[3:5], k[6:8]
    sp = k.find(" ")
    rest = k[sp:sp + 6] if sp != -1 else ""
    ws.cell(row=row, column=3).value = f"{day}/{month}/{year}{rest}"

wb.save(OUT)
