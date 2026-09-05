import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
s1 = wb["Sheet1"]
s2 = wb["Sheet2"]

# Build a Product -> Sold lookup from Sheet2.
sold = {}
for r in range(2, s2.max_row + 1):
    name = s2.cell(r, 1).value
    if name not in (None, ""):
        sold[name] = s2.cell(r, 2).value

# For each out-of-stock product on Sheet1, place its sold quantity, or "" if not found.
for r in range(2, s1.max_row + 1):
    name = s1.cell(r, 1).value
    if name in (None, ""):
        continue
    s1.cell(r, 2).value = sold.get(name, "")

wb.save(OUT)
