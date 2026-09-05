import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
opps = wb["Opportunities"]
data = wb["Data"]

# Stage -> probability lookup from the Data tab (columns A and B).
stage_pct = {}
for row in range(1, data.max_row + 1):
    stage = data.cell(row=row, column=1).value
    pct = data.cell(row=row, column=2).value
    if isinstance(stage, str) and stage.strip() and pct is not None:
        stage_pct[stage.strip()] = pct

# Weighted Value (G) = Est. Revenue (E) * probability for the Pipeline Stage (D).
for row in range(2, opps.max_row + 1):
    stage = opps.cell(row=row, column=4).value
    revenue = opps.cell(row=row, column=5).value
    if isinstance(stage, str) and stage.strip() in stage_pct and isinstance(revenue, (int, float)):
        opps.cell(row=row, column=7, value=revenue * stage_pct[stage.strip()])

wb.save(OUT)
