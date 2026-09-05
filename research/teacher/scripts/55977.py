import os
import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

for row in range(2, ws.max_row + 1):
    date = ws.cell(row, 1).value
    if date is None:
        continue
    weekday = date.weekday()  # Monday = 0 ... Sunday = 6
    ws.cell(row, 2).value = weekday + 1
    ws.cell(row, 3).value = DAY_NAMES[weekday]

wb.save(OUT)
