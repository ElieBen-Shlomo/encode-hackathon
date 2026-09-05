import datetime
import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

header_row = 5
headers = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}
invoice_date_col = headers["Invoice Date"]
invoice_amount_col = headers["Invoice Amount"]
terms_col = headers["Payment Terms"]
received_col = headers["Amount Received"]
remarks_col = headers["Remarks"]

today = datetime.date.today()

row = header_row + 1
while ws.cell(row=row, column=1).value not in (None, ""):
    invoice_date = ws.cell(row=row, column=invoice_date_col).value
    invoice_amount = ws.cell(row=row, column=invoice_amount_col).value
    terms = ws.cell(row=row, column=terms_col).value
    received = ws.cell(row=row, column=received_col).value

    outstanding = invoice_amount - received
    due_date = invoice_date + datetime.timedelta(days=terms)
    overdue_days = None if outstanding == 0 else (today - due_date.date()).days

    if overdue_days is None:
        remark = ""
    elif outstanding < 0:
        remark = "Prepaid"
    elif overdue_days < 90:
        remark = "Call Customer"
    else:
        remark = "Bad Debts"

    ws.cell(row=row, column=remarks_col).value = remark
    row += 1

wb.save(out_path)
