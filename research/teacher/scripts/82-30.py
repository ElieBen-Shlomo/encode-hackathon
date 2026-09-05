import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
raw = wb["Raw Data"]
numbers = wb["Numbers"]

# Scan column by column (A through F), collecting only whole numbers (no
# decimal point), in the order a VBA loop over columns-then-rows would visit
# them.
whole_numbers = []
for col in range(1, 7):
    for row in range(1, raw.max_row + 1):
        v = raw.cell(row=row, column=col).value
        if isinstance(v, (int, float)) and float(v).is_integer():
            whole_numbers.append(int(v))

# Lay them out 6 per row starting at Numbers!A1.
for i, value in enumerate(whole_numbers):
    r, c = divmod(i, 6)
    numbers.cell(row=r + 1, column=c + 1).value = value

wb.save(OUT)
