import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active


def num(v):
    return v if isinstance(v, (int, float)) else None


# Inventory layers oldest -> newest: (units_col, cost_col).
# Beginning inv (B,C) then PO-1..PO-4 (D,E),(F,G),(H,I),(J,K).
layer_cols = [(2, 3), (4, 5), (6, 7), (8, 9), (10, 11)]

for row in range(2, ws.max_row + 1):
    on_hand = num(ws.cell(row=row, column=12).value)  # L: current inventory
    if not on_hand:
        continue
    layers = []
    for units_col, cost_col in layer_cols:
        units = num(ws.cell(row=row, column=units_col).value)
        cost = num(ws.cell(row=row, column=cost_col).value)
        if units and cost is not None:
            layers.append((units, cost))
    # FIFO: oldest units are consumed first, so the on-hand units are the newest.
    remaining = on_hand
    total_cost = 0.0
    for units, cost in reversed(layers):
        take = min(units, remaining)
        total_cost += take * cost
        remaining -= take
        if remaining <= 0:
            break
    ws.cell(row=row, column=13, value=total_cost / on_hand)

wb.save(OUT)
