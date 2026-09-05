import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

# Row 2 holds short column-name codes (e.g. "ToR"); row 1 has that column's
# level range in row 3+ is column A.
header_row = 2
header_map = {}
for col in range(2, ws.max_column + 1):
    name = ws.cell(row=header_row, column=col).value
    if isinstance(name, str) and name.strip():
        header_map[name] = col

def find_label_value(label):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, str) and v.strip().lower() == label:
                return row, col, ws.cell(row=row, column=col + 1).value
    return None

start_row_label, start_col_label, start_level = find_label_value("starting level")
end_row_label, end_col_label, end_level = find_label_value("ending level")
answer_row, answer_col, _ = find_label_value("answer")

# Find the cell (below the header rows) naming which column to sum.
target_col = None
for row in range(header_row + 1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v in header_map:
            target_col = header_map[v]
            break
    if target_col:
        break

# Column A holds sequential "level" numbers; map level -> sheet row.
level_row = {}
for row in range(header_row + 1, ws.max_row + 1):
    level = ws.cell(row=row, column=1).value
    if level is not None:
        level_row[level] = row

r1, r2 = level_row[start_level], level_row[end_level]
lo, hi = min(r1, r2), max(r1, r2)
total = sum(
    ws.cell(row=r, column=target_col).value or 0 for r in range(lo, hi + 1)
)

ws.cell(row=answer_row, column=answer_col + 1).value = total

wb.save(OUT)
