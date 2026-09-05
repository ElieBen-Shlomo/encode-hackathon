import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

lookup = {}
for row in range(2, ws.max_row + 1):
    model = ws.cell(row=row, column=3).value
    cost = ws.cell(row=row, column=4).value
    if model is not None:
        lookup[model] = cost

for row in range(2, ws.max_row + 1):
    model = ws.cell(row=row, column=1).value
    ws.cell(row=row, column=2).value = lookup.get(model)

wb.save(os.environ["OUT_XLSX"])
