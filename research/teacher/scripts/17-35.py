import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["FILTER 5b"]

start = ws["I2"].value
end = ws["J2"].value

# Data range A6:E315 (headers on row 5); filter rows whose date (col A) is
# between the start (I2) and end (J2) criteria, preserving original order.
rows = []
for r in range(6, ws.max_row + 1):
    date = ws.cell(row=r, column=1).value
    if date is None:
        continue
    if start <= date <= end:
        rows.append([ws.cell(row=r, column=c) for c in range(1, 6)])

# Clear any prior filtered output below the header, then write results from I6.
for r in range(6, ws.max_row + 1):
    for c in range(9, 14):
        ws.cell(row=r, column=c).value = None

for i, src_cells in enumerate(rows):
    out_row = 6 + i
    for j, src in enumerate(src_cells):
        dst = ws.cell(row=out_row, column=9 + j, value=src.value)
        dst.number_format = src.number_format

wb.save(OUT)
