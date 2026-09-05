import os
import openpyxl
from openpyxl.styles import Alignment

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

left = Alignment(horizontal="left")

counts = {}
order = []
for r in range(2, ws.max_row + 1):
    item = ws.cell(row=r, column=2).value
    if item in (None, ""):
        continue
    if item not in counts:
        counts[item] = 0
        order.append(item)
    counts[item] += 1

start_row = 3
ws.cell(row=start_row, column=9, value="ITEM").alignment = left
ws.cell(row=start_row, column=10, value="Qty").alignment = left

for i, item in enumerate(order, start=1):
    r = start_row + i
    ws.cell(row=r, column=9, value=item).alignment = left
    ws.cell(row=r, column=10, value=counts[item]).alignment = left

wb.save(out_path)
