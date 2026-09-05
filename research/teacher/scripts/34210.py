import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Forward-fill: each cell in column C takes the most recent non-empty value from
# column A at or above it (the company heading that block belongs to).
current = None
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a not in (None, ""):
        current = a
    ws.cell(r, 3, value=current)

wb.save(os.environ["OUT_XLSX"])
