import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

# Group columns by their row-1 label (cat/dog), rank each group's row-2 values
# descending, and label the top two columns per group "<label>1"/"<label>2".
groups = {}
for col in range(1, ws.max_column + 1):
    label = ws.cell(row=1, column=col).value
    value = ws.cell(row=2, column=col).value
    if label is not None and isinstance(value, (int, float)):
        groups.setdefault(label, []).append((value, col))

for label, entries in groups.items():
    entries.sort(key=lambda e: e[0], reverse=True)
    for rank, (value, col) in enumerate(entries[:2], start=1):
        ws.cell(row=4, column=col).value = f"{label}{rank}"

wb.save(OUT)
