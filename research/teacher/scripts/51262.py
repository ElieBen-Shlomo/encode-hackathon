import os

import openpyxl
from openpyxl.utils import column_index_from_string

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

DATE_COL = column_index_from_string("B")
CAT_COL = column_index_from_string("C")
VAL_COL = column_index_from_string("D")
CRIT_COL = column_index_from_string("F")

# Data rows: contiguous rows with a date in column B.
data_rows = []
r = 6
while ws.cell(row=r, column=DATE_COL).value is not None:
    data_rows.append(r)
    r += 1

# Criteria list: find the "Criteria List" label, collect values below it.
crit_label_row = None
for row in range(1, ws.max_row + 1):
    if ws.cell(row=row, column=CRIT_COL).value == "Criteria List":
        crit_label_row = row
        break

criteria = set()
r = crit_label_row + 1
while r <= ws.max_row and ws.cell(row=r, column=CRIT_COL).value is not None:
    criteria.add(ws.cell(row=r, column=CRIT_COL).value)
    r += 1

# Header months are in row 5, starting at column F, until a blank cell.
header_row = 5
c = CRIT_COL
while ws.cell(row=header_row, column=c).value is not None:
    header_date = ws.cell(row=header_row, column=c).value
    total = 0
    for row in data_rows:
        d = ws.cell(row=row, column=DATE_COL).value
        cat = ws.cell(row=row, column=CAT_COL).value
        if d is not None and (d.year, d.month) == (header_date.year, header_date.month) and cat in criteria:
            total += ws.cell(row=row, column=VAL_COL).value or 0
    ws.cell(row=header_row + 1, column=c).value = total
    c += 1

wb.save(OUT)
