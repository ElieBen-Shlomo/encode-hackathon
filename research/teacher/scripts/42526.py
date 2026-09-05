import os

import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

values = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
top, bottom = max(values), min(values)  # biggest -> +5, smallest -> -5, proportional in between

fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    score = 5 * v / top if v >= 0 else -5 * v / bottom
    cell = ws.cell(row=r, column=2, value=score)
    cell.fill = fill

wb.save(os.environ["OUT_XLSX"])
