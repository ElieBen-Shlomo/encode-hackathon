import os

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
data = wb["Datensatz"]
analyse = wb["Analyse"]
data_vals = openpyxl.load_workbook(OUT, data_only=True)["Datensatz"]

PROD_START = column_index_from_string("D")
N_PRODUCTS = 8

# Collect customer rows (row 4 downward, column C holds "Customer N")
rows = []
r = 4
while data.cell(row=r, column=3).value is not None:
    rows.append(r)
    r += 1

purchases = []  # purchases[p] = set of customer rows where product p (0-indexed) > 0
for p in range(N_PRODUCTS):
    col = PROD_START + p
    buyers = set()
    for row in rows:
        val = data_vals.cell(row=row, column=col).value
        if isinstance(val, (int, float)) and val > 0:
            buyers.add(row)
    purchases.append(buyers)

n_customers = len(rows)
fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

for i in range(N_PRODUCTS):
    for j in range(N_PRODUCTS):
        cell = analyse.cell(row=5 + i, column=PROD_START + j)
        if i == j:
            cell.value = 1
        else:
            both = len(purchases[i] & purchases[j])
            cell.value = both / n_customers
        cell.fill = fill

wb.save(OUT)
