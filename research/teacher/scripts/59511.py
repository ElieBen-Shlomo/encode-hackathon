import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
wb_vals = openpyxl.load_workbook(OUT, data_only=True)
ws = wb["Sheet1"]
ws_vals = wb_vals["Sheet1"]

# Columns A:C define the range brackets (lower, upper, bracket number); A/B
# are formulas, so read their computed values.
brackets = []
for row in range(2, ws.max_row + 1):
    lo = ws_vals.cell(row=row, column=1).value
    hi = ws_vals.cell(row=row, column=2).value
    num = ws.cell(row=row, column=3).value
    if lo is not None and hi is not None:
        brackets.append((lo, hi, num))

for row in range(2, ws.max_row + 1):
    value = ws.cell(row=row, column=5).value
    if not isinstance(value, (int, float)):
        continue
    for lo, hi, num in brackets:
        if lo <= value <= hi:
            ws.cell(row=row, column=6).value = num
            break

wb.save(OUT)
