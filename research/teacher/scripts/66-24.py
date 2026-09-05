import datetime
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
src = wb["Imported Data"]
dst = wb["Items Older than 30 days"]

# Header width: contiguous non-empty cells starting at column A.
n_cols = 0
for col in range(1, src.max_column + 1):
    if src.cell(1, col).value is None:
        break
    n_cols = col

dates = [src.cell(r, 5).value for r in range(2, src.max_row + 1) if src.cell(r, 5).value is not None]
cutoff = max(dates) - datetime.timedelta(days=30)

for col in range(1, n_cols + 1):
    dst.cell(1, col).value = src.cell(1, col).value

out_row = 2
for row in range(2, src.max_row + 1):
    date = src.cell(row, 5).value
    if date is None or not (date < cutoff):
        continue
    for col in range(1, n_cols + 1):
        dst.cell(out_row, col).value = src.cell(row, col).value
        if col == 5:
            dst.cell(out_row, col).number_format = src.cell(row, col).number_format
    out_row += 1

dst.delete_cols(26, 1)

wb.save(OUT)
