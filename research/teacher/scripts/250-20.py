import os
from collections import OrderedDict

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["RNM"]
KEY1, KEY2, QTY = 2, 3, 10  # match on columns B & C; sum column J

# Find the last data row so we ignore the large empty tail.
last = 1
for r in range(2, ws.max_row + 1):
    if any(ws.cell(r, c).value not in (None, "") for c in range(1, 12)):
        last = r

# Group rows by the matching pair in columns B & C, keeping first appearance and summing J.
groups = OrderedDict()
for r in range(2, last + 1):
    b, c = ws.cell(r, KEY1).value, ws.cell(r, KEY2).value
    if b in (None, "") and c in (None, ""):
        continue
    key = (str(b).strip(), str(c).strip())
    j = ws.cell(r, QTY).value or 0
    if key in groups:
        groups[key][0] += j
        groups[key][1].append(r)
    else:
        groups[key] = [j, [r]]

# For each duplicated pair, put the summed value on the first row and mark the
# rest for deletion; delete bottom-up so earlier row indices stay valid.
to_delete = []
for total, occ in groups.values():
    if len(occ) > 1:
        ws.cell(occ[0], QTY).value = total
        to_delete.extend(occ[1:])

for r in sorted(to_delete, reverse=True):
    ws.delete_rows(r, 1)

wb.save(os.environ["OUT_XLSX"])
