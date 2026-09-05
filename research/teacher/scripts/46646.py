import calendar
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

year = ws["B3"].value

for col in range(2, 14):  # B..M -> Jan..Dec
    month = col - 1
    shifts = ws.cell(row=9, column=col).value or 0
    days = calendar.monthrange(int(year), month)[1]
    ws.cell(row=11, column=col).value = shifts / days

wb.save(OUT)
