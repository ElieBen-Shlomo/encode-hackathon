"""Make J23 mirror I12 exactly: blank when I12 is blank, otherwise I12's own value (fixing
the old IF(I12="","") formula which showed FALSE/a 1900 date instead of a clean blank)."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Deal 8"]

i12 = ws["I12"].value
ws["J23"] = None if i12 in (None, "") else i12

wb.save(os.environ["OUT_XLSX"])
