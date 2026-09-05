import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

headers = [ws.cell(row=1, column=col).value for col in range(1, 7)]

for row in range(2, ws.max_row + 1):
    matched = [
        headers[col - 1]
        for col in range(1, 7)
        if str(ws.cell(row=row, column=col).value).strip().upper() == "Y"
    ]
    ws.cell(row=row, column=7, value=", ".join(matched))

wb.save(os.environ["OUT_XLSX"])
