import calendar
import datetime
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

wb_vals = openpyxl.load_workbook(OUT, data_only=True)
ws_vals = wb_vals.active

month_num = {name: i for i, name in enumerate(calendar.month_name) if name}

# Gather (month, qty) for every data row that has a real date in column C.
qty_by_month = {}
for row in range(2, ws_vals.max_row + 1):
    date = ws_vals.cell(row=row, column=3).value
    qty = ws_vals.cell(row=row, column=9).value
    if isinstance(date, datetime.datetime) and isinstance(qty, (int, float)):
        qty_by_month[date.month] = qty_by_month.get(date.month, 0) + qty

for row in range(26, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    m = month_num.get(str(name).strip()) if name else None
    if m is not None:
        ws.cell(row=row, column=2).value = qty_by_month.get(m, 0)

wb.save(OUT)
