import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Exp-DB"]

# Vendor -> category lookup table lives in columns A:B (short "bank speak" prefixes).
vendors = []
r = 1
while ws.cell(row=r, column=1).value is not None:
    vendor = ws.cell(row=r, column=1).value
    category = ws.cell(row=r, column=2).value
    vendors.append((vendor, category))
    r += 1
# Match the longest (most specific) vendor prefix first.
vendors.sort(key=lambda vc: len(vc[0]), reverse=True)

r = 1
while ws.cell(row=r, column=4).value is not None:
    desc = ws.cell(row=r, column=4).value
    category = None
    for vendor, cat in vendors:
        if vendor in desc:
            category = cat
            break
    ws.cell(row=r, column=5).value = category
    r += 1

wb.save(OUT)
