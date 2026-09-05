import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
emp = wb["Emp"]
lookup = wb["Lookup"]
expected = wb["Expected"]

# Lookup table: Variable -> (Category, Subcategory)
lut = {}
r = 2
while lookup.cell(row=r, column=1).value is not None:
    lut[lookup.cell(row=r, column=1).value] = (
        lookup.cell(row=r, column=2).value,
        lookup.cell(row=r, column=3).value,
    )
    r += 1

# Emp table: first 4 columns are the constant keys, remaining columns are variables to unpivot.
KEY_COLS = 4
var_names = []
c = KEY_COLS + 1
while emp.cell(row=1, column=c).value is not None:
    var_names.append(emp.cell(row=1, column=c).value)
    c += 1

out_row = 2
r = 2
while emp.cell(row=r, column=1).value is not None:
    keys = [emp.cell(row=r, column=k).value for k in range(1, KEY_COLS + 1)]
    for i, var in enumerate(var_names):
        value = emp.cell(row=r, column=KEY_COLS + 1 + i).value
        category, subcategory = lut.get(var, (None, None))
        row_vals = keys + [var, category, subcategory, value]
        for col, val in enumerate(row_vals, start=1):
            expected.cell(row=out_row, column=col).value = val
        out_row += 1
    r += 1

wb.save(OUT)
