import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws1 = wb["Sheet1"]
ws2 = wb["Sheet2"]

# Build lookup: Sheet1 column D (Existing Data Fields) -> Sheet1 column B (Deal Name).
lookup = {}
for row in range(2, ws1.max_row + 1):
    key = ws1.cell(row=row, column=4).value
    if isinstance(key, str):
        key = key.strip()
    if key is not None and key not in lookup:
        lookup[key] = ws1.cell(row=row, column=2).value

# For each field in Sheet2 column A, write the matched deal name into column B.
for row in range(2, ws2.max_row + 1):
    key = ws2.cell(row=row, column=1).value
    if key is None:
        continue
    if isinstance(key, str):
        key = key.strip()
    if key in lookup:
        ws2.cell(row=row, column=2, value=lookup[key])

wb.save(OUT)
