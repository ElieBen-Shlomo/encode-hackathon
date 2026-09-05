import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

wb_vals = openpyxl.load_workbook(OUT, data_only=True)
ws_vals = wb_vals.active

msrp = ws_vals["C8"].value
discount = ws_vals["E8"].value

ws["F8"].value = msrp * (1 - discount)

wb.save(OUT)
