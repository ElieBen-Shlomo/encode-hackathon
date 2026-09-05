import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Collect column B values that begin with "PK", in row order.
matches = []
for row in range(1, ws.max_row + 1):
    value = ws.cell(row=row, column=2).value
    if isinstance(value, str) and value.startswith("PK"):
        matches.append(value)

# Write them consecutively into the result column E starting at E4, clearing what follows.
start = 4
for offset, value in enumerate(matches):
    ws.cell(row=start + offset, column=5, value=value)
for row in range(start + len(matches), ws.max_row + 1):
    ws.cell(row=row, column=5, value=None)

wb.save(OUT)
