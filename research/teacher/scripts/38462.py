import os
from datetime import timedelta

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Column B holds elapsed time ([h]:mm). Round each up to the next whole hour in column C.
for row in range(1, ws.max_row + 1):
    v = ws.cell(row=row, column=2).value
    if isinstance(v, timedelta):
        seconds = v.days * 86400 + v.seconds
    elif isinstance(v, (int, float)):
        seconds = round(v * 86400)  # a bare [h]:mm value is a fraction of a day
    else:
        continue
    hours_up = -(-seconds // 3600)  # ceiling division to the whole hour
    cell = ws.cell(row=row, column=3, value=hours_up)
    cell.number_format = "General"

wb.save(OUT)
