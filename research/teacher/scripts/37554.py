import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Build TYPE -> Accounts from the source table (col A = TYPE, col B = Accounts).
counts = {}
for r in range(2, ws.max_row + 1):
    t = ws.cell(row=r, column=1).value
    if t is None or str(t).strip() == "":
        continue
    n = ws.cell(row=r, column=2).value
    key = str(t).strip()
    counts[key] = counts.get(key, 0) + (n if isinstance(n, (int, float)) else 0)

# Column F: exact-match count for each TYPE listed in column D (stop at TOTAL/blank).
for r in range(2, ws.max_row + 1):
    d = ws.cell(row=r, column=4).value
    if d is None or str(d).strip() in ("", "TOTAL"):
        break
    ws.cell(row=r, column=6, value=counts.get(str(d).strip(), 0))

wb.save(OUT)
