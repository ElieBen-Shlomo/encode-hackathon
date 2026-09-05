import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)


def find_sheet(wb, needle):
    needle = needle.lower()
    for name in wb.sheetnames:
        if name.lower() == needle:
            return wb[name]
    raise KeyError(needle)


src = find_sheet(wb, "imported data")
dst = find_sheet(wb, "summary limits")

max_col = src.max_column
header = [src.cell(row=1, column=c).value for c in range(1, max_col + 1)]

out_row = 1
for c, val in enumerate(header, start=1):
    dst.cell(row=out_row, column=c, value=val)
out_row += 1

for r in range(2, src.max_row + 1):
    a_val = src.cell(row=r, column=1).value
    if isinstance(a_val, str) and a_val.startswith("**"):
        for c in range(1, max_col + 1):
            dst.cell(row=out_row, column=c, value=src.cell(row=r, column=c).value)
        out_row += 1

wb.save(out_path)
