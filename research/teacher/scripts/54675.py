import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
NAME = "Gary Stoddard"

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

# Count every occurrence of the exact name within column B (a cell may list
# several names, e.g. "Gary Stoddard, Gary Stoddart"), restricted to rows
# where column A equals 4 — a substring-occurrence count, not just COUNTIFS
# exact-match, since the name can appear alongside other names in one cell.
total = 0
for row in range(2, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    if a is None or b is None:
        continue
    if str(a).strip() != "4":
        continue
    total += str(b).count(NAME)

ws["F3"].value = total
wb.save(OUT)
