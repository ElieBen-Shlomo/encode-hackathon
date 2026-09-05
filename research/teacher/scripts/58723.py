import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

NAME_COL, TIME_COL, RESULT_COL = 1, 9, 13

max_time = {}
for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=NAME_COL).value
    t = ws.cell(row=row, column=TIME_COL).value
    if name is None or t is None:
        continue
    if name not in max_time or t > max_time[name]:
        max_time[name] = t

for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=NAME_COL).value
    t = ws.cell(row=row, column=TIME_COL).value
    if name is None or t is None:
        continue
    ws.cell(row=row, column=RESULT_COL).value = "Latest" if t == max_time[name] else "Not Latest"

wb.save(OUT)
