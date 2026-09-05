import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Column G lists the years to summarize, starting at row 2 and running until the
# data block header (row 13) is reached. For each year, count rows in G14:G185
# whose year matches and whose status in J14:J185 is "open" (case-insensitive,
# since the source data mixes "Open"/"open").
years = []
r = 2
while isinstance(ws.cell(r, 7).value, (int, float)):
    years.append((r, ws.cell(r, 7).value))
    r += 1

for row, year in years:
    count = sum(
        1 for rr in range(14, 186)
        if ws.cell(rr, 7).value == year and str(ws.cell(rr, 10).value or "").strip().lower() == "open"
    )
    ws.cell(row, 9, count)

wb.save(os.environ["OUT_XLSX"])
