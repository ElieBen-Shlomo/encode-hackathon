import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# Split column A into groups of consecutive non-blank cells (blank rows are separators).
groups = []
current = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v is None or (isinstance(v, str) and not v.strip()):
        if current:
            groups.append(current)
            current = []
    else:
        current.append(v)
if current:
    groups.append(current)

# Transpose each group onto its own row, laid out horizontally from column C.
for i, group in enumerate(groups, start=1):
    for j, val in enumerate(group):
        ws.cell(row=i, column=3 + j).value = val

wb.save(OUT)
