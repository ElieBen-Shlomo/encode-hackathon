import os
from datetime import datetime

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

ref_format = ws["A2"].number_format  # existing display format to keep consistent

# Cells with an embedded time value look like '2021-01-18T20:57:45.962Z': a date,
# a literal 'T', a time, and a trailing 'Z'. Strip everything from 'T' onward and
# keep just the date, written as a real datetime so the format still applies.
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=1)
    v = cell.value
    if isinstance(v, str) and "T" in v and v.endswith("Z"):
        date_part = v.split("T", 1)[0]
        cell.value = datetime.strptime(date_part, "%Y-%m-%d")
        cell.number_format = ref_format

wb.save(OUT)
