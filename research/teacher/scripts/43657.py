import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

# Locate each weekly block: a header row containing "Mon" ... "Sun"; the
# actual per-day name entries live two rows below (header, then dates, then names).
header_rows = []
col_start = col_end = None
for row in ws.iter_rows():
    for cell in row:
        if cell.value == "Mon":
            header_rows.append(cell.row)
            if col_start is None:
                col_start = cell.column
                c = col_start
                while ws.cell(row=cell.row, column=c).value not in (None, ""):
                    col_end = c
                    c += 1
            break

data_rows = [r + 2 for r in header_rows]

# Name list is in column L (results written to column K, immediately to its left).
name_col = 12
result_col = 11
row = 2
while ws.cell(row=row, column=name_col).value not in (None, ""):
    name = str(ws.cell(row=row, column=name_col).value).strip().lower()
    count = 0
    for dr in data_rows:
        for c in range(col_start, col_end + 1):
            val = ws.cell(row=dr, column=c).value
            if val and name in str(val).lower():
                count += 1
    ws.cell(row=row, column=result_col, value=count)
    row += 1

wb.save(out_path)
