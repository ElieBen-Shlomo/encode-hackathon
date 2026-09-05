import os

import openpyxl
from openpyxl.styles import PatternFill

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

OCOGS = "OCOGS - Spares - Transfer Price Overhead"
OFSS = "OFSS - ORCL Consulting Prod Cost I/C"
AY, BG, BL, BN = 51, 59, 64, 66
blue = PatternFill(fill_type="solid", start_color="FF0070C0", end_color="FF0070C0")

for row in range(2, ws.max_row + 1):
    bl = ws.cell(row=row, column=BL).value
    bg = ws.cell(row=row, column=BG).value
    ay = ws.cell(row=row, column=AY).value
    if bl == "LAG" and bg in (OCOGS, OFSS):
        value = OCOGS if ay in ("BOA_033E", "BOA_011G") else OFSS
        ws.cell(row=row, column=BN).value = value
        ws.cell(row=row, column=BL).fill = blue

wb.save(OUT)
