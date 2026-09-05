import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])

# On each of the three named sheets, delete rows from row 6 down whose column F
# contains "to be processed". Iterate bottom-up so row shifts are safe.
for name in ("99250", "99251", "99252"):
    ws = wb[name]
    for r in range(ws.max_row, 5, -1):
        v = ws.cell(row=r, column=6).value
        if isinstance(v, str) and "to be processed" in v.lower():
            ws.delete_rows(r)

wb.save(os.environ["OUT_XLSX"])
