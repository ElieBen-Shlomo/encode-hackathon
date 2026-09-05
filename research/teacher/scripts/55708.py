import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

STATUSES = {"in progress", "in review"}

records = []
r = 2
while ws.cell(row=r, column=1).value is not None:
    dept = ws.cell(row=r, column=1).value
    status = ws.cell(row=r, column=2).value
    time = ws.cell(row=r, column=3).value
    records.append((dept, status, time))
    r += 1

calc_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Department" and r > 1)
r = calc_row + 1
while ws.cell(row=r, column=1).value is not None:
    dept = ws.cell(row=r, column=1).value
    values = [t for d, s, t in records
              if d == dept and isinstance(s, str) and s.strip().lower() in STATUSES and t is not None and t >= 6]
    ws.cell(row=r, column=2).value = sum(values) / len(values) if values else None
    r += 1

wb.save(OUT)
