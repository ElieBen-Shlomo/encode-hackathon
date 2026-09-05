import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Track No (column F) restarts at 1 for each new album (column D) and increments
# for each subsequent song within the same album.
prev = None
n = 0
for r in range(2, ws.max_row + 1):
    album = ws.cell(row=r, column=4).value
    if album is None:
        continue
    n = n + 1 if album == prev else 1
    prev = album
    ws.cell(row=r, column=6, value=n)

wb.save(os.environ["OUT_XLSX"])
