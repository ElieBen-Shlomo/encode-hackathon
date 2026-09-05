import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
tracker = wb["Consolidated Tracker"]
existing = wb["Existing Task"]
additions = wb["Additions"]
retired = wb["Retired"]

COLS = 5  # Category, Activity Name, AU, Frequency, Team Member


def read_rows(ws):
    rows = []
    for r in range(3, ws.max_row + 1):
        row = tuple(ws.cell(row=r, column=c).value for c in range(1, COLS + 1))
        if any(v is not None for v in row):
            rows.append(row)
    return rows


combined = read_rows(existing) + read_rows(additions)
to_remove = read_rows(retired)

for row in to_remove:
    if row in combined:
        combined.remove(row)

for i, row in enumerate(combined):
    out_row = 3 + i
    for c, val in enumerate(row, start=1):
        tracker.cell(row=out_row, column=c, value=val)

wb.save(os.environ["OUT_XLSX"])
