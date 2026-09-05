import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

HEADER_ROW = 4
CLIENT_COL = 3  # C: Casino Client
TABLE_START, TABLE_END = HEADER_ROW + 1, HEADER_ROW + 8  # data rows 5-12

# Locate each "Rep N" group: (name_col, pct_col), by scanning the header row.
groups = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=HEADER_ROW, column=col).value
    if isinstance(header, str) and header.strip().startswith("Rep ") and header.strip() != "Rep(s)":
        groups.append((col, col + 2))  # name column, % column two over

# Locate the lookup block: header row with "Rep", "Client", "%".
lookup_header_row = None
for row in range(1, ws.max_row + 1):
    if ws.cell(row=row, column=2).value == "Rep" and ws.cell(row=row, column=3).value == "Client":
        lookup_header_row = row
        break

for row in range(lookup_header_row + 1, ws.max_row + 1):
    rep_name = ws.cell(row=row, column=2).value
    client = ws.cell(row=row, column=3).value
    if not rep_name or not client:
        continue

    pct = None
    for r in range(TABLE_START, TABLE_END + 1):
        if ws.cell(row=r, column=CLIENT_COL).value == client:
            for name_col, pct_col in groups:
                if ws.cell(row=r, column=name_col).value == rep_name:
                    pct = ws.cell(row=r, column=pct_col).value
                    break
            break

    if pct is not None:
        ws.cell(row=row, column=4, value=pct)

wb.save(os.environ["OUT_XLSX"])
