import os
from datetime import datetime, timedelta

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

# Group A: "Year-Month" rows starting at row 3, columns B..AF = day 1..31.
month_rows = {}
r = 3
while True:
    v = ws.cell(row=r, column=1).value
    if v is None or v == "Group B":
        break
    if isinstance(v, str) and "-" in v:
        month_rows[v] = r
    r += 1

# Insert a blank row for any missing month in the Jan..(last present) sequence.
year = min(month_rows)[:4]
last_num = int(max(month_rows)[5:7])
for m in range(1, last_num + 1):
    ym = f"{year}-{m:02d}"
    if ym in month_rows:
        continue
    insert_at = next(rr for mm, rr in sorted(month_rows.items(), key=lambda kv: kv[1]) if mm > ym)
    ws.insert_rows(insert_at)
    ws.cell(row=insert_at, column=1).value = ym
    month_rows = {mm: (rr + 1 if rr >= insert_at else rr) for mm, rr in month_rows.items()}
    month_rows[ym] = insert_at

# Group B: "Group B" label, then a header row, then Date/Data rows.
gb_row = next(rr for rr in range(1, ws.max_row + 1) if ws.cell(row=rr, column=1).value == "Group B")
out_row = gb_row + 2

start_date = datetime(int(year), 1, 1)
end_date = datetime(int(year), 11, 16)  # per instruction: "till 11/16/1951"

d = start_date
while d <= end_date:
    ym = f"{d.year}-{d.month:02d}"
    src_row = month_rows.get(ym)
    val = ws.cell(row=src_row, column=1 + d.day).value if src_row else None
    if val == "M" or (isinstance(val, (int, float)) and val <= -9999):
        val = None  # irrelevant-data sentinels dropped
    elif val == "NA":
        pass  # NA left as-is
    elif val is None:
        val = 0 if d.month == 4 else None  # April: 0s not blanks
    ws.cell(row=out_row, column=1).value = d
    ws.cell(row=out_row, column=2).value = val
    out_row += 1
    d += timedelta(days=1)

wb.save(OUT)
