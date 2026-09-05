import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# Column R remarks: if S is an error -> "Upload"; otherwise if the Difference in U is
# within [-1, 1] -> "Do not Upload"; otherwise (Difference > 1) -> "to Check".
for r in range(2, ws.max_row + 1):
    s = ws.cell(row=r, column=19).value  # S
    u = ws.cell(row=r, column=21).value  # U
    if s is None and u is None:
        continue
    if isinstance(s, str) and s.startswith("#"):
        val = "Upload"
    elif u is not None and -1 <= u <= 1:
        val = "Do not Upload"
    else:
        val = "to Check"
    ws.cell(row=r, column=18, value=val)  # R

wb.save(OUT)
