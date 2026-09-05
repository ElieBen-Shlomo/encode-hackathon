import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Table-1: country names listed under the "Country" header in column A.
countries = []
in_table = False
for row in range(1, ws.max_row + 1):
    v = ws.cell(row=row, column=1).value
    text = str(v).strip() if v is not None else ""
    if text == "Country":
        in_table = True
        continue
    if in_table:
        if not text or text == "Company":
            break
        countries.append(text)

# Table-2: header row starts with "Company"; country columns follow, result column is "No of countries".
comp_header = next(r for r in range(1, ws.max_row + 1)
                   if str(ws.cell(row=r, column=1).value).strip() == "Company")
col_by_country = {}
result_col = None
for c in range(2, ws.max_column + 1):
    h = ws.cell(row=comp_header, column=c).value
    if h is None:
        continue
    h = str(h).strip()
    if h.lower().startswith("no of"):
        result_col = c
    else:
        col_by_country[h] = c

# Count, per company, the Table-1 countries with more than 0 employees.
for r in range(comp_header + 1, ws.max_row + 1):
    company = ws.cell(row=r, column=1).value
    if company is None or not str(company).strip():
        continue
    count = 0
    for country in countries:
        c = col_by_country.get(country)
        if c is not None:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and v > 0:
                count += 1
    ws.cell(row=r, column=result_col, value=count)

wb.save(OUT)
