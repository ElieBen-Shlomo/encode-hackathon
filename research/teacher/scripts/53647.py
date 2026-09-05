import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

# Locate the "Parts" totals row (West total in C, East total in D) above the data block.
totals_row = None
for r in range(1, ws.max_row + 1):
    b = ws.cell(row=r, column=2).value
    if isinstance(b, str) and "parts" in b.lower():
        totals_row = r
        break

c_ref = f"$C${totals_row}"
d_ref = f"$D${totals_row}"

for row in range(7, 17):
    formula = (
        f'=IF(AND(C{row}<>"",D{row}=""),{c_ref}/C{row},'
        f'IF(AND(D{row}<>"",C{row}=""),{d_ref}/D{row},""))'
    )
    ws.cell(row=row, column=5, value=formula)

wb.save(out_path)
