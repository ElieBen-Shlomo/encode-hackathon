"""For each row on Pricing, find the row in 'Package & Weight Data' whose Package Type
matches column C and whose Weight From/To range contains the weight in column D, and copy
its Price into Pricing column E (Post Cost)."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws_price = wb["Pricing"]
ws_data = wb["Package & Weight Data"]

bands = []
for r in range(2, ws_data.max_row + 1):
    pkg, wfrom, wto, price = (ws_data.cell(row=r, column=c).value for c in (1, 2, 3, 4))
    if pkg is not None and price is not None:
        bands.append((pkg, wfrom, wto, price))

for r in range(2, ws_price.max_row + 1):
    pkg = ws_price.cell(row=r, column=3).value
    weight = ws_price.cell(row=r, column=4).value
    if pkg is None or weight is None:
        continue
    for b_pkg, wfrom, wto, price in bands:
        if b_pkg == pkg and wfrom <= weight <= wto:
            ws_price.cell(row=r, column=5).value = price
            break

wb.save(os.environ["OUT_XLSX"])
