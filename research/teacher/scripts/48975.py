"""Copy Input!B (Text) values into Output!B, starting at row 11, for each Input data row
(from row 9 down) whose To Do column (E) is 'yes', in original row order. Output column A
already holds its own sequential numbering and is left untouched."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws_in = wb["Input"]
ws_out = wb["Output"]

matches = []
for row in ws_in.iter_rows(min_row=9, max_col=5):
    text, todo = row[1].value, row[4].value
    if text is not None and isinstance(todo, str) and todo.strip().lower() == "yes":
        matches.append(text)

for i, text in enumerate(matches):
    ws_out.cell(row=11 + i, column=2).value = text

wb.save(os.environ["OUT_XLSX"])
