import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

# Row 3 holds the month headers; data columns run from column B onward.
header_row = 3
first_col = 2
last_col = ws.max_column
out_col = last_col + 1  # first free column right after the data (N when data is B:M)

for row in range(header_row + 1, ws.max_row + 1):
    label = ws.cell(row=row, column=1).value
    if label in (None, ""):
        continue

    numeric_cols = [
        c for c in range(first_col, last_col + 1)
        if isinstance(ws.cell(row=row, column=c).value, (int, float))
    ]
    if len(numeric_cols) < 2:
        continue

    first_col_idx, second_col_idx = numeric_cols[0], numeric_cols[1]
    first_value = ws.cell(row=row, column=first_col_idx).value

    if first_value > 1:
        result = 1
    else:
        result = second_col_idx - first_col_idx

    ws.cell(row=row, column=out_col).value = result

wb.save(out_path)
