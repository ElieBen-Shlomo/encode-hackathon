import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Header row holds both tables' column titles; data starts the row after it.
header_row = next(r for r in range(1, ws.max_row + 1)
                   if any(str(c.value).strip().upper() == "PRICEDATE" for c in ws[r]))
col_pair = next(c.column for c in ws[header_row] if str(c.value).strip().upper() == "CROSS-PARITY IN GBP")
col_rate = next(c.column for c in ws[header_row] if str(c.value).strip().upper() == "MIDRATE")
col_date = next(c.column for c in ws[header_row] if str(c.value).strip().upper() == "PRICEDATE")
col_out = next(c.column for c in ws[header_row] if "CROSS-RATE" in str(c.value).strip().upper())

data_start = header_row + 1
data_end = data_start
while ws.cell(row=data_end + 1, column=col_pair).value not in (None, ""):
    data_end += 1

# Table 2's target currency comes from its own header, e.g. "EUR-XXX CROSS-RATE" -> EUR.
target_ccy = str(ws.cell(row=header_row, column=col_out).value).split("-")[0].strip()

# Rows sharing one PRICEDATE form a block; each block's divisor is its own "<src>-<target>" rate,
# so dragging the ratio down always references the right row instead of a fixed absolute cell.
groups = {}
for r in range(data_start, data_end + 1):
    groups.setdefault(ws.cell(row=r, column=col_date).value, []).append(r)

for rows in groups.values():
    divisor_row = next(r for r in rows if str(ws.cell(row=r, column=col_pair).value).split("-")[-1] == target_ccy)
    divisor = ws.cell(row=divisor_row, column=col_rate).value
    for r in rows:
        ws.cell(row=r, column=col_out, value=ws.cell(row=r, column=col_rate).value / divisor)

wb.save(os.environ["OUT_XLSX"])
