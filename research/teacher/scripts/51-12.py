import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

# Data range A1:M1. Reference starts at A1; scan right until a cell is >= 10%
# above the current reference, count it, then make the *next* cell (right of
# the match) the new reference and resume comparing from the cell after that.
vals = [ws.cell(row=1, column=c).value for c in range(1, 14)]

count = 0
ref = vals[0]
i = 1
while i < len(vals):
    v = vals[i]
    if v is None:
        break
    if ref is not None and v >= ref * 1.10 - 1e-9:
        count += 1
        if i + 1 < len(vals):
            ref = vals[i + 1]
            i += 2
        else:
            break
    else:
        i += 1

ws.cell(row=6, column=2).value = count
wb.save(OUT)
