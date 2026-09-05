import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Running counter in column A: increment for each non-empty column B cell (from B2),
# and leave column A empty on rows where B is blank.
count = 0
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 2).value not in (None, ""):
        count += 1
        ws.cell(r, 1).value = count
    else:
        ws.cell(r, 1).value = None

wb.save(os.environ["OUT_XLSX"])
