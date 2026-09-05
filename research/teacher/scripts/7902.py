import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
grouping = wb["Grouping"]
req = wb["Formula Required"]

N_BUCKETS = 7  # columns C:I in Grouping / D:J in Formula Required

specific = {}   # (group, material) -> [values]
generic = {}    # group -> [values] for rows with no material
first_by_group = {}  # group -> [values], first row seen for that group

r = 2
while grouping.cell(row=r, column=1).value is not None:
    group = grouping.cell(row=r, column=1).value
    material = grouping.cell(row=r, column=2).value or None
    values = [grouping.cell(row=r, column=3 + i).value for i in range(N_BUCKETS)]
    if material is not None:
        specific[(group, material)] = values
    else:
        generic[group] = values
    first_by_group.setdefault(group, values)
    r += 1

for row in range(3, req.max_row + 1):
    group = req.cell(row=row, column=3).value
    material = req.cell(row=row, column=2).value or None
    if group is None:
        continue
    if material is not None and (group, material) in specific:
        values = specific[(group, material)]
    elif group in generic:
        values = generic[group]
    else:
        values = first_by_group.get(group)
    if values is None:
        continue
    for i, val in enumerate(values):
        req.cell(row=row, column=4 + i).value = val

wb.save(OUT)
