import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

employees = []
for r in range(2, ws.max_row + 1):
    hire = ws.cell(row=r, column=1).value
    term = ws.cell(row=r, column=2).value
    if hire is None:
        continue
    employees.append((hire, term))

# Current headcount for a month = hired on/before that month and not yet
# terminated (blank term date, or terminated after that month).
for col in range(5, 17):  # E..P
    month = ws.cell(row=1, column=col).value
    count = sum(1 for hire, term in employees if hire <= month and (term is None or term > month))
    ws.cell(row=2, column=col).value = count

wb.save(OUT)
