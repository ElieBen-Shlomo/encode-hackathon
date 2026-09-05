"""Sum QTY (col B) for rows where PRODUCT ID (col A) equals the example value given in the
instruction (111111), writing the total into D2."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

TARGET_ID = 111111

total = 0
for row in ws.iter_rows(min_row=2, max_col=2):
    pid, qty = row[0].value, row[1].value
    if pid == TARGET_ID and qty is not None:
        total += qty

ws["D2"] = total

wb.save(os.environ["OUT_XLSX"])
