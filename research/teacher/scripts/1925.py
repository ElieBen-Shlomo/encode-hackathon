import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws1 = wb["Sheet1"]
ws2 = wb["Sheet2"]

# Read cached values (C5 date and Sheet2 contents) in case any are formulas.
data = openpyxl.load_workbook(os.environ["IN_XLSX"], data_only=True)
d1, d2 = data["Sheet1"], data["Sheet2"]

# Find the Sheet2 column whose date in row 5 matches Sheet1!C5.
target = d1["C5"].value
match_col = next(c for c in range(1, d2.max_column + 1)
                 if d2.cell(row=5, column=c).value == target)

# Items available in that column (below the date row).
available = {d2.cell(row=r, column=match_col).value
             for r in range(6, d2.max_row + 1)
             if d2.cell(row=r, column=match_col).value is not None}

# For each item in Sheet1 column B, keep it if available, otherwise "NEW".
for r in range(6, ws1.max_row + 1):
    item = ws1.cell(row=r, column=2).value
    if item is None or not str(item).strip():
        continue
    ws1.cell(row=r, column=3, value=item if item in available else "NEW")

wb.save(OUT)
