import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws1 = wb["Sheet1"]
ws2 = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.create_sheet("Sheet2")

# Headers in A2:C2 from Sheet1 row 1.
for col in range(1, 4):
    ws2.cell(row=2, column=col, value=ws1.cell(row=1, column=col).value)

# Extract matching rows: B == 'TELIVISION' or C in ('CLASS III', 'CLASS IV').
out_row = 3
for row in range(2, ws1.max_row + 1):
    item = ws1.cell(row=row, column=2).value
    cls = ws1.cell(row=row, column=3).value
    if item == "TELIVISION" or cls in ("CLASS III", "CLASS IV"):
        for col in range(1, 4):
            ws2.cell(row=out_row, column=col, value=ws1.cell(row=row, column=col).value)
        out_row += 1

wb.save(OUT)
