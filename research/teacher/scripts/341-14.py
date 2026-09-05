import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
src = wb["problem"]
dst = wb["result"]


def to_num(tok):
    try:
        f = float(tok)
        return int(f) if f == int(f) else f
    except ValueError:
        return tok


# Explode each problem row: label in col A, newline-separated numbers in col B become one
# output row each. Labels with no numbers still get one row (B left unfilled).
out = []
for r in range(1, src.max_row + 1):
    label = src.cell(r, 1).value
    if label in (None, ""):
        continue
    raw = src.cell(r, 2).value
    if raw in (None, ""):
        nums = []
    elif isinstance(raw, str):
        nums = [to_num(t) for t in raw.split("\n") if t.strip() != ""]
    else:
        nums = [raw]
    if nums:
        out.extend((label, n) for n in nums)
    else:
        out.append((label, None))

# Rewrite the result sheet in-place (cols A:B).
for row in range(1, dst.max_row + 1):
    dst.cell(row, 1).value = None
    dst.cell(row, 2).value = None
for i, (label, num) in enumerate(out, start=1):
    dst.cell(i, 1, label)
    dst.cell(i, 2, num)

wb.save(OUT)
