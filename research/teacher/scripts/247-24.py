import os

from openpyxl import load_workbook

wb = load_workbook(os.environ["OUT_XLSX"])
main = wb["Main"]
lookup = wb["Lookup"]
NCOL = 13

# Lookup sheet: Employee ID (col A) -> Working Weeks (col B).
weeks = {}
for r in range(2, lookup.max_row + 1):
    k = lookup.cell(r, 1).value
    if k not in (None, ""):
        weeks[str(k).strip()] = lookup.cell(r, 2).value

# Read the real data rows; only A-H are plain inputs (I-L are formulas we recompute).
rows, last0 = [], 1
for r in range(2, main.max_row + 1):
    vals = [main.cell(r, c).value for c in range(1, 9)]
    if all(v in (None, "") for v in vals):
        continue
    last0 = r
    rows.append(vals)


def comp(v):
    return str(v[0] or "").strip()


def loc(v):
    return str(v[1] or "").strip()


# 1) drop Company 'Motorcycle'; 2) drop 'Ahmed Sons' located in 'Canada'.
kept = [v for v in rows if comp(v) != "Motorcycle"
        and not (comp(v) == "Ahmed Sons" and loc(v) == "Canada")]

out = []  # 13-col rows, or None for an inserted blank row
for v in kept:
    # 3) National TV in India gets Bill Rate (col G) = 180.
    g = 180 if (comp(v) == "National TV" and loc(v) == "India") else v[6]
    h = v[7]
    i = g * h              # Resource Bill To client = Bill Rate * Worked Hours
    j = g - 30             # Intercompany Billing Rate = Bill Rate - 30
    k = j * h              # Actual Bill To Client = Intercompany Rate * Worked Hours
    profit = i - k         # Profit
    m = weeks.get(str(v[2]).strip())  # 5) VLOOKUP Working Weeks by Employee ID (col C)
    out.append(v[:6] + [g, h, i, j, k, profit, m])
    if comp(v) == "Ahmed Sons":       # 4) two blank rows after each Ahmed Sons row
        out += [None, None]

# Clear the old block, then write the rebuilt rows from row 2 (blank rows stay empty,
# so no leftover output lands in the inserted/deleted rows).
for r in range(2, max(last0, len(out) + 1) + 1):
    for c in range(1, NCOL + 1):
        main.cell(row=r, column=c).value = None
for idx, row in enumerate(out):
    if row is None:
        continue
    for c, val in enumerate(row, start=1):
        main.cell(row=2 + idx, column=c, value=val)

wb.save(os.environ["OUT_XLSX"])
