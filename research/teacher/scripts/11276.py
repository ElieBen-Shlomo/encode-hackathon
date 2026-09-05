import datetime
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb.active

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# Row 4 holds the dates; row 3 should show the matching short weekday name.
# Fix columns F through AJ (6..36).
for col in range(6, 37):
    date = ws.cell(row=4, column=col).value
    if isinstance(date, datetime.datetime):
        ws.cell(row=3, column=col, value=DAYS[date.weekday()])

wb.save(OUT)
