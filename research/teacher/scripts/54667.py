import datetime
import os

import openpyxl

IN = os.environ["IN_XLSX"]
OUT = os.environ["OUT_XLSX"]

wb_vals = openpyxl.load_workbook(IN, data_only=True)
data_ws = wb_vals["DATA"]
test_vals = wb_vals["Test"]

# Build a lookup of DATA!A -> DATA!E (VLOOKUP(x, DATA!$A:$E, 5, 0)).
lookup = {}
for row in range(1, data_ws.max_row + 1):
    key = data_ws.cell(row, 1).value
    if key is not None and key not in lookup:
        lookup[key] = data_ws.cell(row, 5).value

wb = openpyxl.load_workbook(OUT)
ws = wb["Test"]

for row in range(3, 59):
    # D and E are formula-driven (VLOOKUP by name), so use the cached values.
    d = test_vals.cell(row, 4).value
    e = test_vals.cell(row, 5).value
    if d == e:
        m = test_vals.cell(row, 13).value
        result = m if m is not None else datetime.timedelta(0)
    else:
        c = test_vals.cell(row, 3).value
        result = lookup.get(c, "")
    ws.cell(row, 7).value = result

wb.save(OUT)
