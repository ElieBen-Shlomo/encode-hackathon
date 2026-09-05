import os
import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

for row in ws.iter_rows(min_row=2, max_col=3):
    result = None
    for cell in row:
        if cell.data_type != "e":
            result = cell.value
            break
    ws.cell(row[0].row, 4).value = result

wb.save(OUT)
