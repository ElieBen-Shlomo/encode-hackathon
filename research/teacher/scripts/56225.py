import os
from datetime import datetime

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws_vals = openpyxl.load_workbook(OUT, data_only=True)["data"]
ws = wb["data"]


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        for fmt in ("%m-%d-%y", "%m/%d/%y", "%m-%d-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


name_filter = str(ws_vals["M3"].value).strip().lower()
target_date = as_date(ws_vals["N5"].value)

total = 0.0
r = 2
while ws_vals.cell(row=r, column=1).value is not None:
    name = ws_vals.cell(row=r, column=3).value
    k_val = ws_vals.cell(row=r, column=11).value
    row_date = as_date(ws_vals.cell(row=r, column=1).value)
    if (
        isinstance(name, str)
        and name_filter in name.lower()
        and row_date == target_date
        and isinstance(k_val, (int, float))
    ):
        total += k_val
    r += 1

ws["O3"] = round(total, 6)

wb.save(OUT)
