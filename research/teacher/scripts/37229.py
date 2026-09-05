import datetime
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Month rows B3:B14 -> output row, keyed by (year, month).
month_row = {}
for r in range(3, 15):
    d = ws.cell(r, 2).value
    if isinstance(d, datetime.datetime):
        month_row[(d.year, d.month)] = r

# Accumulate each dated data row (B17:B24, columns C:G) into its month's totals.
totals = {}
for r in range(17, 25):
    d = ws.cell(r, 2).value
    if not isinstance(d, datetime.datetime):
        continue
    out = month_row.get((d.year, d.month))
    if out is None:
        continue
    for c in range(3, 8):  # columns C-G
        v = ws.cell(r, c).value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            totals[(out, c)] = totals.get((out, c), 0) + v

# Write totals to C3:G14; leave the cell empty when the sum is zero.
for r in range(3, 15):
    for c in range(3, 8):
        s = totals.get((r, c), 0)
        ws.cell(r, c).value = s if s else None

wb.save(os.environ["OUT_XLSX"])
