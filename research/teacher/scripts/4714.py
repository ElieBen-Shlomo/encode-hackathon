import os
from collections import defaultdict

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

history = defaultdict(list)
for row in range(2, ws.max_row + 1):
    emp = ws.cell(row, 1).value
    if emp is None:
        break
    hours = ws.cell(row, 4).value or 0
    history[emp].append(hours)
    window = history[emp][-4:]
    if len(window) < 4:
        ws.cell(row, 5).value = "n/a"
    else:
        avg = sum(window) / 4
        ws.cell(row, 5).value = int(avg) if avg == int(avg) else avg

wb.save(OUT)
