import os

import openpyxl

out = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(out)
ws = wb.active
# The 'Points' columns hold IF() formulas that return text ("5", "2", ...); read
# their computed values from a data_only view of the same workbook.
vals = openpyxl.load_workbook(out, data_only=True).active

# 'Section Total' column, and the 'Points' columns to its left (row-3 headers).
total_col = next(c for c in range(1, ws.max_column + 1)
                 if ws.cell(2, c).value == "Section Total")
points_cols = [c for c in range(1, total_col) if ws.cell(3, c).value == "Points"]

# Entry rows carry identity data in columns A-D (rider/horse/age); the Points
# formulas are dragged far below the last entry, so key off that identity.
for r in range(4, ws.max_row + 1):
    if all(ws.cell(r, c).value in (None, "") for c in range(1, 5)):
        continue
    total = 0.0
    for c in points_cols:
        v = vals.cell(r, c).value
        if v not in (None, ""):
            total += float(v)
    ws.cell(r, total_col, value=total)

wb.save(out)
