import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

part = ws.cell(row=2, column=1).value  # A2: the raw part number to total

# From column O onward, data comes in repeating pairs of columns: a raw-part
# number column followed by its quantity column, one pair per completed part.
total = 0
for col in range(15, ws.max_column + 1, 2):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=col).value == part:
            qty = ws.cell(row=row, column=col + 1).value
            if isinstance(qty, (int, float)):
                total += qty

ws.cell(row=2, column=13).value = total  # M2

wb.save(OUT)
