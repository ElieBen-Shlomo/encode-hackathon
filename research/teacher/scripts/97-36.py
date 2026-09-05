import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

SEARCH = "domestic"
last_row = ws.max_row

start = next((r for r in range(1, last_row + 1)
              if isinstance(ws.cell(row=r, column=1).value, str) and SEARCH in ws.cell(row=r, column=1).value.lower()),
             1)

kept = [ws.cell(row=r, column=1).value for r in range(start, last_row + 1)]

for r in range(1, last_row + 1):
    ws.cell(row=r, column=1).value = None

for i, v in enumerate(kept):
    ws.cell(row=1 + i, column=1).value = v

wb.save(OUT)
