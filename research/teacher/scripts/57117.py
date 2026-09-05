import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

for row in range(2, ws.max_row + 1):
    lane = ws.cell(row=row, column=4).value
    if not isinstance(lane, str):
        continue
    if "MA" in lane:
        ws.cell(row=row, column=6).value = "MA"
    elif "BA" in lane:
        ws.cell(row=row, column=6).value = "BA"

wb.save(OUT)
