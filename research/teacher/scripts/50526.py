"""Look up the row in Sheet1 whose column-A value matches the lookup value in B6, then list
every column header (row 1) whose value in that row is > 0, stacked down from B9."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

lookup = ws["B6"].value
max_col = ws.max_column

headers = {c: ws.cell(row=1, column=c).value for c in range(2, max_col + 1)}

target_row = None
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=1).value == lookup:
        target_row = r
        break

matches = []
if target_row is not None:
    for c in range(2, max_col + 1):
        val = ws.cell(row=target_row, column=c).value
        if isinstance(val, (int, float)) and val > 0:
            matches.append(headers[c])

# Clear the previous result list, then write the new one.
for i in range(max_col - 1):
    ws.cell(row=9 + i, column=2).value = None
for i, name in enumerate(matches):
    ws.cell(row=9 + i, column=2).value = name

wb.save(os.environ["OUT_XLSX"])
