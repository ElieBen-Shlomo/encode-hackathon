import os
import openpyxl
from openpyxl.styles import Font, PatternFill

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb["3 straights to 3 boxed"]

fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")

row = 2
while ws.cell(row=row, column=1).value not in (None, ""):
    digits = [ws.cell(row=row, column=c).value for c in (2, 3, 4)]
    if all(d is not None for d in digits):
        boxed = "".join(str(int(d)) for d in sorted(digits))
        cell = ws.cell(row=row, column=7, value=boxed)
        cell.font = Font(bold=True)
        cell.fill = fill
    row += 1

wb.save(out_path)
