import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Total ability (H) is a running total: each year's deposit is added to the
# prior total, then the whole thing grows by that year's efficiency rate.
# Years with no deposit/efficiency simply carry the total forward unchanged.
total = 0.0
for row in range(2, ws.max_row + 1):
    deposit = ws.cell(row=row, column=2).value or 0
    efficiency = ws.cell(row=row, column=3).value or 0
    total = (total + deposit) * (1 + efficiency)
    ws.cell(row=row, column=8).value = round(total, 6)

wb.save(os.environ["OUT_XLSX"])
