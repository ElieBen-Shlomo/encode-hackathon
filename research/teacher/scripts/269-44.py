import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Import Data"] if "Import Data" in wb.sheetnames else wb.active

# Delete the four rows immediately before each "Chassis" in column A.
# Applied top-down and repeated until stable: this removes every full block of
# four rows preceding each marker, packing the "Chassis" rows toward the top.
deleting = True
while deleting:
    deleting = False
    for r in range(5, ws.max_row + 1):  # need four rows above, so r >= 5
        if ws.cell(row=r, column=1).value == "Chassis":
            ws.delete_rows(r - 4, 4)
            deleting = True
            break

wb.save(OUT)
