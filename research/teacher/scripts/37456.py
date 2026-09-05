import os
from datetime import datetime, time

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active  # 'Sample Data'


def as_time(v):
    if isinstance(v, datetime):
        return v.time()
    return v if isinstance(v, time) else time(0, 0)


# Raw transactions: A=Date, B=Time, C=Product, D=Sales (row 4 while a date is present).
tx = []
r = 4
while ws.cell(row=r, column=1).value is not None:
    d = ws.cell(row=r, column=1).value
    dt = datetime.combine(d.date(), as_time(ws.cell(row=r, column=2).value))
    tx.append((ws.cell(row=r, column=3).value, dt, ws.cell(row=r, column=4).value or 0))
    r += 1

# Window boundary times: F1 = 06:00:00 (start), F2 = 05:59:59 (end).
start_t = as_time(ws["F1"].value)
end_t = as_time(ws["F2"].value)

# Product headers of the Optimal Table (row 11, from column G onward).
products = {}
c = 7
while ws.cell(row=11, column=c).value is not None:
    products[c] = ws.cell(row=11, column=c).value
    c += 1

# Date rows of the Optimal Table (from row 12 while a date is present).
date_rows = []
r = 12
while isinstance(ws.cell(row=r, column=6).value, datetime):
    date_rows.append(r)
    r += 1

# Each day that has a following day: sum each product's sales inside the window
# [current_date + 06:00:00, next_date + 05:59:59]. The last date does not calculate.
fmt = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
for i in range(len(date_rows) - 1):
    row = date_rows[i]
    lo = datetime.combine(ws.cell(row=row, column=6).value.date(), start_t)
    hi = datetime.combine(ws.cell(row=date_rows[i + 1], column=6).value.date(), end_t)
    for col, prod in products.items():
        total = sum(s for (p, dt, s) in tx if p == prod and lo <= dt <= hi)
        cell = ws.cell(row=row, column=col, value=total)
        cell.number_format = fmt  # accounting -> currency

wb.save(OUT)
