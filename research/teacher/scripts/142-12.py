import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Sheet1"]

for row in range(1, ws.max_row + 1):
    name = ws.cell(row=row, column=6).value  # Column F
    amount = ws.cell(row=row, column=10).value  # Column J
    if name == "Marble Slab Creamery" and isinstance(amount, (int, float)):
        if float(amount) == int(amount):  # whole number, no cents
            ws.cell(row=row, column=6).value = "Georgia State WH"

wb.save(os.environ["OUT_XLSX"])
