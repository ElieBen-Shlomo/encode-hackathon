import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# The criterion the user inputs sits next to "List of Migrated Customer" (B12).
criterion = ws.cell(row=12, column=2).value

# Walk the customer columns (names in row 3, migrated flags in row 1) and collect
# every customer whose flag equals the criterion.
migrated = []
col = 2
while ws.cell(row=3, column=col).value is not None:
    if ws.cell(row=1, column=col).value == criterion:
        migrated.append(ws.cell(row=3, column=col).value)
    col += 1

# List them downward starting at A13.
for i, name in enumerate(migrated):
    ws.cell(row=13 + i, column=1).value = name

wb.save(OUT)
