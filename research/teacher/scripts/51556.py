import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

rating = ws["S2"].value  # subject impr. rating used for every lookup

for row in range(2, 7):  # P2:P6 -> Q2:Q6
    target_size = ws.cell(row=row, column=16).value  # column P
    values = []
    for r in range(3, ws.max_row + 1):  # data rows: B3:L8
        if ws.cell(row=r, column=13).value != rating:  # column M: impr. rating
            continue
        for c in range(2, 13):  # columns B..L
            if ws.cell(row=2, column=c).value == target_size:
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)):
                    values.append(v)
    ws.cell(row=row, column=17).value = sum(values) / len(values) if values else ""

wb.save(OUT)
