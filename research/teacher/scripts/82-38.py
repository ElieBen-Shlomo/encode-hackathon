import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws_data = wb["DATA"]


def invoice_numbers(sheet):
    """Sequential invoice numbers found directly below each 'INVOICE NO' label."""
    numbers = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "INVOICE NO":
                numbers.append(sheet.cell(row=cell.row + 1, column=cell.column).value)
    return numbers


invoice_cache = {}
pointers = {}

# Group DATA rows: each group runs until (and includes) a "TOTAL" marker row.
groups = []
current = []
for row in range(2, ws_data.max_row + 1):
    current.append(row)
    a_val = ws_data.cell(row=row, column=1).value
    if isinstance(a_val, str) and a_val.strip().upper() == "TOTAL":
        groups.append(current)
        current = []
if current:
    groups.append(current)

for group in groups:
    source_name = next(
        (ws_data.cell(row=r, column=5).value for r in group if ws_data.cell(row=r, column=5).value),
        None,
    )
    if source_name not in wb.sheetnames or source_name == ws_data.title:
        continue  # unmatched sheet name: leave original values untouched

    if source_name not in invoice_cache:
        invoice_cache[source_name] = invoice_numbers(wb[source_name])
        pointers[source_name] = 0

    numbers = invoice_cache[source_name]
    idx = pointers[source_name]
    if idx >= len(numbers):
        continue  # ran out of invoice numbers on the source sheet

    invoice = numbers[idx]
    pointers[source_name] = idx + 1
    for r in group:
        ws_data.cell(row=r, column=4).value = invoice

wb.save(out_path)
