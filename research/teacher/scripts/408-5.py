import copy
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active
wsv = openpyxl.load_workbook(OUT, data_only=True)[ws.title]

max_row, max_col = ws.max_row, ws.max_column
REF_COL, AMOUNT_COL, MATCH_COL = 3, 4, 5  # Column C, Column D, Column E per the instruction

# Sum Column D across every row where Column E is exactly zero (this also includes the
# 'BR1 Sales' row's own current amount, since its Column E is zero too).
zero_sum = 0.0
sales_row = None
for r in range(2, max_row + 1):
    if wsv.cell(row=r, column=MATCH_COL).value == 0:
        zero_sum += wsv.cell(row=r, column=AMOUNT_COL).value or 0
    ref = ws.cell(row=r, column=REF_COL).value
    if isinstance(ref, str) and "BR1 Sales" in ref:
        sales_row = r

# Keep the header, the 'BR1 Sales' row, and every other row that isn't blank and isn't a
# zero-match row.
keep_rows = [1] + [r for r in range(2, max_row + 1)
                    if not all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, max_col + 1))
                    and (r == sales_row or wsv.cell(row=r, column=MATCH_COL).value != 0)]


def snapshot(r):
    out = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        value = wsv.cell(row=r, column=c).value
        out.append([value, copy.copy(cell.font), copy.copy(cell.border), copy.copy(cell.fill),
                    cell.number_format, copy.copy(cell.alignment)])
    return out


rows_out = [snapshot(r) for r in keep_rows]
rows_out[keep_rows.index(sales_row)][AMOUNT_COL - 1][0] = zero_sum

for out_r, row_data in enumerate(rows_out, start=1):
    for c, (value, font, border, fill, fmt, align) in enumerate(row_data, start=1):
        cell = ws.cell(row=out_r, column=c)
        cell.value = value
        cell.font, cell.border, cell.fill = font, border, fill
        cell.number_format, cell.alignment = fmt, align

# Blank every row now beyond the kept range.
for r in range(len(rows_out) + 1, max_row + 1):
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).value = None

wb.save(OUT)
