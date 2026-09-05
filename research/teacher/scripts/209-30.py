import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Data to Import"]

# Remove the last three characters from each postcode in column C (data starts at C2).
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=3)
    value = cell.value
    if isinstance(value, str) and value:
        cell.value = value[:-3]

wb.save(os.environ["OUT_XLSX"])
