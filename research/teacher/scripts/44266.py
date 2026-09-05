import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
wb_vals = openpyxl.load_workbook(OUT, data_only=True)
data = wb["Data"]
lookup = wb["Lookup"]
lookup_vals = wb_vals["Lookup"]

# Row 4 marks the start of each year's block of columns; the "Total Expenses"
# value (row 14) sits somewhere within that block, before the next year starts.
year_starts = sorted(
    (col, data.cell(row=4, column=col).value)
    for col in range(1, data.max_column + 1)
    if data.cell(row=4, column=col).value is not None
)
year_col = {}
for i, (col, year) in enumerate(year_starts):
    end = year_starts[i + 1][0] if i + 1 < len(year_starts) else data.max_column + 1
    for c in range(col, end):
        if data.cell(row=14, column=c).value is not None:
            year_col[year] = c
            break

for row in range(2, lookup.max_row + 1):
    year = lookup_vals.cell(row=row, column=1).value
    col = year_col.get(year)
    value = data.cell(row=14, column=col).value if col is not None else ""
    lookup.cell(row=row, column=2).value = value if value is not None else ""

wb.save(OUT)
