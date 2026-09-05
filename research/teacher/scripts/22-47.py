import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["sheet1"]

# Source data lives in stacked blocks in columns A(ITEM) B(NAME) C(REF), each block
# led by an 'ITEM/NAME/REF' header. Collect unique items, skipping headers, empty
# cells, and duplicates (identical NAME and REF).
HEADERS = {"ITEM", "NAME", "REF"}
seen = set()
items = []
for r in range(1, ws.max_row + 1):
    name = ws.cell(r, 2).value
    ref = ws.cell(r, 3).value
    if name is None or ref is None:
        continue
    if str(name).strip() in HEADERS:
        continue
    key = (name, ref)
    if key in seen:
        continue
    seen.add(key)
    items.append((name, ref))

# Final output in F(ITEM) G(NAME) H(REF), sorted by REF (column H) lowest to highest.
items.sort(key=lambda x: x[1])
for i, (name, ref) in enumerate(items, start=1):
    ws.cell(i + 1, 6, i)
    ws.cell(i + 1, 7, name)
    ws.cell(i + 1, 8, ref)

wb.save(os.environ["OUT_XLSX"])
