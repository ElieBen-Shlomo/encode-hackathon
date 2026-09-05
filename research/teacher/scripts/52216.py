import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
inputs = wb["INPUTS"]
sheet2 = wb["Sheet2"]

# INPUTS!A14 holds the Metro/Regional selector; each expense row's label lives
# in column A. Build a lookup from Sheet2 keyed on (selector, expense label)
# -> its 5 year values (columns C:G), so C:G can be filled correctly no
# matter which row/column the original INDEX-MATCH formula drifted onto.
lookup = {}
for r in range(1, sheet2.max_row + 1):
    kind = sheet2.cell(row=r, column=1).value
    label = sheet2.cell(row=r, column=2).value
    if kind is None or label is None:
        continue
    values = [sheet2.cell(row=r, column=c).value for c in range(3, 8)]
    lookup[(kind, label)] = values

selector = inputs.cell(row=14, column=1).value

for row in (15, 16):
    label = inputs.cell(row=row, column=1).value
    values = lookup.get((selector, label))
    if values is None:
        continue
    for offset, col in enumerate(range(3, 8)):
        inputs.cell(row=row, column=col).value = values[offset]

wb.save(OUT)
