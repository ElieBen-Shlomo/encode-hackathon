import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

ref = ws["G2"].value

total = 0
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    if a == ref or b == ref:
        for col in (3, 4, 5):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                total += val

ws["H2"] = total
wb.save(out_path)
