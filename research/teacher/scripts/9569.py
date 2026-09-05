import datetime
import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb.active

wb_vals = openpyxl.load_workbook(OUT, data_only=True)
ws_vals = wb_vals.active

# Month columns run from the first date header onward; find where they end.
start_col = 5
end_col = start_col
while isinstance(ws.cell(row=8, column=end_col + 1).value, datetime.datetime):
    end_col += 1
target_col = end_col + 1  # column right after the last month = "Earliest GM"

for row in range(9, ws.max_row + 1):
    earliest = None
    for col in range(start_col, end_col + 1):
        v = ws_vals.cell(row=row, column=col).value
        if isinstance(v, (int, float)) and v != 0:
            earliest = v
            break
    if earliest is not None:
        ws.cell(row=row, column=target_col).value = earliest

wb.save(OUT)
