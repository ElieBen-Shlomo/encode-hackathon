import os
import re
from collections import defaultdict

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
out_ws = wb["OUT CAS"]
sources = [s for s in wb.sheetnames if s != out_ws.title]

# Each bucket: (account-ref prefix, starting column of its ITEM/REF/AMOUNT triple).
BUCKETS = [("PUR PURCHASE", 1), ("PUR PAID", 5), ("CASH SS", 9), ("EXPRR", 13)]

# Collect every ledger line from the other sheets, bucketed by account type.
buckets = defaultdict(list)
for name in sources:
    ws = wb[name]
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(r, 2).value
        if ref is None:
            continue
        ref = str(ref).strip()
        prefix = next((p for p, _ in BUCKETS if ref.startswith(p)), None)
        if prefix is None:
            continue
        debit, credit = ws.cell(r, 3).value, ws.cell(r, 4).value
        buckets[prefix].append((ref, debit if debit is not None else credit))


def item_number(ref):
    return int(re.findall(r"\d+", ref)[-1])


# Clear old data in each bucket's three columns before repopulating.
for _, start_col in BUCKETS:
    for r in range(2, out_ws.max_row + 1):
        for c in (start_col, start_col + 1, start_col + 2):
            out_ws.cell(r, c).value = None

for prefix, start_col in BUCKETS:
    items = sorted(buckets[prefix], key=lambda t: item_number(t[0]))
    row = 2
    for i, (ref, amount) in enumerate(items, start=1):
        out_ws.cell(row, start_col, i)
        out_ws.cell(row, start_col + 1, ref)
        out_ws.cell(row, start_col + 2, amount)
        row += 1
    out_ws.cell(row, start_col, "TOTAL")
    out_ws.cell(row, start_col + 2, sum(a for _, a in items))

wb.save(OUT)
