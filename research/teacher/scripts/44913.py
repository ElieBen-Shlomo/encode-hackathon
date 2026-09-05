import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

id_col, amt_col, out_col = 1, 2, 4

# Find the max AMOUNT per UNIQUE ID group.
max_by_id = {}
for r in range(2, ws.max_row + 1):
    uid, amt = ws.cell(r, id_col).value, ws.cell(r, amt_col).value
    if uid is None or not isinstance(amt, (int, float)):
        continue
    if uid not in max_by_id or amt > max_by_id[uid]:
        max_by_id[uid] = amt

# Show the amount only on the row(s) hitting that ID's max; leave every other row blank.
for r in range(2, ws.max_row + 1):
    uid, amt = ws.cell(r, id_col).value, ws.cell(r, amt_col).value
    ws.cell(r, out_col).value = amt if uid in max_by_id and amt == max_by_id[uid] else None

wb.save(OUT)
