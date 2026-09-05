"""Score rows 5-7 (columns B-E) from the top-3 source rows: per column, the row(s) tied for
the max get 6 (alone), 3 each (two-way tie), or 2 each (three-way tie); everyone else gets 0.
Rows are matched to their source by the label in column A, not by fixed offsets."""
import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

SOURCE_ROWS = [1, 2, 3]
TARGET_ROWS = [5, 6, 7]
POINTS = {1: 6, 2: 3, 3: 2}

labels = {r: ws.cell(row=r, column=1).value for r in SOURCE_ROWS}
target_for_label = {ws.cell(row=r, column=1).value: r for r in TARGET_ROWS}

for col in range(2, 6):  # B..E
    values = {r: ws.cell(row=r, column=col).value for r in SOURCE_ROWS}
    best = max(values.values())
    winners = [r for r, v in values.items() if v == best]
    points = POINTS[len(winners)]
    for r in SOURCE_ROWS:
        target_row = target_for_label[labels[r]]
        ws.cell(row=target_row, column=col).value = points if r in winners else 0

wb.save(os.environ["OUT_XLSX"])
