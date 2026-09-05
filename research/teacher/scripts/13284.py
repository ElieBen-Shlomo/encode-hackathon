import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
base = wb["Base"]
streets = wb["Streets"]

# Collect the Streets ranges that have a populated Start (column C). A blank End
# (column D) leaves the range open-ended (up to infinity).
ranges = []
for r in range(2, streets.max_row + 1):
    name = streets.cell(row=r, column=2).value
    start = streets.cell(row=r, column=3).value
    end = streets.cell(row=r, column=4).value
    result = streets.cell(row=r, column=6).value
    if name is None or not isinstance(start, (int, float)):
        continue
    hi = end if isinstance(end, (int, float)) else float("inf")
    ranges.append((str(name).strip().lower(), start, hi, result))

# For each Base row match the street name and number range; on overlaps keep the
# range with the highest Start (the most specific lower bound).
for r in range(2, base.max_row + 1):
    street = base.cell(row=r, column=3).value
    number = base.cell(row=r, column=4).value
    answer = ""
    if street is not None and isinstance(number, (int, float)):
        key = str(street).strip().lower()
        best = None
        for name, lo, hi, result in ranges:
            if name == key and lo <= number <= hi:
                if best is None or lo > best[0]:
                    best = (lo, result)
        if best is not None:
            answer = best[1]
    base.cell(row=r, column=5).value = answer

wb.save(OUT)
