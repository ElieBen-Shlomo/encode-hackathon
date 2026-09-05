import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Statement"]

# Find the first "Invoice No." in column A and delete every row above it so that
# it lands in row 1.
target = None
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if isinstance(v, str) and v.strip() == "Invoice No.":
        target = r
        break

if target and target > 1:
    ws.delete_rows(1, target - 1)

wb.save(os.environ["OUT_XLSX"])
