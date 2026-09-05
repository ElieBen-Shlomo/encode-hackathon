import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
name = next((s for s in wb.sheetnames if s.lower() == "final recon items"), wb.sheetnames[0])
ws = wb[name]

# Delete rows whose column D value is effectively zero (|value| < 0.011, absorbing the
# tiny +/-0.01 precision noise). The non-numeric text header in D1 is skipped.
THRESHOLD = 0.011
to_delete = [r for r in range(2, ws.max_row + 1)
             if isinstance(ws.cell(r, 4).value, (int, float))
             and not isinstance(ws.cell(r, 4).value, bool)
             and abs(ws.cell(r, 4).value) < THRESHOLD]

for r in sorted(to_delete, reverse=True):
    ws.delete_rows(r, 1)

wb.save(os.environ["OUT_XLSX"])
