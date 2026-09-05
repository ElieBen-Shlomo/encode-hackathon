"""Forecast total in column E: before 09:45 take column B (Actual Value), at/after 09:45
take column C (Forecast Value); 0/blank results stay blank. Sum into the Total row, and
fill E2:E<total row> with #FFC000."""
import datetime
import os

import openpyxl
from openpyxl.styles import PatternFill

CUTOFF = datetime.time(9, 45)


def as_time(v):
    if isinstance(v, datetime.datetime):
        return v.time()
    if isinstance(v, datetime.time):
        return v
    return None


wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
wb_vals = openpyxl.load_workbook(os.environ["OUT_XLSX"], data_only=True)
ws = wb.active
wsv = wb_vals[ws.title]

fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

total_row = None
first_row = None
running_total = 0
for r in range(2, ws.max_row + 1):
    a_val = wsv.cell(row=r, column=1).value
    t = as_time(a_val)
    if t is not None:
        first_row = first_row or r
        b_val = wsv.cell(row=r, column=2).value
        c_val = wsv.cell(row=r, column=3).value
        chosen = b_val if t < CUTOFF else c_val
        if chosen in (None, 0):
            ws.cell(row=r, column=5).value = None
        else:
            ws.cell(row=r, column=5).value = chosen
            running_total += chosen
    elif isinstance(a_val, str) and a_val.strip().lower() == "total":
        total_row = r

if total_row is not None:
    ws.cell(row=total_row, column=5).value = running_total

last_row = total_row or ws.max_row
for r in range(first_row or 2, last_row + 1):
    ws.cell(row=r, column=5).fill = fill

wb.save(os.environ["OUT_XLSX"])
