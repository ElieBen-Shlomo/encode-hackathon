import os
import re
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

max_col = ws.max_column
row = 2
while ws.cell(row=row, column=1).value not in (None, ""):
    text = str(ws.cell(row=row, column=1).value)
    tags = re.findall(r"#(\w+)", text)
    for c in range(2, max_col + 1):
        ws.cell(row=row, column=c).value = None
    for i, tag in enumerate(tags):
        ws.cell(row=row, column=2 + i, value=tag)
    row += 1

wb.save(out_path)
