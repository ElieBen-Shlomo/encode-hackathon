import os
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

pairs = {}
for row in range(2, ws.max_row + 1):
    date = ws.cell(row=row, column=5).value
    value = ws.cell(row=row, column=6).value
    if date is not None:
        pairs[date] = value

for row in range(2, ws.max_row + 1):
    timeline_date = ws.cell(row=row, column=1).value
    if timeline_date in pairs:
        ws.cell(row=row, column=2).value = timeline_date
        ws.cell(row=row, column=3).value = pairs[timeline_date]
    else:
        ws.cell(row=row, column=2).value = None
        ws.cell(row=row, column=3).value = None

wb.save(os.environ["OUT_XLSX"])
