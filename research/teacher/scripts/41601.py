import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Students"]

# For each student row, the name in column A names a sheet holding that
# student's age in C2 - look it up and drop the value into column E.
for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name in wb.sheetnames:
        ws.cell(row=row, column=5, value=wb[name]["C2"].value)

wb.save(OUT)
