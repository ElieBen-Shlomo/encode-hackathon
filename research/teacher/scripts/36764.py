import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
s1 = wb["Sheet1"]
s2 = wb["Sheet2"]

# SUMIF: total Sheet2!R2:R363 (column R = 18) over rows whose column A matches Sheet1!A2.
key = str(s1["A2"].value).strip().lower()
total = 0
for r in range(2, 364):
    a = s2.cell(r, 1).value
    if a is not None and str(a).strip().lower() == key:
        v = s2.cell(r, 18).value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            total += v

s1["C2"] = total
wb.save(os.environ["OUT_XLSX"])
