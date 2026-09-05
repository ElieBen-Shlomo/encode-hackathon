import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet2"]

BLANKY_STR = {"-", "$", "$0", "$0.0", "0"}


def is_blanky(v):
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() in BLANKY_STR
    if isinstance(v, (int, float)):
        return v == 0
    return False


first_row, last_row, ncols = 2, ws.max_row, 11

kept = []
for row in range(first_row, last_row + 1):
    j = ws.cell(row=row, column=10).value
    k = ws.cell(row=row, column=11).value
    if is_blanky(j) and is_blanky(k):
        continue
    kept.append([ws.cell(row=row, column=c).value for c in range(1, ncols + 1)])

for row in range(first_row, last_row + 1):
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).value = None

for i, vals in enumerate(kept):
    row = first_row + i
    for col, v in enumerate(vals, start=1):
        ws.cell(row=row, column=col).value = v

wb.save(OUT)
