import os

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

exclude = ws.cell(row=2, column=1).value

items = []
r = 2
while ws.cell(row=r, column=2).value is not None:
    items.append(ws.cell(row=r, column=2).value)
    r += 1
last_row = r - 1

filtered = [item for item in items if item != exclude]

for i in range(2, last_row + 1):
    idx = i - 2
    ws.cell(row=i, column=3).value = filtered[idx] if idx < len(filtered) else None

wb.save(OUT)
