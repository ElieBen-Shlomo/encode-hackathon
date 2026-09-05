import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet2"]

# Total price (column B) per serial number (column A).
totals = {}
for r in range(2, ws.max_row + 1):
    serial = ws.cell(row=r, column=1).value
    if serial is None:
        continue
    totals[serial] = totals.get(serial, 0) + (ws.cell(row=r, column=2).value or 0)

# Write each serial's total only on its first row; leave later rows blank.
seen = set()
for r in range(2, ws.max_row + 1):
    serial = ws.cell(row=r, column=1).value
    if serial is not None and serial not in seen:
        ws.cell(row=r, column=3, value=totals[serial])
        seen.add(serial)
    else:
        ws.cell(row=r, column=3, value=None)

wb.save(OUT)
