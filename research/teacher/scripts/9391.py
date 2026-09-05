import datetime
import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
front = wb["Front"]
data = wb["Data"]

DATA_VALUE_COL = 6  # column F, per the instruction

# Parse Data into date blocks: a row with a date in column A (and nothing in
# B) starts a block; the next row is the "Agent Name" header; subsequent
# rows (until a blank column A) are per-agent values.
lookup = {}
r = 1
while r <= data.max_row:
    val = data.cell(row=r, column=1).value
    if isinstance(val, datetime.datetime):
        block_date = val
        r += 2  # skip the "Agent Name" header row
        while data.cell(row=r, column=1).value not in (None, ""):
            agent = data.cell(row=r, column=1).value
            lookup[(block_date, agent)] = data.cell(row=r, column=DATA_VALUE_COL).value
            r += 1
    else:
        r += 1

# Fill Front!B2:.. for each date column and each agent row.
col = 2
while front.cell(row=1, column=col).value is not None:
    date = front.cell(row=1, column=col).value
    row = 2
    while front.cell(row=row, column=1).value is not None:
        agent = front.cell(row=row, column=1).value
        front.cell(row=row, column=col).value = lookup.get((date, agent))
        row += 1
    col += 1

wb.save(out_path)
