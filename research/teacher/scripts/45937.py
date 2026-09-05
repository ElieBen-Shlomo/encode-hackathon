import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

wb_vals = openpyxl.load_workbook(out_path, data_only=True)
ws_vals = wb_vals.active

# kms-range legend (instruction: "kms ranges are specified in cells H20:I22")
RANGE_FIRST_ROW = 20
RANGE_LAST_ROW = 22
# category-column header row for the legend (row above the % row)
CAT_HEADER_ROW = 18

# Build category letter -> column index map from the legend header row.
cat_to_col = {}
for col in range(ws.max_column, 0, -1):
    v = ws.cell(row=CAT_HEADER_ROW, column=col).value
    if isinstance(v, str) and len(v.strip()) == 1 and v.strip().isalpha():
        cat_to_col[v.strip()] = col
    if v == "Categories":
        break

# Walk the data rows starting at row 7 while a Category has been assigned in column D.
row = 7
while ws_vals.cell(row=row, column=4).value not in (None, ""):
    kms = ws_vals.cell(row=row, column=3).value
    category = ws_vals.cell(row=row, column=4).value
    pct = None
    if isinstance(kms, (int, float)) and category in cat_to_col:
        for r in range(RANGE_FIRST_ROW, RANGE_LAST_ROW + 1):
            lo = ws.cell(row=r, column=8).value  # H
            hi = ws.cell(row=r, column=9).value  # I
            if lo is None:
                continue
            if hi is None or kms <= hi:
                if kms >= lo:
                    pct = ws.cell(row=r, column=cat_to_col[category]).value
                    break
    ws.cell(row=row, column=5).value = pct
    row += 1

wb.save(out_path)
