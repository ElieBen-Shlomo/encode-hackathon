import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active


def tiers(income):
    # Tier 1: $0.15 per dollar over 60k up to 90k; Tier 2: $0.25 over 90k up to 120k;
    # Tier 3: $0.35 per dollar over 120k.
    t1 = 0.15 * max(0, min(income, 90000) - 60000)
    t2 = 0.25 * max(0, min(income, 120000) - 90000)
    t3 = 0.35 * max(0, income - 120000)
    return t1, t2, t3


# Quarter rows 6-9: Income in col B, Tier 1/2/3 in cols C/D/E.
for r in range(6, 10):
    income = ws.cell(r, 2).value or 0
    t1, t2, t3 = tiers(income)
    ws.cell(r, 3).value = t1
    ws.cell(r, 4).value = t2
    ws.cell(r, 5).value = t3

wb.save(OUT)
