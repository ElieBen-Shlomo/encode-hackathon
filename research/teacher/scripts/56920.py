import os
import re
import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

PATTERN = re.compile(r"955\d{7}|CHK\d{7}")

for row in range(3, ws.max_row + 1):
    text = ws.cell(row=row, column=2).value
    if not text:
        continue
    match = PATTERN.search(str(text))
    if match:
        ws.cell(row=row, column=3, value=match.group())

wb.save(os.environ["OUT_XLSX"])
