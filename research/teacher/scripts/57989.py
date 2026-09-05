import os
import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Locate the "Synthese" section marker in column A.
syn_row = None
for row in range(1, ws.max_row + 1):
    if ws.cell(row, 1).value == "Synthese":
        syn_row = row
        break

header_row = syn_row + 1  # weekday names for the synthesis table
data_start = syn_row + 2

# Main table: driver name (col A) -> {col index: weekday name}, rows 2..syn_row-1.
main_headers = {}
for col in range(2, ws.max_column + 1):
    val = ws.cell(1, col).value
    if val is not None:
        main_headers[col] = val

main_rows = {}
for row in range(2, syn_row):
    driver = ws.cell(row, 1).value
    if driver is not None:
        main_rows[driver] = row

# Synthesis table columns: weekday name -> column index.
syn_cols = {}
for col in range(2, ws.max_column + 1):
    val = ws.cell(header_row, col).value
    if val is not None:
        syn_cols[val] = col

for row in range(data_start, ws.max_row + 1):
    driver = ws.cell(row, 1).value
    if driver is None:
        break
    src_row = main_rows.get(driver)
    for weekday, syn_col in syn_cols.items():
        count = 0
        if src_row is not None:
            for col, name in main_headers.items():
                if name == weekday and ws.cell(src_row, col).value not in (None, ""):
                    count += 1
        ws.cell(row, syn_col).value = count

wb.save(OUT)
