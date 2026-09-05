import os
from collections import Counter

import openpyxl
from openpyxl.styles import Border, PatternFill, Side

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Count each (Mat Doc, Doc No) combination in columns E/F (data starts row 5).
counts = Counter()
for row in range(5, ws.max_row + 1):
    mat, doc = ws.cell(row=row, column=5).value, ws.cell(row=row, column=6).value
    if mat is not None and doc is not None:
        counts[(mat, doc)] += 1

# Remark column G: "Multiple" when the combination repeats, else "Single".
for row in range(5, ws.max_row + 1):
    mat, doc = ws.cell(row=row, column=5).value, ws.cell(row=row, column=6).value
    if mat is not None and doc is not None:
        ws.cell(row=row, column=7, value="Multiple" if counts[(mat, doc)] > 1 else "Single")

# Strip all fill colours and borders, then re-apply thin all-borders to populated
# cells except those reading "Multiple" or "Single".
thin = Side(style="thin")
all_borders = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows():
    for cell in row:
        cell.fill = PatternFill(fill_type=None)
        if cell.value is not None and str(cell.value).strip() not in ("Multiple", "Single"):
            cell.border = all_borders
        else:
            cell.border = Border()

wb.save(OUT)
