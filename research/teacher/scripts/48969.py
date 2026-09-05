import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active


def find_label(text):
    """Return (row, col) of the first cell whose value equals text."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == text:
                return cell.row, cell.column
    return None


def read_block(name_row, name_col):
    """Read a Name/date-columns table starting at the 'Name' header cell."""
    dates = {}
    col = name_col + 1
    while ws.cell(row=name_row, column=col).value is not None:
        dates[ws.cell(row=name_row, column=col).value] = col
        col += 1
    values = {}
    r = name_row + 1
    while ws.cell(row=r, column=name_col).value not in (None, ""):
        row_name = ws.cell(row=r, column=name_col).value
        values[row_name] = {
            d: ws.cell(row=r, column=c).value for d, c in dates.items()
        }
        r += 1
    return dates, values


s1_row, s1_col = find_label("Source 1")
s2_row, s2_col = find_label("Source 2")

# Locate the "Name" header of the Result block: it's on the same row as
# Source 1's header, to the right of the Source-1 date columns.
name_row = s1_row
name_col = s1_col + 1
result_name_col = None
for row in ws.iter_rows(min_row=name_row, max_row=name_row):
    for cell in row:
        if cell.value == "Name" and cell.column > name_col:
            result_name_col = cell.column
            break

_, source1 = read_block(s1_row, name_col)
_, source2 = read_block(s2_row, s2_col + 1)
result_dates, _ = read_block(name_row, result_name_col)

r = name_row + 1
while ws.cell(row=r, column=result_name_col).value not in (None, ""):
    name = ws.cell(row=r, column=result_name_col).value
    s1_vals = source1.get(name, {})
    s2_vals = source2.get(name, {})
    for date, col in result_dates.items():
        match = date in s1_vals and date in s2_vals and s1_vals[date] == s2_vals[date]
        ws.cell(row=r, column=col).value = bool(match)
    r += 1

wb.save(out_path)
