import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Sheet1"]
targets = {"@9T", "SAL", "T9A"}
NCOL = ws.max_column

# Read every row once, then keep the header plus data rows whose Group (col A) is
# not one of the target groups (dropping trailing empties too).
rows = list(ws.iter_rows(min_row=1, max_col=NCOL, values_only=True))
last0 = max((i for i, r in enumerate(rows, start=1) if any(v not in (None, "") for v in r)), default=1)
kept = [r for r in rows[1:]
        if any(v not in (None, "") for v in r)
        and not (isinstance(r[0], str) and r[0].strip() in targets)]

# Rewrite the survivors directly under the header, then blank the leftover tail.
for i, r in enumerate(kept, start=2):
    for c, v in enumerate(r, start=1):
        ws.cell(row=i, column=c, value=v)
for r in range(2 + len(kept), last0 + 1):
    for c in range(1, NCOL + 1):
        ws.cell(row=r, column=c).value = None

wb.save(os.environ["OUT_XLSX"])
