import os

import openpyxl
from openpyxl.styles import Border, Font, Side

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Column B lists "location: qty" pairs separated by commas. Count the locations that do NOT
# start with 'X' or 'Z' (those prefixes mark pallet locations to exclude).
for row in range(2, ws.max_row + 1):
    raw = ws.cell(row, 2).value
    if not isinstance(raw, str) or not raw.strip():
        continue
    count = 0
    for segment in raw.split(","):
        location = segment.split(":", 1)[0].strip()
        if location and not location.upper().startswith(("X", "Z")):
            count += 1
    ws.cell(row, 3).value = count

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in range(2, 7):
    cell = ws.cell(row, 3)
    cell.border = border
    cell.font = Font(name="Courier New", size=9)

wb.save(OUT)
