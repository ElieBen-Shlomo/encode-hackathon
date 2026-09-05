import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["NOMINA"]

# Columns: D=WORKER, G=DEDUCTIONS, H=ADDITIONS, J=DEBT.
# Debt carries the previous week's debt for the same worker, then adds the
# new addition and subtracts the new deduction.
debt = {}
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=1).value == "Total":
        break
    worker = ws.cell(row=r, column=4).value
    if worker is None:
        continue
    deduction = ws.cell(row=r, column=7).value or 0
    addition = ws.cell(row=r, column=8).value or 0
    debt[worker] = debt.get(worker, 0) + addition - deduction
    ws.cell(row=r, column=10, value=debt[worker])

wb.save(OUT)
