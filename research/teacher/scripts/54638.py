import os
import openpyxl
from openpyxl.styles import PatternFill

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb["Sheet1"]

wb_vals = openpyxl.load_workbook(in_path, data_only=True)
ws_vals = wb_vals["Sheet1"]

# Unique, order-preserving list of non-blank values from A2:A150.
seen = []
for row in range(2, 151):
    val = ws_vals.cell(row=row, column=1).value
    if val is None or str(val).strip() == "":
        continue
    val = str(val).strip()
    if val not in seen:
        seen.append(val)

fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
for i, row in enumerate(range(2, 151)):
    value = seen[i] if i < len(seen) else None
    cell = ws.cell(row=row, column=2, value=value)
    cell.fill = fill

ws.sheet_view.showGridLines = False
wb["Sheet2"].sheet_state = "hidden"

wb.save(out_path)
