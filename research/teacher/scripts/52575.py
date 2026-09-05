"""Mark 'y' in column F on the first row of each new (employee, week) run — i.e. where the
week number (col E) differs from the row above — if that employee ever has both a 'Sick
Day' (col C) and an 'Overtime' (col D) entry for that week number anywhere in the sheet
(a week number can recur in separate non-contiguous date ranges)."""
import os
from collections import defaultdict

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

last_row = ws.max_row

flags = defaultdict(lambda: {"sick": False, "overtime": False})
for r in range(2, last_row + 1):
    emp = ws.cell(row=r, column=2).value
    week = ws.cell(row=r, column=5).value
    key = (emp, week)
    if ws.cell(row=r, column=3).value == "Sick Day":
        flags[key]["sick"] = True
    if ws.cell(row=r, column=4).value == "Overtime":
        flags[key]["overtime"] = True

prev_week = None
for r in range(2, last_row + 1):
    emp = ws.cell(row=r, column=2).value
    week = ws.cell(row=r, column=5).value
    is_boundary = week != prev_week
    prev_week = week
    if not is_boundary:
        continue
    f = flags[(emp, week)]
    ws.cell(row=r, column=6).value = "y" if f["sick"] and f["overtime"] else None

wb.save(os.environ["OUT_XLSX"])
