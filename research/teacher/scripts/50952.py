import os
import re

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Build id -> token set from Comments column (split on any non-alphanumeric char).
records = []
for row in range(2, ws.max_row + 1):
    rid = ws.cell(row, 1).value
    comment = ws.cell(row, 2).value
    if rid is None or comment is None:
        continue
    tokens = {t.lower() for t in re.split(r"[^A-Za-z0-9]+", str(comment)) if t}
    records.append((rid, tokens))

for row in range(2, ws.max_row + 1):
    p1 = ws.cell(row, 5).value
    p2 = ws.cell(row, 6).value
    if p1 is None or p2 is None:
        continue
    needed = {str(p1).lower(), str(p2).lower()}
    found = None
    for rid, tokens in records:
        if needed <= tokens:
            found = rid
            break
    ws.cell(row, 7).value = found

wb.save(OUT)
