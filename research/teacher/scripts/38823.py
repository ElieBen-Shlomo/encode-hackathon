import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# For each search term in column H, sum Units Sold (C) over rows whose Date (A) is
# within the [E4, F4] range and whose Fabric list (B) contains that search term.
start = ws["E4"].value
end = ws["F4"].value

for r in range(4, ws.max_row + 1):
    term = ws.cell(row=r, column=8).value  # H
    if term is None or str(term).strip() == "":
        continue
    total = 0
    for dr in range(3, ws.max_row + 1):
        date = ws.cell(row=dr, column=1).value    # A
        fabric = ws.cell(row=dr, column=2).value  # B
        units = ws.cell(row=dr, column=3).value   # C
        if date is None or fabric is None or units is None:
            continue
        if start <= date <= end and str(term) in str(fabric):
            total += units
    ws.cell(row=r, column=9, value=total)  # I

wb.save(OUT)
