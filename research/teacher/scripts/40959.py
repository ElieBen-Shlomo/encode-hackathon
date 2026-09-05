import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

# Column C is "Marketing" whenever Column A is hyphenated (e.g. "Sales-Marketing") or
# Column B contains "General"; otherwise it just echoes the single-department name.
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=1).value
    if name in (None, ""):
        continue
    sub_dept = ws.cell(row=r, column=2).value
    hyphenated = "-" in str(name)
    flagged_general = isinstance(sub_dept, str) and "General" in sub_dept
    ws.cell(row=r, column=3, value="Marketing" if hyphenated or flagged_general else str(name))

wb.save(os.environ["OUT_XLSX"])
