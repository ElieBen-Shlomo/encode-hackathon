"""For each race (a block of rows starting where col B 'Runner' == 1), transpose the Odds
(col G) values of its first C (col C 'Runners') rows across columns H, I, J... on the
race's start row, preserving blanks as blanks. Then shade H2:S28, drop decimals on whole
numbers in that range, and center-align it."""
import os

import openpyxl
from openpyxl.styles import Alignment, PatternFill

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=2).value != 1:
        continue
    runners = ws.cell(row=r, column=3).value
    if not isinstance(runners, (int, float)):
        continue
    for i in range(int(runners)):
        odds = ws.cell(row=r + i, column=7).value
        ws.cell(row=r, column=8 + i).value = odds

fill = PatternFill(fill_type="solid", start_color="E2EFDA", end_color="E2EFDA")
center = Alignment(horizontal="center", vertical="center")
for row in ws["H2:S28"]:
    for cell in row:
        cell.fill = fill
        cell.alignment = center
        if isinstance(cell.value, (int, float)) and float(cell.value) == int(cell.value):
            cell.number_format = "0"

wb.save(os.environ["OUT_XLSX"])
