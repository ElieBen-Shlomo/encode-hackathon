import os

import openpyxl
from openpyxl.utils import column_index_from_string

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
src = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
dst = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.create_sheet("Sheet2")
dst.delete_rows(1, dst.max_row)  # start clean in case of a rerun

n_cols = src.max_column
count_col = column_index_from_string("AG")  # the duplicate count, per the instruction
headers = [src.cell(1, c).value for c in range(1, n_cols + 1)]
for c, h in enumerate(headers, start=1):
    dst.cell(1, c).value = h

# Walk Sheet1's data rows (until column A runs out) and duplicate each one per its count,
# inserted directly below the previous copy, preserving the original row order.
out_row, r = 2, 2
while src.cell(r, 1).value not in (None, ""):
    values = [src.cell(r, c).value for c in range(1, n_cols + 1)]
    try:
        copies = max(int(values[count_col - 1]), 1)
    except (TypeError, ValueError):
        copies = 1
    for _ in range(copies):
        for c, v in enumerate(values, start=1):
            dst.cell(out_row, c).value = v
        out_row += 1
    r += 1

wb.save(OUT)
