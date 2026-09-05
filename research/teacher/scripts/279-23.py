import os
import re

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb["Sheet1"]

# For every cell in column B: keep only what is inside parentheses (dropping the
# parentheses and everything outside them). Blank cells stay blank; cells with no
# parentheses are left unchanged.
for r in range(1, ws.max_row + 1):
    cell = ws.cell(row=r, column=2)
    if cell.value is None:
        continue
    m = re.search(r"\(([^)]*)\)", str(cell.value))
    if m:
        cell.value = m.group(1)

wb.save(OUT)
