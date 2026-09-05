import os
import openpyxl

out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

max_row, max_col = ws.max_row, ws.max_column

# Header row: column A blank, several columns to the right hold company names.
header_row = None
for r in range(1, max_row + 1):
    a_val = ws.cell(row=r, column=1).value
    texts = [ws.cell(row=r, column=c).value for c in range(2, max_col + 1)]
    if a_val in (None, "") and sum(1 for t in texts if isinstance(t, str) and t.strip()) >= 2:
        header_row = r
        break

headers = {}
for c in range(2, max_col + 1):
    v = ws.cell(row=header_row, column=c).value
    if isinstance(v, str) and v.strip():
        headers[v.strip()] = c

# Data rows directly below the header, until the first fully blank row.
data_rows = []
r = header_row + 1
while r <= max_row:
    vals = [ws.cell(row=r, column=c).value for c in headers.values()]
    if all(v is None for v in vals):
        break
    data_rows.append(r)
    r += 1

# Summary rows below: company name in column A, total goes in column B.
for r in range(data_rows[-1] + 1, max_row + 1):
    name = ws.cell(row=r, column=1).value
    if not isinstance(name, str) or not name.strip():
        continue
    col = headers.get(name.strip())
    if col is None:
        continue
    total = sum(
        ws.cell(row=dr, column=col).value
        for dr in data_rows
        if isinstance(ws.cell(row=dr, column=col).value, (int, float))
    )
    ws.cell(row=r, column=2, value=total)

wb.save(out_path)
