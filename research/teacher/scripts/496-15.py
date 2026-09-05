import os
import re

import openpyxl

OUT = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(OUT, data_only=False)
ws = wb["Sheet1"]

for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=5).value  # column E
    if not val or "check" not in str(val).lower():
        continue
    digits = re.findall(r"\d+", str(val))
    if not digits:
        continue
    ws.cell(row=row, column=4).value = int(digits[-1])
    ws.cell(row=row, column=5).value = "Check Paid"

wb.save(OUT)
