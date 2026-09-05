import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb.active

values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]

# Every group of 3 source rows becomes one record spread across columns D, E, F.
for i in range(0, len(values) - len(values) % 3, 3):
    out_row = i // 3 + 1
    for j, v in enumerate(values[i:i + 3]):
        ws.cell(row=out_row, column=4 + j, value=v)

wb.save(os.environ["OUT_XLSX"])
