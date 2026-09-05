import os

import openpyxl

wb = openpyxl.load_workbook(os.environ["OUT_XLSX"])
ws = wb["Sheet1"]
NCOL = ws.max_column

# Locate the sort columns dynamically by header text (row 1) instead of hard-coding.
header = {ws.cell(1, c).value: c for c in range(1, NCOL + 1)}
task = header["Task"]
resp = header["Responsibility"]

# Collect the data rows and sort by Task, then Responsibility, both ascending.
data = []
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, NCOL + 1)]
    if any(v not in (None, "") for v in vals):
        data.append(vals)

data.sort(key=lambda v: (str(v[task - 1]), str(v[resp - 1])))

for i, vals in enumerate(data):
    for c, v in enumerate(vals, start=1):
        ws.cell(row=2 + i, column=c, value=v)

wb.save(os.environ["OUT_XLSX"])
