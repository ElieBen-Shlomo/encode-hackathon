import os

import openpyxl

OUT = os.environ["OUT_XLSX"]
wb = openpyxl.load_workbook(OUT)
ws = wb.active

# Keyword phrases with their search volumes, from column A/B.
keywords = []
r = 2
while ws.cell(r, 1).value is not None:
    keywords.append((str(ws.cell(r, 1).value).strip().lower(), ws.cell(r, 2).value))
    r += 1

# For each root word in column D, sum the volume of every keyword that contains
# the root as a complete word (not merely a substring), write to column E.
r = 2
while ws.cell(r, 4).value is not None:
    root = str(ws.cell(r, 4).value).strip().lower()
    total = sum(vol for kw, vol in keywords if root in kw.split())
    ws.cell(r, 5, total)
    r += 1

wb.save(OUT)
