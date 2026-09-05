import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

wb_vals = openpyxl.load_workbook(out_path, data_only=True)
ws_vals = wb_vals.active

# Row 1 has merged header blocks (e.g. "TOTAL", dates); row 2 has the
# per-column category (SPEND / SALES / REVENUE / GP) within each block.
blocks = []
total_cols = None
for merged in ws.merged_cells.ranges:
    if merged.min_row == 1:
        cols = list(range(merged.min_col, merged.max_col + 1))
        label = ws.cell(row=1, column=merged.min_col).value
        if isinstance(label, str) and label.strip().upper() == "TOTAL":
            total_cols = cols
        else:
            blocks.append(cols)

data_row = 3

for col in total_cols:
    category = ws.cell(row=2, column=col).value
    total = 0
    for block in blocks:
        for c in block:
            if ws.cell(row=2, column=c).value == category:
                v = ws_vals.cell(row=data_row, column=c).value
                if isinstance(v, (int, float)):
                    total += v
    ws.cell(row=data_row, column=col).value = total

wb.save(out_path)
