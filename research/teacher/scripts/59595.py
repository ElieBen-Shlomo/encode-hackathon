import datetime
import os
import openpyxl

in_path = os.environ["IN_XLSX"]
out_path = os.environ["OUT_XLSX"]

wb = openpyxl.load_workbook(out_path)
ws = wb.active

# Data starts at row 4: column A = date, column B = points, column C = rolling 7-day total.
data = []
row = 4
while ws.cell(row=row, column=1).value not in (None, ""):
    data.append((row, ws.cell(row=row, column=1).value, ws.cell(row=row, column=2).value))
    row += 1

for row, date, _points in data:
    total = 0
    for _r, d, p in data:
        if d <= date and (date - d) <= datetime.timedelta(days=7):
            total += p
    ws.cell(row=row, column=3).value = total

wb.save(out_path)
